#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
dbt_to_kimball_excel.py  —  Generate a Kimball-style Excel catalog + star & lineage diagrams
===========================================================================================

Version: 2.0.1

What this does
--------------
Turns a compiled **dbt Core** project into a Kimball-style Excel workbook with:

1) Summary tab
   • One row per model: Kind, Materialization, Relation, FQN, Tags, Description.
   • Frozen header, blue header bg, bold header, auto-filter, wrapped long text.

2) (Optional) Relationships tab  [default: included — disable with --no-relationships]
   • One row per dbt `relationships` test (FromModel/FromColumn → ToModel/ToColumn).
   • Mirrors exactly what dbt tests assert.

3) Star Map tab
   • Aggregated rows of (FactModel, DimensionModel, FactFKColumn(s), DimKeyColumn(s)).
   • This is the canonical mapping that the star diagrams use (after options).

4) One tab per model (Dimension/Fact/Other)
   • **Table meta block** at the top (labels **without '#'** and **bold**):
       Model, Kind, Materialization, Relation, Tags, SurrogateKey, BusinessKey, PrimaryKey, Description
     - Surrogate/Business/Primary use **original casing** from your YAML.
     - If PrimaryKey is missing in YAML, we **fall back to SurrogateKey** for display.
   • **Column grid**: Column, DataType, Nullable?, Description, Tests, IsPK?, IsFK?, Source
     - Heuristic PK/FK from tests; Nullable? from not_null; YAML overlay for data_type/nullable/description.
   • Frozen column header, blue header + bold, auto-filter, wrapped Description/Tests, “green-bar” banding.

5) (Optional) **Star-diagram** tabs (PNG via Matplotlib, embedded in new sheets)
   • Per fact model: fact = rounded yellow rectangle, dims = rounded **light green** rectangles.
   • **Bold** labels for both fact and dimensions.
   • FK column names printed along spokes (wrapped if long).
   • Two-ring layout if many dimensions (configurable).
   • Sheet titles like **“Star-Fact FactOrderLine”** (safe & ≤31 chars).

6) (New) **Lineage** and **Pipelines** tabs (enable with `--lineage`)
   • **Lineage**: one row per dependency edge (upstream → downstream), with tiers, materializations, and paths.
   • **Pipelines**: one row per **mart** model with lists of upstream **stage** and **base** models (closure).

7) (New) Optional **Lineage-diagram** tabs (enable with `--lineage-diagrams` together with `--lineage`)
   • Left→Right layered DAG per mart: Base (left, pale blue) → Stage (middle, lavender) → Mart (right, yellow).
   • Rounded rectangles, arrows, bold labels. Embedded as PNG per mart.

Sheet tab style
---------------
We can’t change Excel’s rounded “pill” shape, but we color tabs:
  - Summary / Relationships / Star Map / Lineage / Pipelines: blue
  - Dimension model tabs: purple
  - Fact model tabs: peach
  - Star/Lineage diagram tabs: green

Inputs
------
• --manifest  target/manifest.json   (required)
• --catalog   target/catalog.json    (required)
• --schemas   CSV/globs of schema YAMLs (optional) to overlay:
    - Reads model-level `meta.surrogate_key`, `meta.business_key`, `meta.primary_key`
      (keeps original case for display; also stores lowercase for matching).
    - Reads column-level `meta.data_type`, `meta.nullable`, and `description`.

Install
-------
    pip install pandas openpyxl pyyaml matplotlib

Typical use
-----------
    python scripts/dbt_to_kimball_excel.py ^
      --manifest .\target\manifest.json ^
      --catalog  .\target\catalog.json ^
      --schemas ".\models\mart\schemas.yml,.\models\stage\schemas.yml" ^
      --exclude-sheet-prefixes base_,stage_ ^
      --exclude-sheet-tags stage,logging ^
      --exclude-sheet-path-globs "models/base/*,models/stage/*,models/logging/*" ^
      --star-diagrams --diagram-two-rings 12 --diagram-font-scale 0.95 --diagram-dpi 220 ^
      --lineage --lineage-diagrams ^
      --out .\documentation\kimball_model_catalog.xlsx

Important flags & behavior
--------------------------
• --include-views + --materializations TABLE,INCREMENTAL: combined filter logic for per-model tabs.
• Exclude per-model tabs via:
  --exclude-sheet-prefixes / --exclude-sheet-tags / --exclude-sheet-materializations / --exclude-sheet-path-globs
• --no-relationships            : omit the Relationships sheet (Star Map & diagrams still generated).
• --include-missing-dims        : keep Star Map rows whose target dim isn’t in manifest (suffix “(missing)”).
                                  Diagrams still skip missing dims (robust).
• --star-diagrams               : enable star diagram tabs.
• --diagram-two-rings N         : inner ring up to N dims, rest outer ring.
• --diagram-no-wrap-labels      : disable label wrapping in star diagrams.
• --diagram-max-label-chars INT : wrapping width per line (default 18).

(New) Tiering & lineage flags
-----------------------------
• --lineage                      : generate Lineage and Pipelines sheets.
• --lineage-diagrams             : also generate per-mart lineage diagram tabs (requires --lineage).
• Tier classification:
    By path globs:
      --tier-base-globs  "models/base/*,models/bronze/*"
      --tier-stage-globs "models/stage/*,models/silver/*"
      --tier-mart-globs  "models/mart/*,models/gold/*"
    By tags:
      --tier-base-tags  base
      --tier-stage-tags stage
      --tier-mart-tags  mart
  If none provided, defaults to path-prefix heuristic:
    models/base/* → base, models/stage/* → stage, models/mart/* → mart.

Limitations / assumptions
-------------------------
• Relationship edges are derived from dbt `relationships` tests in `manifest.json`.
• If a test references a dimension not present in the manifest:
  - Default: skip in Star Map & diagrams.
  - With --include-missing-dims: include in Star Map (suffix “(missing)”); diagrams still skip.
• Lineage edges are derived from `depends_on.nodes` (model→model) as compiled by dbt.

Changelog highlights
--------------------
2.0.1
  - BUGFIX: Pipelines sheet no longer fails when there are no mart rows (safe empty handling & sorting).
  - Robust lineage-diagram loop against column-name drift.
  - Keeps all 2.0.0 features unchanged.

2.0.0
  - Added Lineage & Pipelines sheets, tier classification, and optional Lineage diagrams.

"""

import argparse, json, re, sys, fnmatch, math, tempfile, os
from collections import defaultdict, OrderedDict
from pathlib import Path

import pandas as pd
try:
    import yaml
except Exception:
    yaml = None


# =============================================================================
# Basic helpers
# =============================================================================

def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node: dict) -> bool:
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags) -> str:
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    raw = (base or 'Sheet').strip()
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', raw)[:31] or 'Sheet'
    used_lower = {u.lower() for u in used}
    candidate = cleaned; i = 1
    while candidate.lower() in used_lower:
        suffix = f'_{i}'
        candidate = (cleaned[:31 - len(suffix)] + suffix)
        i += 1
    used.add(candidate)
    return candidate

def _normalize_to_model(raw):
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"""ref\(\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    m = re.match(r"""source\(\s*['"][^'"]+['"]\s*,\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    return s

def extract_tests_for_model(manifest, node_id):
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test':
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends:
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_model = _normalize_to_model(raw_to)
            rel_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')):
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    db, schema, alias = model_relation_identifiers(node)
    key = f'{db}.{schema}.{alias}'.lower()
    for k, v in (catalog.get('nodes') or {}).items():
        if k.lower() == key:
            return v.get('columns') or {}
    return {}

def build_relationship_rows_with_deps(manifest):
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test':
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname:
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_model = _normalize_to_model(raw_to)
        to_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
        dep_models = []
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model':
                dep_models.append((n.get('alias') or n.get('name')))
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'DepModels': dep_models,'ToModel': to_model,'ToColumn': to_field,
                     'FromModel': None,'FromColumn': from_column,'TestName': tname})
    return rows

def parse_csv_set(s: str):
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}


# =============================================================================
# YAML overlays (preserve display casing + lowercase for matching)
# =============================================================================

try:
    import yaml  # just in case re-imported above
except Exception:
    yaml = None

def load_yaml_overlays(paths_csv: str):
    overlays = {}
    if not paths_csv or not yaml: return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc: continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname: continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})

                def _norm_list(v):
                    if v is None: return []
                    if isinstance(v, (list, tuple)): return [str(x) for x in v]
                    return [str(v)]

                sk_disp = _norm_list(mm.get('surrogate_key'))
                bk_disp = _norm_list(mm.get('business_key'))
                pk_disp = _norm_list(mm.get('primary_key'))

                entry['model_meta_display'] = {
                    'surrogate_key': sk_disp,
                    'business_key':  bk_disp,
                    'primary_key':   pk_disp,
                }
                entry['model_meta_match'] = {
                    'surrogate_key': [s.lower() for s in sk_disp],
                    'business_key':  [s.lower() for s in bk_disp],
                    'primary_key':   [s.lower() for s in pk_disp],
                }

                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {'data_type': dtype, 'nullable': nullable, 'description': desc}.items() if v is not None}
    return overlays


# =============================================================================
# Diagram utilities — wrap only at camel-case boundaries (no spaces inserted)
# =============================================================================

_CAMEL_SPLIT_RE = re.compile(
    r'(?<=[A-Za-z])(?=[A-Z][a-z])|(?<=[a-z])(?=[A-Z])|(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])'
)

def _camel_chunks(s: str):
    if not s: return []
    idxs = [0]
    for m in _CAMEL_SPLIT_RE.finditer(s):
        idxs.append(m.start())
    idxs.append(len(s))
    return [s[idxs[i]:idxs[i+1]] for i in range(len(idxs)-1)]

def _wrap_preserving_string(s: str, max_chars: int, max_lines: int = 2):
    chunks = _camel_chunks(s)
    if not chunks:
        return ['']
    lines, i = [], 0
    for _ in range(max_lines):
        cur = ''
        while i < len(chunks):
            cand = cur + chunks[i]
            if len(cand) <= max_chars or cur == '':
                cur = cand; i += 1
            else:
                break
        if cur:
            lines.append(cur)
        else:
            lines.append((chunks[i][:max_chars-1] + '…') if max_chars > 1 else '…'); i += 1
        if i >= len(chunks): break
    if i < len(chunks):
        last = lines[-1]
        lines[-1] = (last[:max(1, max_chars-1)] + '…') if len(last) >= max_chars else (last + '…')
    return lines

def _clip_to_circle(cx, cy, r, x, y):
    dx, dy = x - cx, y - cy
    d = math.hypot(dx, dy) or 1e-6
    return cx + r * dx / d, cy + r * dy / d

def _clip_to_rect(cx, cy, w, h, x, y):
    dx, dy = x - cx, y - cy
    if dx == 0 and dy == 0:
        return cx, cy
    tx = (w/2) / abs(dx) if dx != 0 else float('inf')
    ty = (h/2) / abs(dy) if dy != 0 else float('inf')
    t = min(tx, ty)
    return cx + dx * t, cy + dy * t

def _palette_soft():
    return {
        'fact': '#FFF2CC',
        'dim_uniform': '#DFF5E1',
        'edge': '#8A8A8A',
        'node_edge': '#222222',
        'shadow': (0,0,0,0.08),
        'edge_text': '#333333',
    }

def draw_star_png(
    fact_label, fact_family, spokes, outfile,
    *,
    shape='roundrect',
    two_rings_threshold=None,
    font_scale=1.0,
    dpi=220,
    color_scheme='soft',
    wrap_labels=True,
    max_label_chars=18,
):
    try:
        import importlib
        plt = importlib.import_module('matplotlib.pyplot')
        from matplotlib.patches import FancyBboxPatch
    except Exception:
        return False

    if not spokes:
        return False

    PALETTE = _palette_soft()

    f_fact = int(13 * font_scale)
    f_dim  = int(10 * font_scale)
    f_edge = max(8, int(9 * font_scale))

    dims = spokes; n = len(dims)
    inner_max = two_rings_threshold if (two_rings_threshold and n > two_rings_threshold) else None
    inner = dims if not inner_max else dims[:inner_max]
    outer = [] if not inner_max else dims[inner_max:]

    inner_radius = 3.4 if len(inner) <= 12 else 3.8
    outer_radius = inner_radius + 1.1 if outer else None

    figsize = (7.5, 7.5)
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111)
    ax.set_aspect('equal'); ax.axis('off')
    cx, cy = 0.0, 0.0
    shadow_offset = 0.09

    fact_lines = (_wrap_preserving_string(fact_label, max_label_chars+4, 2)
                  if wrap_labels else [fact_label])
    longest = max(len(l) for l in fact_lines)
    rect_w = max(2.9, 0.13 * longest + 1.2)
    rect_h = (1.3 if len(fact_lines) == 1 else 1.55 + 0.42*max(0, len(fact_lines)-1))

    from matplotlib.patches import FancyBboxPatch as FBP
    ax.add_patch(FBP((cx - rect_w/2 + shadow_offset, cy - rect_h/2 - shadow_offset), rect_w, rect_h,
                     boxstyle="round,pad=0.02,rounding_size=0.25",
                     linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
    ax.add_patch(FBP((cx - rect_w/2, cy - rect_h/2), rect_w, rect_h,
                     boxstyle="round,pad=0.02,rounding_size=0.25",
                     linewidth=1.6, edgecolor=PALETTE['node_edge'], facecolor=PALETTE['fact'], zorder=2))

    if len(fact_lines) == 1:
        ax.text(cx, cy, fact_lines[0], ha='center', va='center', fontsize=f_fact, fontweight='bold', zorder=3)
    else:
        y0 = cy + (0.22 if len(fact_lines) == 2 else 0.33)
        for i, line in enumerate(fact_lines):
            ax.text(cx, y0 - i*0.34, line, ha='center', va='center',
                    fontsize=f_fact, fontweight='bold', zorder=3)

    def ring_positions(count, radius):
        if count <= 0: return []
        angles = [2*math.pi*i/count for i in range(count)]
        return [(cx + radius*math.cos(a), cy + radius*math.sin(a), a) for i, a in enumerate(angles)]

    inner_pos = ring_positions(len(inner), inner_radius)
    outer_pos = ring_positions(len(outer), outer_radius) if outer else []

    def draw_dim_node(x, y, label):
        node_edge = PALETTE['node_edge']
        face = PALETTE['dim_uniform']
        lines = (_wrap_preserving_string(label, max_label_chars, 2) if wrap_labels else [label])
        longest = max(len(l) for l in lines)
        w = max(2.1, 0.12 * longest + 0.9)
        h = (1.1 if len(lines) == 1 else 1.35)
        r = 0.35

        ax.add_patch(FBP((x - w/2 + shadow_offset, y - h/2 - shadow_offset), w, h,
                         boxstyle=f"round,pad=0.02,rounding_size={r}", linewidth=0,
                         facecolor=PALETTE['shadow'], zorder=1))
        ax.add_patch(FBP((x - w/2, y - h/2), w, h,
                         boxstyle=f"round,pad=0.02,rounding_size={r}",
                         linewidth=1.6, edgecolor=node_edge, facecolor=face, zorder=2))

        if len(lines) == 1:
            ax.text(x, y, lines[0], ha='center', va='center', fontsize=f_dim, fontweight='bold', zorder=3)
        else:
            y0 = y + 0.16
            for i, line in enumerate(lines):
                ax.text(x, y0 - i*0.28, line, ha='center', va='center', fontsize=f_dim, fontweight='bold', zorder=3)

        return w, h

    def draw_spokes(positions, items):
        for (x, y, _), spec in zip(positions, items):
            w, h = draw_dim_node(x, y, spec['dim_label'])
            rx, ry = _clip_to_rect(cx, cy, rect_w, rect_h, x, y)
            cx2, cy2 = _clip_to_circle(x, y, min(w, h)/2 - 0.02, cx, cy)
            ax.plot([rx, cx2], [ry, cy2], color=PALETTE['edge'], linewidth=1.3, zorder=1)
            if spec['edge_labels']:
                label = ', '.join([l for l in spec['edge_labels'] if l])
                midx = rx*0.35 + cx2*0.65
                midy = ry*0.35 + cy2*0.65
                dx, dy = cx2 - rx, cy2 - ry
                L = math.hypot(dx, dy) or 1.0
                off = 0.17
                ex = midx - off * dy / L
                ey = midy + off * dx / L
                if len(label) > 32:
                    parts, cur, total = [], [], 0
                    for tok in label.split(', '):
                        if total + len(tok) > 28 and cur:
                            parts.append(', '.join(cur)); cur = [tok]; total = len(tok)
                        else:
                            cur.append(tok); total += (len(tok) + 2)
                    if cur: parts.append(', '.join(cur))
                    label = '\n'.join(parts)
                ax.text(ex, ey, label, ha='center', va='center', fontsize=f_edge, color=PALETTE['edge_text'], zorder=3)

    inner_pos = ring_positions(len(inner), inner_radius)
    draw_spokes(inner_pos, inner)
    if outer:
        outer_pos = ring_positions(len(outer), outer_radius)
        draw_spokes(outer_pos, outer)

    fig.tight_layout(pad=0.55)
    fig.savefig(outfile, dpi=dpi)
    plt.close(fig)
    return True


# =============================================================================
# Lineage helpers
# =============================================================================

def _match_any_glob(path_str: str, globs_lower: set) -> bool:
    p = (path_str or '').lower()
    for g in globs_lower:
        if fnmatch.fnmatch(p, g):
            return True
    return False

def classify_tier(node, *, base_globs, stage_globs, mart_globs, base_tags, stage_tags, mart_tags):
    path_str = (node.get('path') or '').lower()
    tags = {t.lower() for t in (node.get('tags') or [])}
    if _match_any_glob(path_str, base_globs) or (tags & base_tags):
        return 'base'
    if _match_any_glob(path_str, stage_globs) or (tags & stage_tags):
        return 'stage'
    if _match_any_glob(path_str, mart_globs) or (tags & mart_tags):
        return 'mart'
    if not base_globs and not stage_globs and not mart_globs:
        if path_str.startswith('models/base/'):  return 'base'
        if path_str.startswith('models/stage/'): return 'stage'
        if path_str.startswith('models/mart/'):  return 'mart'
    return 'other'

def lineage_edges(manifest):
    edges = []
    nodes = manifest.get('nodes', {})
    for _, n in nodes.items():
        if not node_is_model(n):
            continue
        to_alias = (n.get('alias') or n.get('name') or '').lower()
        for dep in (n.get('depends_on') or {}).get('nodes') or []:
            dn = nodes.get(dep)
            if dn and node_is_model(dn):
                from_alias = (dn.get('alias') or dn.get('name') or '').lower()
                edges.append((from_alias, to_alias))
    return edges


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true')
    parser.add_argument('--materializations', type=str, default='')
    parser.add_argument('--schemas', type=str, default='')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='')
    parser.add_argument('--exclude-sheet-tags', type=str, default='')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='')
    parser.add_argument('--no-relationships', action='store_true',
                        help='Omit the Relationships sheet.')
    parser.add_argument('--include-missing-dims', action='store_true',
                        help='Include Star Map rows where the dimension is missing; diagrams still skip.')

    # Star diagram flags
    parser.add_argument('--star-diagrams', action='store_true')
    parser.add_argument('--diagram-shape', choices=['roundrect','circle'], default='roundrect')
    parser.add_argument('--diagram-dedupe-roles', action='store_true')
    parser.add_argument('--diagram-two-rings', type=int, default=None)
    parser.add_argument('--diagram-font-scale', type=float, default=1.0)
    parser.add_argument('--diagram-dpi', type=int, default=240)
    parser.add_argument('--diagram-color-scheme', choices=['soft'], default='soft')
    parser.add_argument('--diagram-legend', action='store_true')
    parser.add_argument('--diagram-no-wrap-labels', action='store_true')
    parser.add_argument('--diagram-max-label-chars', type=int, default=18)

    # Lineage flags
    parser.add_argument('--lineage', action='store_true', help='Generate lineage sheets (Lineage, Pipelines).')
    parser.add_argument('--lineage-diagrams', action='store_true', help='Generate per-mart lineage diagrams.')
    parser.add_argument('--tier-base-globs', type=str, default='')
    parser.add_argument('--tier-stage-globs', type=str, default='')
    parser.add_argument('--tier-mart-globs', type=str, default='')
    parser.add_argument('--tier-base-tags', type=str, default='')
    parser.add_argument('--tier-stage-tags', type=str, default='')
    parser.add_argument('--tier-mart-tags', type=str, default='')

    args = parser.parse_args()

    manifest = load_json(args.manifest); catalog = load_json(args.catalog)

    # Filter models (materializations + include-views)
    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    alias_to_node, name_to_alias, alias_to_display = {}, {}, {}
    for n in filtered:
        alias_display = (n.get('alias') or n.get('name') or '')
        alias_lower = alias_display.lower()
        name_lower  = (n.get('name') or '').lower()
        alias_to_node[alias_lower] = n
        alias_to_display[alias_lower] = alias_display
        name_to_alias[alias_lower] = alias_lower
        if name_lower:
            name_to_alias[name_lower] = alias_lower

    def canon(token: str):
        if not token: return None
        t = str(token).strip().strip('"').lower()
        return name_to_alias.get(t, t)

    # Relationships → normalized rows
    raw_rel_rows = build_relationship_rows_with_deps(manifest)
    norm_rel_rows = []
    for r in raw_rel_rows:
        dep_aliases = [canon(x) for x in (r.get('DepModels') or [])]
        to_alias    = canon(r.get('ToModel'))
        from_alias = None
        for d in dep_aliases:
            if d and d != to_alias:
                from_alias = d; break
        if not from_alias and dep_aliases:
            from_alias = dep_aliases[0]
        if not to_alias or not from_alias:
            continue
        norm_rel_rows.append({
            'FromModel': from_alias,
            'FromColumn': r.get('FromColumn'),
            'ToModel': to_alias,
            'ToColumn': r.get('ToColumn'),
            'TestName': r.get('TestName'),
        })

    spokes_raw = defaultdict(list)
    for r in norm_rel_rows:
        frm = r['FromModel']; to = r['ToModel']; col = r.get('FromColumn')
        if not frm or not to:
            continue
        if (to not in alias_to_display) and (not args.include_missing_dims):
            continue
        if to in alias_to_display:
            spokes_raw[frm].append((to, col))

    yaml_overlays = load_yaml_overlays(args.schemas)

    ex_prefixes   = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags       = parse_csv_set(args.exclude_sheet_tags)
    ex_mats       = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XLImage

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    alt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    TAB_BLUE   = "5B9BD5"
    TAB_PURPLE = "C9C2F3"
    TAB_PEACH  = "F8CBAD"
    TAB_GREEN  = "A9D18E"

    def style_table(ws, top_header_row: int, width_map: dict, wrap_cols: list, freeze=True):
        if freeze:
            ws.freeze_panes = f"A{top_header_row+1}"
        max_col = ws.max_column
        for col_idx in range(1, max_col+1):
            c = ws.cell(row=top_header_row, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(vertical="center")
        last_col_letter = ws.cell(row=top_header_row, column=max_col).column_letter
        ws.auto_filter.ref = f"A{top_header_row}:{last_col_letter}{ws.max_row}"
        for r in range(top_header_row+1, ws.max_row+1):
            for letter in wrap_cols:
                col_index = ws[letter + str(top_header_row)].column
                ws.cell(row=r, column=col_index).alignment = Alignment(wrap_text=True, vertical="top")
        for letter, width in width_map.items():
            ws.column_dimensions[letter].width = width

    widths_summary = {"A": 26, "B": 12, "C": 16, "D": 16, "E": 16, "F": 36, "G": 28, "H": 36, "I": 24, "J": 50}
    widths_rel     = {"A": 28, "B": 30, "C": 28, "D": 30, "E": 26}
    widths_star    = {"A": 28, "B": 28, "C": 36, "D": 30}
    widths_lineage = {"A":28,"B":10,"C":10,"D":38,"E":4,"F":28,"G":10,"H":10,"I":38}
    widths_pipes   = {"A":30,"B":60,"C":60}

    # Summary
    summary_rows = []
    for alias_lower, n in alias_to_node.items():
        db, schema, alias_disp = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias_lower, n.get('tags', []))
        summary_rows.append({'Model': alias_disp,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias_disp}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    df_summary = pd.DataFrame(summary_rows, columns=['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description'])
    if not df_summary.empty:
        df_summary = df_summary.sort_values(['Kind','Model'])

    # Relationships
    df_rel = pd.DataFrame([
        {'FromModel': alias_to_display.get(r['FromModel'], r['FromModel']),
         'FromColumn': r.get('FromColumn'),
         'ToModel': alias_to_display.get(r['ToModel'], r['ToModel']) if r['ToModel'] in alias_to_display else (r['ToModel'] or '(missing)'),
         'ToColumn': r.get('ToColumn'),
         'TestName': r.get('TestName')}
        for r in norm_rel_rows
    ]) if norm_rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])

    # Star Map
    star_map = OrderedDict()
    for r in norm_rel_rows:
        f = alias_to_display.get(r['FromModel'], r['FromModel'])
        if r['ToModel'] in alias_to_display:
            d = alias_to_display.get(r['ToModel'], r['ToModel'])
        else:
            if not args.include_missing_dims:
                continue
            d_raw = r['ToModel'] or '(missing)'
            d = f"{d_raw}(missing)" if not d_raw.endswith('(missing)') else d_raw
        e = star_map.setdefault((f,d), {'FactModel': f, 'DimensionModel': d, 'FactFKColumn': set(), 'DimKeyColumn': set()})
        if r.get('FromColumn'):
            e['FactFKColumn'].add(r.get('FromColumn'))
        if r.get('ToColumn'):
            e['DimKeyColumn'].add(r.get('ToColumn'))

    df_star = (
        pd.DataFrame([
            {'FactModel': e['FactModel'],
             'DimensionModel': e['DimensionModel'],
             'FactFKColumn': ', '.join(sorted(e['FactFKColumn'])) if e['FactFKColumn'] else '',
             'DimKeyColumn': ', '.join(sorted(e['DimKeyColumn'])) if e['DimKeyColumn'] else ''}
            for e in star_map.values()
        ], columns=['FactModel','DimensionModel','FactFKColumn','DimKeyColumn'])
        if star_map else
        pd.DataFrame(columns=['FactModel','DimensionModel','FactFKColumn','DimKeyColumn'])
    )

    used_sheet_names = set()

    # ---- Tiering & Lineage ----
    base_globs  = parse_csv_set(args.tier_base_globs)
    stage_globs = parse_csv_set(args.tier_stage_globs)
    mart_globs  = parse_csv_set(args.tier_mart_globs)
    base_tags   = parse_csv_set(args.tier_base_tags)
    stage_tags  = parse_csv_set(args.tier_stage_tags)
    mart_tags   = parse_csv_set(args.tier_mart_tags)

    alias_to_tier = {}
    for a, n in alias_to_node.items():
        alias_to_tier[a] = classify_tier(
            n,
            base_globs=base_globs, stage_globs=stage_globs, mart_globs=mart_globs,
            base_tags=base_tags, stage_tags=stage_tags, mart_tags=mart_tags
        )

    all_edges = lineage_edges(manifest)
    edges = [(u, v) for (u, v) in all_edges if (u in alias_to_node and v in alias_to_node)]

    def _mat(n): return (n.get('config') or {}).get('materialized')
    df_lineage = pd.DataFrame(columns=['FromModel','FromTier','FromMat','FromPath','→','ToModel','ToTier','ToMat','ToPath'])
    if args.lineage and edges:
        rows = []
        for u, v in edges:
            nu, nv = alias_to_node[u], alias_to_node[v]
            rows.append({
                'FromModel': alias_to_display.get(u, u),
                'FromTier':  alias_to_tier.get(u, 'other'),
                'FromMat':   _mat(nu),
                'FromPath':  nu.get('path'),
                '→': '→',
                'ToModel':   alias_to_display.get(v, v),
                'ToTier':    alias_to_tier.get(v, 'other'),
                'ToMat':     _mat(nv),
                'ToPath':    nv.get('path'),
            })
        df_lineage = pd.DataFrame(rows, columns=['FromModel','FromTier','FromMat','FromPath','→','ToModel','ToTier','ToMat','ToPath'])

    # Build Pipelines (safe for empty)
    df_pipes = pd.DataFrame(columns=['MartModel','Stages','Bases'])
    parents = defaultdict(set)
    if args.lineage and edges:
        children = defaultdict(set)
        for u, v in edges:
            parents[v].add(u); children[u].add(v)

        mart_aliases = [a for a, t in alias_to_tier.items() if t == 'mart']
        rows = []
        for m in mart_aliases:
            seen, frontier = set(), [m]
            stages, bases = set(), set()
            while frontier:
                nxt = []
                for x in frontier:
                    for p in parents.get(x, []):
                        if p in seen: continue
                        seen.add(p); nxt.append(p)
                        tier = alias_to_tier.get(p, 'other')
                        if tier == 'stage': stages.add(alias_to_display.get(p, p))
                        elif tier == 'base': bases.add(alias_to_display.get(p, p))
                frontier = nxt
            rows.append({
                'MartModel': alias_to_display.get(m, m),
                'Stages': ', '.join(sorted(stages)),
                'Bases':  ', '.join(sorted(bases)),
            })
        # ---- BUGFIX: guard empty & column existence before sort ----
        if rows:
            df_pipes = pd.DataFrame(rows, columns=['MartModel','Stages','Bases'])
            if 'MartModel' in df_pipes.columns and not df_pipes.empty:
                df_pipes = df_pipes.sort_values('MartModel')
        else:
            df_pipes = pd.DataFrame(columns=['MartModel','Stages','Bases'])

    # ---- Write Excel ----
    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        wb = writer.book; ws = wb['Summary']
        ws.sheet_properties.tabColor = TAB_BLUE
        style_table(ws, top_header_row=1, width_map=widths_summary,
                    wrap_cols=['A','F','G','H','I','J'], freeze=True)

        if not args.no_relationships and (df_rel is not None) and (not df_rel.empty):
            df_rel.to_excel(writer, index=False, sheet_name='Relationships')
            ws = wb['Relationships']; ws.sheet_properties.tabColor = TAB_BLUE
            style_table(ws, top_header_row=1, width_map=widths_rel,
                        wrap_cols=['A','B','C','D','E'], freeze=True)

        if not df_star.empty:
            df_star.to_excel(writer, index=False, sheet_name='Star Map')
            ws = wb['Star Map']; ws.sheet_properties.tabColor = TAB_BLUE
            style_table(ws, top_header_row=1, width_map=widths_star,
                        wrap_cols=['A','B','C','D'], freeze=True)

        if args.lineage and not df_lineage.empty:
            df_lineage.to_excel(writer, index=False, sheet_name='Lineage')
            ws = wb['Lineage']; ws.sheet_properties.tabColor = TAB_BLUE
            style_table(ws, top_header_row=1, width_map=widths_lineage,
                        wrap_cols=['D','I'], freeze=True)

        if args.lineage and not df_pipes.empty:
            df_pipes.to_excel(writer, index=False, sheet_name='Pipelines')
            ws = wb['Pipelines']; ws.sheet_properties.tabColor = TAB_BLUE
            style_table(ws, top_header_row=1, width_map=widths_pipes,
                        wrap_cols=['B','C'], freeze=True)

        # Per-model sheets
        for alias_lower, n in alias_to_node.items():
            db, schema, alias_disp = model_relation_identifiers(n)
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''

            def should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                     prefixes, tags_excl, mats_excl, path_globs) -> bool:
                if any(alias_lower.startswith(p) for p in prefixes): return True
                if any(t in tags_lower for t in tags_excl): return True
                if mat_lower in mats_excl: return True
                ps = (path_str or '')
                for g in path_globs:
                    if fnmatch.fnmatch(ps.lower(), g): return True
                return False

            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            model_sheet_name = safe_sheet_name(alias_disp or 'Model', used_sheet_names)

            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            overlay_entry = (yaml_overlays.get(alias_lower, {}) or {})
            overlay_cols = overlay_entry.get('columns', {}) or {}
            meta_disp = overlay_entry.get('model_meta_display', {}) or {}
            meta_match = overlay_entry.get('model_meta_match', {}) or {}

            sk_display = meta_disp.get('surrogate_key', []) or []
            bk_display = meta_disp.get('business_key', []) or []
            pk_display = meta_disp.get('primary_key', []) or []
            pk_display_final = pk_display if pk_display else sk_display

            pk_match = set((meta_match.get('primary_key', []) or []))
            sk_match = set((meta_match.get('surrogate_key', []) or []))
            explicit_pk_names = {*pk_match, *sk_match}

            rows = []
            catalog_cols = get_catalog_columns_for_model(catalog, n)
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or (meta.get('type') or '')
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            db, schema, alias_disp = model_relation_identifiers(n)
            info_rows = [
                {'Column': 'Model',         'DataType': alias_disp,                                'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Kind',          'DataType': classify_model_kind(alias_lower, n.get('tags', [])), 'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Materialization','DataType': (n.get('config') or {}).get('materialized'),        'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Relation',      'DataType': f'{db}.{schema}.{alias_disp}',             'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Tags',          'DataType': ','.join(n.get('tags') or []),             'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'SurrogateKey',  'DataType': ', '.join(sk_display),                     'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'BusinessKey',   'DataType': ', '.join(bk_display),                     'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'PrimaryKey',    'DataType': ', '.join(pk_display_final),               'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Description',   'DataType': (n.get('description') or ''),              'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            df_meta.to_excel(writer, index=False, sheet_name=model_sheet_name, header=False, startrow=0)
            start_row = len(df_meta.index) + 1
            df.to_excel(writer, index=False, sheet_name=model_sheet_name, startrow=start_row)

            ws = wb[model_sheet_name]
            kind = classify_model_kind(alias_lower, n.get('tags', []))
            ws.sheet_properties.tabColor = TAB_PEACH if kind == 'Fact' else (TAB_PURPLE if kind == 'Dimension' else TAB_BLUE)

            ws.freeze_panes = f"A{start_row+2}"
            hdr = start_row + 1

            for col_idx in range(1, 8+1):
                c = ws.cell(row=hdr, column=col_idx)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(vertical="center")

            last_col_letter = ws.cell(row=hdr, column=8).column_letter
            ws.auto_filter.ref = f"A{hdr}:{last_col_letter}{ws.max_row}"

            for r in range(hdr+1, ws.max_row+1):
                ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")

            from openpyxl.styles import Font as _Font
            for r in range(1, len(df_meta.index)+1):
                ws.cell(row=r, column=1).font = _Font(bold=True)

            for r in range(hdr+1, ws.max_row+1):
                if (r - (hdr+1)) % 2 == 1:
                    for c in range(1, 8+1):
                        ws.cell(row=r, column=c).fill = alt_fill

            for col_letter, width in {"A":28,"B":20,"C":10,"D":80,"E":28,"F":6,"G":6,"H":12}.items():
                ws.column_dimensions[col_letter].width = width

        # Star diagram tabs
        if args.star_diagrams:
            tmpdir = tempfile.mkdtemp(prefix="dbt_star_")
            wb = writer.book
            used = set(wb.sheetnames)

            for center_alias_lower, pairs in sorted(spokes_raw.items()):
                if args.diagram_dedupe_roles:
                    by_dim = OrderedDict()
                    for dim_alias_lower, fk_col in pairs:
                        by_dim.setdefault(dim_alias_lower, []).append(fk_col or '')
                    spokes = []
                    for dim_alias_lower, fk_cols in by_dim.items():
                        spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                       'dim_family': 'uniform',
                                       'edge_labels': [c for c in fk_cols if c]})
                else:
                    spokes = []
                    for dim_alias_lower, fk_col in pairs:
                        spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                       'dim_family': 'uniform',
                                       'edge_labels': [fk_col] if fk_col else []})

                if not spokes:
                    continue

                fact_label = alias_to_display.get(center_alias_lower, center_alias_lower)
                img_path = os.path.join(tmpdir, f"{center_alias_lower}.png")

                drew = draw_star_png(
                    fact_label, 'other', spokes, img_path,
                    shape='roundrect',
                    two_rings_threshold=args.diagram_two_rings,
                    font_scale=args.diagram_font_scale,
                    dpi=args.diagram_dpi,
                    color_scheme=args.diagram_color_scheme,
                    wrap_labels=not args.diagram_no_wrap_labels,
                    max_label_chars=max(10, args.diagram_max_label_chars),
                )
                if not drew:
                    continue

                proposed = f"Star-Fact {fact_label}"
                if any(c in proposed for c in r'[]:*?\/'):
                    proposed = f"Star Fact {fact_label}"
                sheet_name = safe_sheet_name(proposed, used)

                ws = wb.create_sheet(title=sheet_name)
                ws.sheet_properties.tabColor = TAB_GREEN
                try:
                    img = XLImage(img_path)
                    ws.add_image(img, "A1")
                except Exception:
                    pass

                if args.diagram_legend:
                    ws["J2"] = ("Legend: dimensions shown in light green; fact in light yellow.\n"
                                "Spoke text = fact FK column(s).")

        # Lineage diagrams (robust to column-label drift)
        if args.lineage and args.lineage_diagrams and not df_pipes.empty:
            import importlib
            try:
                plt = importlib.import_module('matplotlib.pyplot')
                from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
            except Exception:
                plt = None

            if plt:
                tmpdir2 = tempfile.mkdtemp(prefix="dbt_lineage_")
                wb = writer.book
                used = set(wb.sheetnames)

                disp_to_alias = {v: k for k, v in alias_to_display.items()}

                def _get_any(row, *names):
                    for n in names:
                        if n in row:
                            return row[n]
                        for k in row.index:
                            if str(k).strip().lower() == str(n).strip().lower():
                                return row[k]
                    return row.iloc[0]

                warned_cols = False

                for _, prow in df_pipes.iterrows():
                    if 'MartModel' not in df_pipes.columns and not warned_cols:
                        print(f"[lineage-diagrams] Pipelines columns detected: {list(df_pipes.columns)}")
                        warned_cols = True

                    mart_disp = _get_any(prow, 'MartModel')

                    m_alias = disp_to_alias.get(mart_disp)
                    if not m_alias:
                        m_alias = next((a for d, a in disp_to_alias.items()
                                        if str(d).lower() == str(mart_disp).lower()), None)
                    if not m_alias:
                        continue

                    upstream = set(); frontier = [m_alias]
                    while frontier:
                        nxt = []
                        for x in frontier:
                            for p in parents.get(x, []):
                                if p in upstream: continue
                                upstream.add(p); nxt.append(p)
                        frontier = nxt

                    tiers = {'base': [], 'stage': [], 'mart': [m_alias]}
                    for u in upstream:
                        t = alias_to_tier.get(u, 'other')
                        if t in ('base','stage','mart'):
                            tiers[t].append(u)

                    if not tiers['base'] and not tiers['stage']:
                        continue

                    x_pos = {'base': 0.5, 'stage': 3.5, 'mart': 6.5}
                    y_gap = 0.9

                    fig = plt.figure(figsize=(7.6, 3.6))
                    ax = fig.add_subplot(111); ax.set_aspect('equal'); ax.axis('off')

                    def draw_node(x, y, label, face, bold=True):
                        lines = _wrap_preserving_string(label, 18, 2)
                        w, h, r = 2.6, 1.0, 0.25
                        ax.add_patch(FancyBboxPatch((x-w/2+0.07, y-h/2-0.07), w, h,
                                                    boxstyle=f"round,pad=0.02,rounding_size={r}",
                                                    linewidth=0, facecolor=(0,0,0,0.08)))
                        ax.add_patch(FancyBboxPatch((x-w/2, y-h/2), w, h,
                                                    boxstyle=f"round,pad=0.02,rounding_size={r}",
                                                    linewidth=1.4, edgecolor="#222", facecolor=face))
                        if len(lines)==1:
                            ax.text(x, y, lines[0], ha='center', va='center', fontsize=10, fontweight='bold' if bold else None)
                        else:
                            ax.text(x, y+0.16, lines[0], ha='center', va='center', fontsize=10, fontweight='bold' if bold else None)
                            ax.text(x, y-0.12, lines[1], ha='center', va='center', fontsize=10, fontweight='bold' if bold else None)

                    PA = _palette_soft()
                    node_face = {'base': '#D9E1F2', 'stage': '#E4D9F2', 'mart': PA['fact']}

                    coords = {}
                    for tier in ('base','stage','mart'):
                        nodes_here = tiers[tier]
                        if not nodes_here: continue
                        y0 = (len(nodes_here)-1)*y_gap/2
                        for i, a in enumerate(sorted(nodes_here)):
                            y = 1.8 - (i*y_gap - y0)
                            x = x_pos[tier]
                            label = alias_to_display.get(a, a)
                            draw_node(x, y, label, node_face[tier], bold=True)
                            coords[a] = (x, y)

                    from matplotlib.patches import FancyArrowPatch as FAP
                    for u, v in edges:
                        if u in coords and v in coords:
                            x1, y1 = coords[u]; x2, y2 = coords[v]
                            ax.add_patch(FAP((x1+1.3, y1), (x2-1.3, y2),
                                             arrowstyle='-|>', mutation_scale=10,
                                             linewidth=1.2, color='#7A7A7A'))

                    fig.tight_layout(pad=0.4)
                    png = os.path.join(tmpdir2, f"lineage_{m_alias}.png")
                    fig.savefig(png, dpi=220); plt.close(fig)

                    proposed = f"Lineage-{mart_disp}"
                    if any(c in proposed for c in r'[]:*?\/'):
                        proposed = f"Lineage {mart_disp}"
                    lname = safe_sheet_name(proposed, used)

                    ws = wb.create_sheet(title=lname)
                    ws.sheet_properties.tabColor = TAB_GREEN
                    try:
                        img = XLImage(png); ws.add_image(img, "A1")
                    except Exception:
                        pass

    print(f"Wrote: {args.out}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
