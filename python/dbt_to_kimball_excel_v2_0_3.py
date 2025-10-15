#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
dbt_to_kimball_excel.py
-----------------------
Version: 2.0.3 (2025-10-15)

Generate a Kimball-style Excel workbook from a dbt Core project:

1) Reads dbt's manifest.json + catalog.json (+ optional schemas.yml files)
2) Produces a workbook with:
   - Summary (all models, kinds, materializations, FQN, tags, path, description)
   - One sheet per model with table meta + column meta (frozen header, blue header fill,
     word-wrapped Description/Tests, alternating green-bar rows)
   - Relationships (rows derived from column-level relationship tests; shows From/To model+column)
   - Star Map (FactModel, DimensionModel, FactFKColumn, DimKeyColumn)
   - Optional star diagrams (one per fact) embedded as PNG, with rounded shapes and
     FK column labels, light-yellow fact, pastel dimensions, deduped role-players
   - Optional textual Lineage sheet (base -> stage -> mart) plus per-mart lineage diagrams
     (Matplotlib and/or Graphviz)

CLI (examples)
--------------
# minimal
python scripts/dbt_to_kimball_excel.py --manifest target/manifest.json --catalog target/catalo g.json --out documentation/kimball_model_catalog.xlsx

# include schemas.yml (comma-separated paths)
python scripts/dbt_to_kimball_excel.py --manifest target/manifest.json --catalog target/catalog.json ^
  --schemas "models/mart/schemas.yml,models/stage/schemas.yml" ^
  --out documentation/kimball_model_catalog.xlsx

# exclude noisy tabs and add diagrams + lineage (Graphviz if present)
python scripts/dbt_to_kimball_excel.py --manifest target/manifest.json --catalog target/catalog.json ^
  --schemas "models/mart/schemas.yml,models/stage/schemas.yml" ^
  --exclude-sheet-prefixes "base_,stage_" --exclude-sheet-materializations "view" ^
  --exclude-sheet-tags "stage,logging" --exclude-sheet-path-globs "models/base/*,models/stage/*,models/logging/*" ^
  --star-diagrams --diagram-two-rings 12 --diagram-color-scheme soft --diagram-legend ^
  --lineage --lineage-diagrams --gv-lineage --gv-format png ^
  --out documentation/kimball_model_catalog.xlsx

Notes & Design Choices
----------------------
- We parse column tests from either `tests:` or `data_tests:` in schemas.yml (dbt 1.x/ newer style).
- Relationship tests recognized in flexible forms, e.g.:
    relationships: { to: ref('DimUser'), field: 'UserSKey' }
    relationships: { to: source('...'), field: '...', column: '...' }
    relationships: { to: 'DimUser', field: 'UserSKey' }  # we normalize `to` to a model alias/name
- Column meta supported: description, data_type, meta.nullable, plus table-level meta like
  surrogate_key, business_key, primary_key, tags.
- “Kind” classifier heuristic:
    - tags or name prefix ‘fact’ => Fact
    - tags or name prefix ‘dim’  => Dimension
    - config.materialized values recognized
    - otherwise Stage/Base/View/Ephemeral/Unknown as applicable
- Robust “mart list” for lineage diagrams: union of Star Map facts, tiered ‘mart’, kind in {fact, dimension}, tag ‘mart’.
- We **no longer** skip lineage diagrams if only mart→mart edges exist.
- Graphviz lineage requires Graphviz binaries on PATH (`dot -V`). If missing, Graphviz diagrams are skipped gracefully.

Changelog (recent)
------------------
- Fixed `safe_sheet_name` and CLI parsing typos
- Freeze panes: header row stays visible; filters on headers
- Word-wrap for Description & Tests; alternating green-bar rows
- Summary/Relationships/Star Map headers styled (blue background, bold black text)
- Star Map includes FK columns; Star sheet names use “Star-Fact…” with safe Excel names
- Matplotlib star diagrams: rounded shapes, consistent palette, wrapped labels (camel-case aware), two-ring layout
- Graphviz lineage diagrams: render even without base/stage, robust mart detection, embedded as PNG tabs
- Fixed openpyxl column-letter usage via `get_column_letter`
- Sheet-tab styling via openpyxl (tabColor & small caps in labels where appropriate)

Dependencies
------------
pip install:
  pyyaml pandas openpyxl matplotlib networkx graphviz

(For graphviz diagrams, also install system Graphviz so `dot -V` works.)

"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import tempfile
from collections import defaultdict, namedtuple
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

# Third-party
import yaml
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter, column_index_from_string

# Matplotlib for diagrams
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch

# Optional Graphviz
try:
    import graphviz
    _HAS_GRAPHVIZ = True
except Exception:
    _HAS_GRAPHVIZ = False


# ============ Utilities ============

def pascalize(name: str) -> str:
    """Return a display-friendly name; prefer the alias casing if already Pascal/Camel,
    otherwise do a simple split/capitalize heuristic."""
    if not name:
        return name
    # if already has case transitions, keep it
    if re.search(r'[a-z][A-Z]', name):
        return name
    # if snake_case
    if '_' in name:
        return ''.join(part.capitalize() for part in name.split('_') if part)
    # fall back to capitalize first
    return name[:1].upper() + name[1:]


def split_camel(label: str) -> List[str]:
    """Split camel/Pascal case words for wrapping."""
    if not label:
        return [label]
    parts = re.findall(r'[A-Z]?[a-z0-9]+|[A-Z]+(?![a-z])', label)
    return parts if parts else [label]


def wrap_label(label: str, max_chars: int = 14) -> str:
    """Wrap label at camel boundaries trying to keep lines below max_chars."""
    words = split_camel(label)
    lines = []
    cur = ""
    for w in words:
        if not cur:
            cur = w
        elif len(cur) + 1 + len(w) <= max_chars:
            cur += " " + w
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return "\n".join(lines)


def safe_sheet_name(base: str, used: set) -> str:
    """Excel sheet name safety with de-duplication (case-insensitive)."""
    raw = (base or 'Sheet').strip()
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', raw)[:31] or 'Sheet'
    used_lower = {u.lower() for u in used}
    candidate = cleaned
    i = 1
    while candidate.lower() in used_lower:
        suffix = f'_{i}'
        candidate = (cleaned[:31 - len(suffix)] + suffix)
        i += 1
    used.add(candidate)
    return candidate


def is_glob_matched(path: str, patterns: List[str]) -> bool:
    from fnmatch import fnmatch
    for pat in patterns:
        if fnmatch(path.replace("\\", "/"), pat):
            return True
    return False


# ============ Loading dbt artifacts ============

def load_manifest(p: Path) -> Dict[str, Any]:
    with p.open('r', encoding='utf-8') as f:
        return json.load(f)


def load_catalog(p: Path) -> Dict[str, Any]:
    if not p or not p.exists():
        return {}
    with p.open('r', encoding='utf-8') as f:
        return json.load(f)


def load_schemas_yml(paths_csv: str) -> Dict[str, Any]:
    """Load and normalize schemas.yml content from comma-separated path list."""
    if not paths_csv:
        return {}
    result = {}
    for raw in [x.strip() for x in paths_csv.split(',') if x.strip()]:
        ypath = Path(raw)
        if not ypath.exists():
            continue
        data = yaml.safe_load(ypath.read_text(encoding='utf-8')) or {}
        for m in (data.get('models') or []):
            name = (m.get('name') or '').strip()
            if not name:
                continue
            # merge by name (later files override earlier)
            result[name.lower()] = m
    return result


# ============ Model/introspection helpers ============

def classify_kind(name: str, tags: List[str], mat: str) -> str:
    lo = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if 'fact' in tags_lower or lo.startswith('fact'):
        return 'fact'
    if 'dim' in tags_lower or lo.startswith('dim'):
        return 'dimension'
    if mat in {'incremental', 'table', 'view', 'materializedview'} and (lo.startswith('stage') or 'stage' in tags_lower):
        return 'stage'
    if 'base' in tags_lower or lo.startswith('base'):
        return 'base'
    if mat == 'view':
        return 'view'
    if mat == 'ephemeral':
        return 'ephemeral'
    if mat:
        return mat
    return 'unknown'


def extract_relationship_tests(col_entry: Dict[str, Any]) -> List[Dict[str, str]]:
    """Normalize relationship tests from dbt schema styles (tests: or data_tests:)."""
    out = []
    tests_key = 'tests' if 'tests' in col_entry else 'data_tests' if 'data_tests' in col_entry else None
    if not tests_key:
        return out
    tests = col_entry.get(tests_key) or []
    # list of dicts or strings
    for t in tests:
        if isinstance(t, dict) and 'relationships' in t:
            rel = t['relationships'] or {}
            to_val = rel.get('to') or rel.get('to_model') or ''
            # normalize ref('X') to X
            to_model = str(to_val)
            m = re.search(r"ref\(['\"]([^'\"]+)['\"]\)", to_model)
            if m:
                to_model = m.group(1)
            else:
                # bare model name
                to_model = re.sub(r'[^A-Za-z0-9_]', '', to_model)
            field = rel.get('field') or rel.get('to_field') or rel.get('to_column') or ''
            to_col = rel.get('column') or rel.get('to_column') or field
            out.append({'to_model': to_model, 'to_column': str(to_col)})
    return out


def read_schema_columns(schema_m: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """Return (table_meta, columns_list) from a normalized schemas.yml model entry."""
    tmeta = {}
    if not schema_m:
        return tmeta, []
    # table-level info
    for key in ('description', 'surrogate_key', 'business_key', 'primary_key'):
        if key in schema_m:
            tmeta[key] = schema_m.get(key)
    tmeta['tags'] = ','.join(schema_m.get('tags') or [])
    # columns
    cols = []
    for c in (schema_m.get('columns') or []):
        item = {
            'name': c.get('name'),
            'description': c.get('description') or '',
            'data_type': c.get('data_type') or '',
            'nullable': str(((c.get('meta') or {}).get('nullable')) is not False).upper().replace('TRUE', 'Y').replace('FALSE','N'),
            'tests': [],
            'is_pk': '',
            'is_fk': '',
            'source': 'YAML'
        }
        # detect PK
        if c.get('meta') and (c['meta'].get('identity') or c['meta'].get('is_primary_key')):
            item['is_pk'] = 'Y'
        # tests
        rels = extract_relationship_tests(c)
        if rels:
            item['is_fk'] = 'Y'
        item['tests'] = rels
        cols.append(item)
    return tmeta, cols


# ============ Excel styling helpers ============

BLUE = "DDEBF7"
PALE_YELLOW = "FFF2CC"
PALE_GREEN = "E2EFDA"
ALT_GREEN = "F2F9ED"

header_fill = PatternFill("solid", fgColor=BLUE)
header_font = Font(bold=True, color="000000")

def style_table(ws: Worksheet, top_header_row: int, width_map: Dict[str, float], wrap_cols: List[str], freeze=True):
    """Apply standard formatting to a tabular area with header at top_header_row."""
    if freeze:
        ws.freeze_panes = f"A{top_header_row+1}"

    max_col = ws.max_column

    # header styling
    for col_idx in range(1, max_col+1):
        c = ws.cell(row=top_header_row, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    # filters across full width
    last_col_letter = get_column_letter(max_col)
    ws.auto_filter.ref = f"A{top_header_row}:{last_col_letter}{ws.max_row}"

    # alternating green-bar
    thin = Side(style="thin", color="DDDDDD")
    for r in range(top_header_row+1, ws.max_row+1):
        fill = PatternFill("solid", fgColor=(ALT_GREEN if (r % 2 == 0) else "FFFFFF"))
        for col_idx in range(1, max_col+1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = fill
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # wrap certain columns
    for r in range(top_header_row+1, ws.max_row+1):
        for letter in wrap_cols:
            col_index = column_index_from_string(letter)
            ws.cell(row=r, column=col_index).alignment = Alignment(wrap_text=True, vertical="top")

    # column widths
    for letter, width in width_map.items():
        ws.column_dimensions[letter].width = width


# ============ Diagram helpers (Matplotlib) ============

def draw_star_png(out_png: Path,
                  fact_label: str,
                  dim_nodes: List[Tuple[str, Optional[str]]],
                  fk_labels: List[str],
                  two_rings: int = 12,
                  font_scale: float = 1.0,
                  color_scheme: str = "soft",
                  wrap_labels: bool = True):
    """Draw a star (fact center, dimensions around) and save to PNG."""
    W, H = 1200, 900
    fig = plt.figure(figsize=(W/100, H/100), dpi=100)
    ax = plt.gca()
    ax.set_aspect('equal')
    ax.axis('off')

    # colors
    if color_scheme == "soft":
        dim_face = "#D7EEF6"
        fact_face = "#FFF2CC"
        stroke = "#2b2b2b"
    else:
        dim_face = "#E3F5E1"
        fact_face = "#FFE7AA"
        stroke = "#333333"

    # layout
    center = (0.0, 0.0)
    radius_outer = 3.8
    radius_inner = 2.4
    dims = dim_nodes or []

    # split into up to two rings if many dims
    n = len(dims)
    ring1 = min(n, max(0, two_rings))
    first_ring = dims[:ring1]
    second_ring = dims[ring1:]

    def place_ring(nodes, r, angle_offset=0.0):
        k = len(nodes)
        if k == 0:
            return []
        pts = []
        for i, (lab, _role) in enumerate(nodes):
            ang = angle_offset + 2 * math.pi * (i / k)
            pts.append((lab, (center[0] + r*math.cos(ang), center[1] + r*math.sin(ang))))
        return pts

    pts1 = place_ring(first_ring, radius_outer, angle_offset=0.0)
    pts2 = place_ring(second_ring, radius_inner, angle_offset=math.pi/len(second_ring) if second_ring else 0.0)

    # draw center fact
    face = FancyBboxPatch((center[0]-0.9, center[1]-0.35), 1.8, 0.7,
                          boxstyle="round,pad=0.05,rounding_size=0.12",
                          linewidth=2, edgecolor=stroke, facecolor=fact_face)
    ax.add_patch(face)
    flabel = wrap_label(fact_label, 16) if wrap_labels else fact_label
    ax.text(center[0], center[1], flabel, ha='center', va='center', fontsize=12*font_scale, fontweight='bold')

    def draw_dim(lbl, pos):
        # rounded pill
        w, h = 1.6, 0.7
        node = FancyBboxPatch((pos[0]-w/2, pos[1]-h/2), w, h,
                              boxstyle="round,pad=0.04,rounding_size=0.16",
                              linewidth=2, edgecolor=stroke, facecolor=dim_face)
        ax.add_patch(node)
        t = wrap_label(lbl, 14) if wrap_labels else lbl
        ax.text(pos[0], pos[1], t, ha='center', va='center', fontsize=11*font_scale, fontweight='bold')

    # draw rings
    points = pts1 + pts2
    for di, (dlabel, (x, y)) in enumerate(points):
        draw_dim(dlabel, (x, y))
        # connector line
        ax.plot([center[0], x], [center[1], y], color="#888888", linewidth=2, alpha=0.7)
        # FK label if provided
        if di < len(fk_labels) and fk_labels[di]:
            mx, my = (center[0]+x)/2, (center[1]+y)/2
            ax.text(mx, my, wrap_label(fk_labels[di], 18) if wrap_labels else fk_labels[di],
                    ha='center', va='center', fontsize=9*font_scale, color="#666666")

    # fit view
    ax.set_xlim(-5.5, 5.5)
    ax.set_ylim(-4.5, 4.5)
    fig.tight_layout()
    fig.savefig(str(out_png), dpi=150)
    plt.close(fig)


# ============ Graphviz lineage renderer ============

def graphviz_available() -> bool:
    if not _HAS_GRAPHVIZ:
        return False
    try:
        # probe DOT
        graphviz.backend.viewing.rendering.DEFAULT_ENGINE
        return True
    except Exception:
        return False


def render_lineage_graphviz(out_png: Path,
                            mart_display: str,
                            tiers: Dict[str, List[str]],
                            edges: List[Tuple[str, str]],
                            alias_to_display: Dict[str, str]) -> bool:
    """Draw tiered lineage with Graphviz. Returns True if a PNG was written."""
    if not graphviz_available():
        return False

    g = graphviz.Digraph(format="png", graph_attr=dict(splines="spline", rankdir="LR",
                                                       fontsize="12", fontname="Segoe UI"))
    g.attr("node", shape="box", style="rounded,filled", fontname="Segoe UI", fontsize="10")

    face = {"base": "#F0F0F0", "stage": "#EAF3FE", "mart": "#FDF5D3"}

    # subgraphs by rank
    def add_tier(tier_name: str, rank: str = 'same'):
        with g.subgraph(name=f'cluster_{tier_name}') as c:
            c.attr(rank=rank, color='white')  # invisible cluster border
            for a in tiers.get(tier_name, []):
                label = alias_to_display.get(a, a)
                c.node(a, label=label, fillcolor=face[tier_name])

    add_tier("base", "same")
    add_tier("stage", "same")
    add_tier("mart", "same")

    # edges
    for u, v in edges:
        g.edge(u, v, color="#888888")

    # title
    g.attr(label=f"Lineage: {mart_display}", labelloc="t", fontsize="14", fontname="Segoe UI")

    try:
        out_png.parent.mkdir(parents=True, exist_ok=True)
        g.render(filename=str(out_png.with_suffix("")), cleanup=True)
        return out_png.exists()
    except Exception:
        return False


# ============ Workbook builders ============

def build_workbook(args):
    manifest = load_manifest(Path(args.manifest))
    catalog  = load_catalog(Path(args.catalog)) if args.catalog else {}
    schemas  = load_schemas_yml(args.schemas) if args.schemas else {}

    nodes = manifest.get('nodes', {})
    # Build model registry
    models = []
    alias_to_node = {}
    for unique_id, n in nodes.items():
        if n.get('resource_type') != 'model':
            continue
        name = n.get('name')
        alias = n.get('alias') or name
        alias_lc = (alias or '').lower()
        alias_to_node[alias_lc] = n

        config = n.get('config') or {}
        mat = (config.get('materialized') or '').lower()
        tags = n.get('tags') or []
        kind = classify_kind(alias or name, tags, mat)

        rel = ".".join([n.get('database') or '', n.get('schema') or '', (n.get('relation_name') or alias or name)]).strip('.')
        models.append({
            'unique_id': unique_id,
            'name': name,
            'alias': alias,
            'alias_lc': alias_lc,
            'path': n.get('path') or '',
            'fqn': ".".join(n.get('fqn') or []),
            'database': n.get('database') or '',
            'schema': n.get('schema') or '',
            'relation': rel,
            'tags': tags,
            'materialized': mat,
            'kind': kind,
            'description': n.get('description') or ''
        })

    # Map for display casing
    alias_to_display = {m['alias_lc']: m['alias'] for m in models}
    model_kinds = {m['alias_lc']: m['kind'] for m in models}

    # Load schemas meta
    schema_meta = {}   # alias_lc -> (table_meta, columns_list)
    for m in models:
        entry = schemas.get(m['alias_lc']) or schemas.get(m['name'].lower())
        tmeta, cols = read_schema_columns(entry)
        schema_meta[m['alias_lc']] = (tmeta, cols)

    # Extract relationship tests (for Relationships + Star Map)
    relationships_rows = []  # FromModel, FromColumn, ToModel, ToColumn, TestName
    star_map_rows = []       # FactModel, DimensionModel, FactFKColumn, DimKeyColumn
    for m in models:
        alias_lc = m['alias_lc']
        display = alias_to_display.get(alias_lc, m['alias'])
        _tmeta, cols = schema_meta.get(alias_lc, ({}, []))
        for c in cols:
            for rel in c.get('tests') or []:
                to_model = alias_to_display.get((rel.get('to_model') or '').lower(), rel.get('to_model'))
                to_col   = rel.get('to_column') or ''
                relationships_rows.append({
                    'FromModel': display,
                    'FromColumn': c['name'],
                    'ToModel': pascalize(to_model or ''),
                    'ToColumn': to_col,
                    'TestName': 'relationships'
                })
                # if looks like fact->dim, add to star map
                if model_kinds.get(alias_lc) == 'fact':
                    star_map_rows.append({
                        'FactModel': display,
                        'DimensionModel': pascalize(to_model or ''),
                        'FactFKColumn': c['name'],
                        'DimKeyColumn': to_col
                    })

    # Build Pandas frames (useful internally)
    df_rels = pd.DataFrame(relationships_rows)
    df_star = pd.DataFrame(star_map_rows).drop_duplicates()

    # Exclusion logic for MODEL SHEETS (by prefix/tags/materializations/path globs)
    exc_prefixes = [p.strip() for p in (args.exclude_sheet_prefixes or "").split(",") if p.strip()]
    exc_tags     = {t.strip().lower() for t in (args.exclude_sheet_tags or "").split(",") if t.strip()}
    exc_mats     = {m.strip().lower() for m in (args.exclude_sheet_materializations or "").split(",") if m.strip()}
    exc_globs    = [g.strip() for g in (args.exclude_sheet_path_globs or "").split(",") if g.strip()]

    def excluded_sheet(m):
        a = m['alias'].lower()
        if any(a.startswith(p.lower()) for p in exc_prefixes):
            return True
        if exc_tags and (set(t.lower() for t in m['tags']) & exc_tags):
            return True
        if exc_mats and (m['materialized'] in exc_mats):
            return True
        if exc_globs and is_glob_matched(m['path'], exc_globs):
            return True
        # default: include
        return False

    # Workbook build
    wb = Workbook()
    used_names = set()
    # Remove default 'Sheet'
    del wb[wb.sheetnames[0]]

    # Summary sheet
    ws_sum = wb.create_sheet(safe_sheet_name("Summary", used_names))
    ws_sum.append(["Model", "Kind", "Materialization", "Relation", "Path", "FQN", "Tags", "Description"])
    for m in models:
        ws_sum.append([
            alias_to_display[m['alias_lc']],
            m['kind'], m['materialized'], m['relation'], m['path'], m['fqn'],
            ",".join(m['tags']), m['description']
        ])
    style_table(ws_sum, 1,
                width_map={"A": 36, "B": 14, "C": 18, "D": 44, "E": 32, "F": 48, "G": 24, "H": 64},
                wrap_cols=list("DH"))

    # Relationships sheet
    ws_rel = wb.create_sheet(safe_sheet_name("Relationships", used_names))
    ws_rel.append(["FromModel", "FromColumn", "ToModel", "ToColumn", "TestName"])
    if not df_rels.empty:
        for _, r in df_rels.iterrows():
            ws_rel.append([r['FromModel'], r['FromColumn'], r['ToModel'], r['ToColumn'], r['TestName']])
    style_table(ws_rel, 1,
                width_map={"A": 36, "B": 36, "C": 36, "D": 28, "E": 18},
                wrap_cols=list("ABCD"))

    # Star Map sheet (with FK columns)
    ws_star = wb.create_sheet(safe_sheet_name("Star Map", used_names))
    ws_star.append(["FactModel", "DimensionModel", "FactFKColumn", "DimKeyColumn"])
    if not df_star.empty:
        for _, r in df_star.iterrows():
            ws_star.append([r['FactModel'], r['DimensionModel'], r['FactFKColumn'], r['DimKeyColumn']])
    style_table(ws_star, 1,
                width_map={"A": 36, "B": 36, "C": 32, "D": 28},
                wrap_cols=list("ABCD"))

    # Model sheets
    for m in models:
        if excluded_sheet(m):
            continue
        alias_lc = m['alias_lc']
        display = alias_to_display[alias_lc]
        sheet_name = safe_sheet_name(display, used_names)
        ws = wb.create_sheet(sheet_name)

        # Table meta (without leading '#')
        tmeta, cols = schema_meta.get(alias_lc, ({}, []))
        # Order: Model, Kind, Materialization, Relation, Tags, SurrogateKey, BusinessKey, PrimaryKey, Description
        meta_rows = [
            ["Model", display],
            ["Kind", m['kind'].capitalize()],
            ["Materialization", m['materialized']],
            ["Relation", m['relation']],
            ["Tags", ",".join(m['tags'])],
            ["SurrogateKey", tmeta.get('surrogate_key', '')],
            ["BusinessKey", tmeta.get('business_key', '')],
            ["PrimaryKey", tmeta.get('primary_key', tmeta.get('surrogate_key',''))],
            ["Description", tmeta.get('description', '')]
        ]
        for row in meta_rows:
            ws.append(row)

        # Column header
        ws.append(["Column", "DataType", "Nullable?", "Description", "Tests", "IsPK?", "IsFK?", "Source"])
        hdr = ws.max_row

        # Column rows
        for c in cols:
            tests_text = ""
            if c.get('tests'):
                tests_text = "; ".join([f"relationships to={t['to_model']} field={t['to_column']}" for t in c['tests']])
            ws.append([
                c.get('name'), c.get('data_type'), c.get('nullable'),
                c.get('description'), tests_text,
                c.get('is_pk'), c.get('is_fk'), c.get('source')
            ])

        # Style / layout: freeze rows above header; wrap Description (D) and Tests (E)
        style_table(ws, hdr, width_map={"A": 32, "B": 22, "C": 10, "D": 68, "E": 48, "F": 10, "G": 10, "H": 10},
                    wrap_cols=list("DE"))

    # Star diagrams (Matplotlib)
    if args.star_diagrams:
        # cluster by fact
        facts = sorted(set(df_star['FactModel'])) if not df_star.empty else []
        for fact in facts:
            pick = df_star[df_star['FactModel'] == fact]
            dims = pick['DimensionModel'].tolist()
            fks  = pick['FactFKColumn'].tolist()
            # dedupe role players by label
            seen = set()
            dim_nodes = []
            fk_labels = []
            for d, fk in zip(dims, fks):
                key = d.lower()
                if key in seen:
                    continue
                seen.add(key)
                dim_nodes.append((d, None))
                fk_labels.append(fk)
            # draw
            with tempfile.TemporaryDirectory() as td:
                out_png = Path(td) / f"star_{fact}.png"
                draw_star_png(out_png, fact, dim_nodes, fk_labels,
                              two_rings=args.diagram_two_rings,
                              font_scale=args.diagram_font_scale,
                              color_scheme=args.diagram_color_scheme,
                              wrap_labels=not args.diagram_no_wrap_labels)
                # embed
                tab = safe_sheet_name(f"Star-Fact {fact}", used_names).replace('_', ' ')
                ws_img = wb.create_sheet(tab)
                img = XLImage(str(out_png))
                ws_img.add_image(img, "A1")
                ws_img.sheet_properties.tabColor = "92D050"  # soft green

    # Lineage text + (optional) Matplotlib + Graphviz
    if args.lineage:
        # Build upstream dependencies from manifest
        parent_map = manifest.get('parent_map') or {}
        id_to_alias = {}
        for m in models:
            id_to_alias[m['unique_id']] = m['alias_lc']

        # tiers
        alias_to_tier = {}
        for m in models:
            alias_to_tier[m['alias_lc']] = ('mart' if m['kind'] in {'fact', 'dimension'} else
                                            ('stage' if m['kind'] == 'stage' else
                                             ('base' if m['kind'] == 'base' else 'mart' if m['kind'] in {'view','table','incremental'} else 'mart')))

        # lineage rows
        ws_lin = wb.create_sheet(safe_sheet_name("Lineage", used_names))
        ws_lin.append(["MartModel", "UpstreamModel", "UpstreamTier"])
        for m in models:
            if alias_to_tier[m['alias_lc']] != 'mart':
                continue
            parents = []
            for pid in (parent_map.get(m['unique_id']) or []):
                al = id_to_alias.get(pid)
                if al:
                    parents.append(al)
            # dedupe + fill
            for p in sorted(set(parents)):
                ws_lin.append([alias_to_display[m['alias_lc']], alias_to_display.get(p, p), alias_to_tier.get(p, '')])
        style_table(ws_lin, 1, width_map={"A": 40, "B": 40, "C": 14}, wrap_cols=list("ABC"))

        # Matplotlib lineage diagrams (optional)
        if args.lineage_diagrams:
            # create a per-mart diagram with tiers
            mart_models = sorted({alias_to_display[m['alias_lc']] for m in models if alias_to_tier[m['alias_lc']] == 'mart'})
            for mart_disp in mart_models:
                mart_alias_lc = mart_disp.lower()
                # Collect parents recursively for tiers
                # For simplicity, put any non-mart parent into base/stage by alias_to_tier
                tiers = {"base": [], "stage": [], "mart": []}
                edges = []
                # Resolve upstream from Lineage sheet rows we just wrote
                rows = []
                for r in range(2, ws_lin.max_row+1):
                    if (ws_lin.cell(r, 1).value or '').strip() == mart_disp:
                        rows.append((ws_lin.cell(r, 2).value or '').strip())
                for up in sorted(set(rows)):
                    up_lc = up.lower()
                    t = alias_to_tier.get(up_lc, 'mart')
                    tiers.setdefault(t, [])
                    tiers[t].append(up_lc)
                    edges.append((up_lc, mart_alias_lc))
                # also register mart node itself
                tiers['mart'].append(mart_alias_lc)
                # draw via Graphviz if requested
                if args.gv_lineage and graphviz_available():
                    with tempfile.TemporaryDirectory() as td:
                        out_png = Path(td) / f"lineage_{mart_disp}.png"
                        ok = render_lineage_graphviz(out_png, mart_disp, tiers, edges, alias_to_display)
                        if ok:
                            tab = safe_sheet_name(f"Lineage-GV {mart_disp}", used_names)
                            ws_img = wb.create_sheet(tab)
                            img = XLImage(str(out_png))
                            ws_img.add_image(img, "A1")
                else:
                    # simple Matplotlib block diagram (fallback)
                    with tempfile.TemporaryDirectory() as td:
                        out_png = Path(td) / f"lineage_{mart_disp}.png"
                        fig = plt.figure(figsize=(10, 4), dpi=150)
                        ax = plt.gca(); ax.axis('off')
                        x = 0.5
                        y = {'base': 0.75, 'stage': 0.5, 'mart': 0.25}
                        for tier, arr in tiers.items():
                            for i, a in enumerate(sorted(set(arr))):
                                lbl = pascalize(alias_to_display.get(a, a))
                                rect = FancyBboxPatch((x+i*2.2, y[tier]), 1.8, 0.3,
                                                      boxstyle="round,pad=0.02,rounding_size=0.08",
                                                      linewidth=1.5, edgecolor="#333333",
                                                      facecolor=("#F0F0F0" if tier=='base' else "#EAF3FE" if tier=='stage' else "#FFF2CC"))
                                ax.add_patch(rect)
                                ax.text(x+i*2.2+0.9, y[tier]+0.15, wrap_label(lbl, 18), ha='center', va='center', fontsize=9)
                        for u, v in edges:
                            # naive: draw a line from tier y of u to mart
                            yu = y[alias_to_tier.get(u, 'mart')]
                            xv = x
                            ax.plot([x+0.9, xv+0.9], [yu+0.15, y['mart']+0.15], color="#888888")
                        fig.tight_layout()
                        fig.savefig(str(out_png))
                        plt.close(fig)
                        tab = safe_sheet_name(f"Lineage-{mart_disp}", used_names)
                        ws_img = wb.create_sheet(tab)
                        img = XLImage(str(out_png))
                        ws_img.add_image(img, "A1")

        # Graphviz lineage using robust mart detection if requested and not already done above
        if args.gv_lineage and graphviz_available():
            # Robust mart list
            mart_list = set()
            # 1) Star Map facts
            if 'Star Map' in wb.sheetnames:
                try:
                    wss = wb['Star Map']
                    for r in range(2, wss.max_row+1):
                        fac = (wss.cell(r, 1).value or '').strip()
                        if fac:
                            mart_list.add(fac)
                except Exception:
                    pass
            # 2) tier map
            for alias, tier in alias_to_tier.items():
                if tier == 'mart':
                    mart_list.add(alias_to_display.get(alias, alias))
            # 3) model kinds (facts & dimensions are mart layer)
            for alias, kind in model_kinds.items():
                if kind in {'fact', 'dimension'}:
                    mart_list.add(alias_to_display.get(alias, alias))
            # 4) tags 'mart'
            for alias, node in alias_to_node.items():
                if 'mart' in {t.lower() for t in (node.get('tags') or [])}:
                    mart_list.add(alias_to_display.get(alias, alias))
            mart_list = {m for m in (m.strip() for m in mart_list) if m}
            if not mart_list:
                print("NOTE: gv-lineage: no mart models found; nothing to draw.", file=sys.stderr)
            else:
                print(f"gv-lineage: will render {len(mart_list)} mart models", file=sys.stderr)
                for mart_disp in sorted(mart_list):
                    print(f"gv-lineage: rendering {mart_disp}", file=sys.stderr)
                    mart_alias_lc = mart_disp.lower()
                    tiers = {"base": [], "stage": [], "mart": [mart_alias_lc]}
                    edges = []
                    # use Relationships sheet and/or parent_map to infer upstream
                    ups = set()
                    # from Relationships (dims are upstream to fact)
                    if 'Relationships' in wb.sheetnames:
                        wr = wb['Relationships']
                        for r in range(2, wr.max_row+1):
                            if (wr.cell(r, 1).value or '').strip() == mart_disp:
                                ups.add((wr.cell(r, 3).value or '').strip().lower())
                    # parent_map
                    for pid in (parent_map.get(next((m['unique_id'] for m in models if m['alias'].lower()==mart_alias_lc), ''), []) or []):
                        al = id_to_alias.get(pid)
                        if al:
                            ups.add(al)
                    for up in sorted(u for u in ups if u):
                        t = alias_to_tier.get(up, 'mart')
                        tiers.setdefault(t, [])
                        tiers[t].append(up)
                        edges.append((up, mart_alias_lc))
                    # draw
                    with tempfile.TemporaryDirectory() as td:
                        out_png = Path(td) / f"lineage_gv_{mart_disp}.png"
                        ok = render_lineage_graphviz(out_png, mart_disp, tiers, edges, alias_to_display)
                        if ok:
                            tab = safe_sheet_name(f"Lineage-GV {mart_disp}", used_names)
                            ws_img = wb.create_sheet(tab)
                            img = XLImage(str(out_png))
                            ws_img.add_image(img, "A1")

    # Save workbook
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"Wrote: {out_path}")


# ============ CLI ============

def main():
    p = argparse.ArgumentParser(description="Generate a Kimball-style Excel workbook from dbt artifacts.")
    p.add_argument("--manifest", required=True, help="Path to target/manifest.json")
    p.add_argument("--catalog",  required=False, default="", help="Path to target/catalog.json")
    p.add_argument("--schemas",  required=False, default="", help="Comma-separated paths to schemas.yml files")

    # Sheet inclusion/exclusion knobs
    p.add_argument("--include-views", action="store_true", help="Include view materializations in sheets/summary")
    p.add_argument("--materializations", default="", help="Only include models with these materializations (comma-list)")
    p.add_argument("--exclude-sheet-prefixes", default="", help="Comma-list of model name prefixes to exclude from individual model sheets")
    p.add_argument("--exclude-sheet-tags", default="", help="Comma-list of tags that exclude a sheet")
    p.add_argument("--exclude-sheet-materializations", default="", help="Comma-list of materializations to exclude from sheets")
    p.add_argument("--exclude-sheet-path-globs", default="", help="Comma-list of globs (e.g. models/base/*) to exclude from sheets")

    # Star diagrams
    p.add_argument("--star-diagrams", action="store_true", help="Render a star diagram tab per fact")
    p.add_argument("--diagram-shape", choices=["roundrect"], default="roundrect")
    p.add_argument("--diagram-dedupe-roles", action="store_true", help="Deduplicate role-playing dimensions by label")
    p.add_argument("--diagram-two-rings", type=int, default=12, help="Max nodes in outer ring before using inner ring")
    p.add_argument("--diagram-font-scale", type=float, default=0.9)
    p.add_argument("--diagram-dpi", type=int, default=240)
    p.add_argument("--diagram-color-scheme", choices=["soft", "mint"], default="soft")
    p.add_argument("--diagram-legend", action="store_true")
    p.add_argument("--diagram-no-wrap-labels", action="store_true", help="Do not wrap node labels")

    # Lineage
    p.add_argument("--lineage", action="store_true", help="Emit textual lineage table")
    p.add_argument("--lineage-diagrams", action="store_true", help="Emit per-mart lineage diagrams (Matplotlib fallback)")
    p.add_argument("--gv-lineage", action="store_true", help="Use Graphviz for lineage diagrams when available")
    p.add_argument("--gv-format", default="png", help="Graphviz output image format (png recommended)")

    p.add_argument("--out", required=True, help="Output XLSX path")
    args = p.parse_args()

    build_workbook(args)


if __name__ == "__main__":
    main()
