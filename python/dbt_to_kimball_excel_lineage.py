#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
dbt_to_kimball_excel.py
-----------------------

Generate a Kimball-style Excel workbook from a dbt Core project.

What it does
============
1) Reads dbt's `manifest.json` + `catalog.json` (+ optional `schemas.yml` files)
2) Produces a workbook with:
   • Summary tab (Model, Kind, Materialization, Relation, Path, FQN, Tags, Description)
   • One tab per model:
       - Table meta first (Model, Kind, Materialization, Relation, Tags, SurrogateKey,
         BusinessKey, PrimaryKey, Description)
       - Then a header row for column metadata followed by all columns
       - Formatting: frozen header, blue header fill, bold header text, auto-filter,
         green-bar alternating rows, wrapped Description & Tests
   • Relationships tab (FromModel, FromColumn, ToModel, ToColumn, TestName)
   • Star Map tab (FactModel, DimensionModel, FactFKColumn, DimKeyColumn)
   • Optional star diagrams (one tab per fact) with rounded shapes & FK edge labels
   • Optional Lineage tab (textual base→stage→mart)
   • Optional lineage diagrams per mart (Matplotlib fallback and/or Graphviz if available)

Why both “Relationships” and “Star Map”?
=======================================
- Relationships: a raw, row-per-test view of column-level relationship tests from schemas.yml.
  It’s exhaustive, including non fact→dim relationships.
- Star Map: distilled fact→dimension pairs + the exact FK column used on the fact
  (pulled from Relationships). This is used to render star diagrams and to seed
  the “what is a mart?” set for lineage.

dbt schema compatibility
========================
- Supports both `tests:` and `data_tests:` keys under columns in schemas.yml.
- Relationship test shapes accepted:
    relationships:
      to: ref('DimUser')    # or 'DimUser', or source(...)
      field: UserSKey
      # optionally: column/to_column
- Table-level schema extras observed when present:
    surrogate_key, business_key, primary_key, description, tags.

Heuristics
==========
- Model Kind:
    • fact: tags include 'fact' OR name/alias startswith 'fact'
    • dimension: tags include 'dim' OR name/alias startswith 'dim'
    • stage/base by tag or prefix; else defer to materialization where sensible
- “Mart layer” for diagrams/lineage includes facts & dimensions (and, if needed, tables/views).

Graphviz
========
- If Graphviz binaries are installed (verify with `dot -V`) and the `graphviz` python
  package is present, `--gv-lineage` will embed Graphviz-rendered lineage diagrams.
- If Graphviz is unavailable, Matplotlib fallback is used (when `--lineage-diagrams` is set).

Install
=======
pip install:
  pyyaml pandas openpyxl matplotlib networkx graphviz pillow

(For Graphviz diagrams, install system Graphviz so `dot -V` works.)

Examples
========
# Minimal:
python scripts/dbt_to_kimball_excel.py --manifest target/manifest.json --catalog target/catalog.json \
  --out documentation/kimball_model_catalog.xlsx

# With schemas and exclusions + diagrams + lineage:
python scripts/dbt_to_kimball_excel.py \
  --manifest target/manifest.json --catalog target/catalog.json \
  --schemas "models/mart/schemas.yml,models/stage/schemas.yml" \
  --exclude-sheet-prefixes "base_,stage_" \
  --exclude-sheet-materializations "view" \
  --exclude-sheet-tags "stage,logging" \
  --exclude-sheet-path-globs "models/base/*,models/stage/*,models/logging/*" \
  --star-diagrams --diagram-two-rings 12 --diagram-font-scale 0.9 \
  --diagram-color-scheme soft --diagram-legend \
  --lineage --lineage-diagrams --gv-lineage --gv-format png \
  --out documentation/kimball_model_catalog.xlsx

Recent fixes (kept working behavior)
====================================
- Fixed `safe_sheet_name()` typo; safe de-dup with case-insensitive check.
- Replaced deprecated `.column_letter` usage with `get_column_letter()` and
  `column_index_from_string()` in styling helpers.
- Star & lineage images now embedded **from memory** (Pillow `Image`) to avoid
  temp file lifetime issues during workbook save.
- More robust mart detection for lineage diagrams (union of Star Map facts,
  tier map, kind in {fact, dimension}, and tag 'mart').
- Removed guards that skipped lineage diagrams if only mart→mart edges exist.

"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import tempfile
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

# Matplotlib (headless)
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch

# Optional Graphviz
try:
    import graphviz
    _HAS_GRAPHVIZ = True
except Exception:
    _HAS_GRAPHVIZ = False


# ------------------------
# Utility helpers
# ------------------------

def pascalize(name: str) -> str:
    if not name:
        return name
    # preserve mixed-case words
    if re.search(r'[a-z][A-Z]', name):
        return name
    if '_' in name:
        return ''.join(p.capitalize() for p in name.split('_') if p)
    return name[:1].upper() + name[1:]


def split_camel(label: str) -> List[str]:
    if not label:
        return [label]
    parts = re.findall(r'[A-Z]?[a-z0-9]+|[A-Z]+(?![a-z])', label)
    return parts or [label]


def wrap_label(label: str, max_chars: int = 14) -> str:
    words = split_camel(label)
    lines, cur = [], ""
    for w in words:
        if not cur:
            cur = w
        elif len(cur) + 1 + len(w) <= max_chars:
            cur += " " + w
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return "\n".join(lines)


def safe_sheet_name(base: str, used: set) -> str:
    """Excel sheet names are <=31 chars and cannot include :\/?*[]"""
    raw = (base or 'Sheet').strip()
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', raw)[:31] or 'Sheet'
    used_lower = {u.lower() for u in used}
    candidate = cleaned
    i = 1
    while candidate.lower() in used_lower:
        suffix = f'_{i}'
        candidate = (cleaned[:31 - len(suffix)] + suffix)
        i += 1
    used.add(candidate)
    return candidate


def is_glob_matched(path: str, patterns: List[str]) -> bool:
    from fnmatch import fnmatch
    norm = path.replace("\\", "/")
    return any(fnmatch(norm, p) for p in patterns)


# ------------------------
# Load dbt artifacts
# ------------------------

def load_manifest(p: Path) -> Dict[str, Any]:
    return json.loads(p.read_text(encoding="utf-8"))


def load_catalog(p: Optional[Path]) -> Dict[str, Any]:
    if not p or not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def load_schemas_yml(paths_csv: str) -> Dict[str, Any]:
    """Merge models from multiple schemas.yml files keyed by lowercased model name."""
    out = {}
    if not paths_csv:
        return out
    for raw in [x.strip() for x in paths_csv.split(',') if x.strip()]:
        p = Path(raw)
        if not p.exists():
            continue
        doc = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
        for m in (doc.get('models') or []):
            name = (m.get('name') or '').strip()
            if name:
                out[name.lower()] = m
    return out


# ------------------------
# Schema + kind helpers
# ------------------------

def classify_kind(name_or_alias: str, tags: List[str], mat: str) -> str:
    lo = (name_or_alias or '').lower()
    tl = {t.lower() for t in (tags or [])}
    if 'fact' in tl or lo.startswith('fact'):
        return 'fact'
    if 'dim' in tl or lo.startswith('dim'):
        return 'dimension'
    if 'stage' in tl or lo.startswith('stage'):
        return 'stage'
    if 'base' in tl or lo.startswith('base'):
        return 'base'
    if mat in ('view', 'table', 'incremental', 'materializedview'):
        return mat
    return 'unknown'


def extract_relationship_tests(col_entry: Dict[str, Any]) -> List[Dict[str, str]]:
    out = []
    tkey = 'tests' if 'tests' in col_entry else 'data_tests' if 'data_tests' in col_entry else None
    if not tkey:
        return out
    for t in (col_entry.get(tkey) or []):
        if isinstance(t, dict) and 'relationships' in t:
            rel = t['relationships'] or {}
            to_val = rel.get('to') or rel.get('to_model') or ''
            to_model = str(to_val)
            # normalize ref('X') or source('...') -> X (best-effort)
            m = re.search(r"ref\(['\"]([^'\"]+)['\"]\)", to_model)
            if m:
                to_model = m.group(1)
            else:
                to_model = re.sub(r'[^A-Za-z0-9_]', '', to_model)
            field = rel.get('field') or rel.get('to_field') or rel.get('to_column') or ''
            to_col = rel.get('column') or rel.get('to_column') or field
            out.append({'to_model': to_model, 'to_column': str(to_col or '')})
    return out


def read_schema_columns(schema_m: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    tmeta: Dict[str, Any] = {}
    cols: List[Dict[str, Any]] = []
    if not schema_m:
        return tmeta, cols

    # table-level meta
    for k in ('description', 'surrogate_key', 'business_key', 'primary_key'):
        if k in schema_m:
            tmeta[k] = schema_m.get(k)
    tmeta['tags'] = ','.join(schema_m.get('tags') or [])

    # columns
    for c in (schema_m.get('columns') or []):
        meta = c.get('meta') or {}
        nullable = meta.get('nullable')
        if nullable is None:
            nullable_flag = ''  # unknown
        else:
            nullable_flag = 'Y' if bool(nullable) else 'N'

        item = {
            'name': c.get('name'),
            'description': c.get('description') or '',
            'data_type': c.get('data_type') or '',
            'nullable': nullable_flag,
            'tests': [],
            'is_pk': 'Y' if (meta.get('identity') or meta.get('is_primary_key')) else '',
            'is_fk': '',
            'source': 'YAML',
        }
        rels = extract_relationship_tests(c)
        if rels:
            item['is_fk'] = 'Y'
        item['tests'] = rels
        cols.append(item)
    return tmeta, cols


# ------------------------
# Excel styling
# ------------------------

BLUE = "DDEBF7"
ALT_GREEN = "F2F9ED"

header_fill = PatternFill("solid", fgColor=BLUE)
header_font = Font(bold=True, color="000000")

def style_table(ws: Worksheet, top_header_row: int, width_map: Dict[str, float], wrap_cols: List[str], freeze=True):
    if freeze:
        ws.freeze_panes = f"A{top_header_row+1}"

    max_col = ws.max_column

    # header styling
    for col_idx in range(1, max_col+1):
        c = ws.cell(row=top_header_row, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    # filter range
    last_col_letter = get_column_letter(max_col)
    ws.auto_filter.ref = f"A{top_header_row}:{last_col_letter}{ws.max_row}"

    # alternating green-bar rows, thin borders
    thin = Side(style="thin", color="DDDDDD")
    for r in range(top_header_row+1, ws.max_row+1):
        filler = PatternFill("solid", fgColor=(ALT_GREEN if r % 2 == 0 else "FFFFFF"))
        for col_idx in range(1, max_col+1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = filler
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # wrap selected columns
    for r in range(top_header_row+1, ws.max_row+1):
        for letter in wrap_cols:
            ws.cell(row=r, column=column_index_from_string(letter)).alignment = Alignment(wrap_text=True, vertical="top")

    # widths
    for letter, width in width_map.items():
        ws.column_dimensions[letter].width = width


# ------------------------
# Star diagram (Matplotlib) -> Pillow Image (memory-safe)
# ------------------------

def draw_star_image(fact_label: str,
                    dim_nodes: List[Tuple[str, Optional[str]]],
                    fk_labels: List[str],
                    two_rings: int = 12,
                    font_scale: float = 1.0,
                    color_scheme: str = "soft",
                    wrap_labels: bool = True) -> "PIL.Image.Image":
    W, H = 1200, 900
    fig = plt.figure(figsize=(W/100, H/100), dpi=100)
    ax = plt.gca()
    ax.set_aspect('equal')
    ax.axis('off')

    if color_scheme == "soft":
        dim_face = "#D7EEF6"
        fact_face = "#FFF2CC"
        stroke = "#2b2b2b"
    else:
        dim_face = "#E3F5E1"
        fact_face = "#FFE7AA"
        stroke = "#333333"

    center = (0.0, 0.0)
    radius_outer = 3.8
    radius_inner = 2.4
    dims = dim_nodes or []

    n = len(dims)
    ring1 = min(n, max(0, two_rings))
    first_ring = dims[:ring1]
    second_ring = dims[ring1:]

    def place_ring(nodes, r, angle_offset=0.0):
        k = len(nodes)
        if k == 0:
            return []
        pts = []
        for i, (lab, _role) in enumerate(nodes):
            ang = angle_offset + 2 * math.pi * (i / k)
            pts.append((lab, (center[0] + r*math.cos(ang), center[1] + r*math.sin(ang))))
        return pts

    pts1 = place_ring(first_ring, radius_outer, angle_offset=0.0)
    pts2 = place_ring(second_ring, radius_inner, angle_offset=math.pi/len(second_ring) if second_ring else 0.0)

    # center fact
    face = FancyBboxPatch((center[0]-0.9, center[1]-0.35), 1.8, 0.7,
                          boxstyle="round,pad=0.05,rounding_size=0.12",
                          linewidth=2, edgecolor=stroke, facecolor=fact_face)
    ax.add_patch(face)
    flabel = wrap_label(fact_label, 16) if wrap_labels else fact_label
    ax.text(center[0], center[1], flabel, ha='center', va='center', fontsize=12*font_scale, fontweight='bold')

    def draw_dim(lbl, pos):
        w, h = 1.6, 0.7
        node = FancyBboxPatch((pos[0]-w/2, pos[1]-h/2), w, h,
                              boxstyle="round,pad=0.04,rounding_size=0.16",
                              linewidth=2, edgecolor=stroke, facecolor=dim_face)
        ax.add_patch(node)
        t = wrap_label(lbl, 14) if wrap_labels else lbl
        ax.text(pos[0], pos[1], t, ha='center', va='center', fontsize=11*font_scale, fontweight='bold')

    points = pts1 + pts2
    for di, (dlabel, (x, y)) in enumerate(points):
        draw_dim(dlabel, (x, y))
        ax.plot([center[0], x], [center[1], y], color="#888888", linewidth=2, alpha=0.7)
        if di < len(fk_labels) and fk_labels[di]:
            mx, my = (center[0]+x)/2, (center[1]+y)/2
            ax.text(mx, my, wrap_label(fk_labels[di], 18) if wrap_labels else fk_labels[di],
                    ha='center', va='center', fontsize=9*font_scale, color="#666666")

    ax.set_xlim(-5.5, 5.5)
    ax.set_ylim(-4.5, 4.5)
    try:
        fig.tight_layout()
    except Exception:
        pass

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return PILImage.open(buf)


# ------------------------
# Graphviz lineage -> Pillow Image (memory-safe)
# ------------------------

def graphviz_available() -> bool:
    if not _HAS_GRAPHVIZ:
        return False
    # Try building a trivial graph to ensure backend is usable
    try:
        graphviz.Digraph().source
        return True
    except Exception:
        return False


def render_lineage_graphviz_image(mart_display: str,
                                  tiers: Dict[str, List[str]],
                                  edges: List[Tuple[str, str]],
                                  alias_to_display: Dict[str, str]) -> Optional["PIL.Image.Image"]:
    if not graphviz_available():
        return None
    g = graphviz.Digraph(format="png", graph_attr=dict(splines="spline", rankdir="LR",
                                                       fontsize="12", fontname="Segoe UI"))
    g.attr("node", shape="box", style="rounded,filled", fontname="Segoe UI", fontsize="10")
    face = {"base": "#F0F0F0", "stage": "#EAF3FE", "mart": "#FFF2CC"}

    def add_tier(tier_name: str, rank: str = 'same'):
        with g.subgraph(name=f'cluster_{tier_name}') as c:
            c.attr(rank=rank, color='white')
            for a in tiers.get(tier_name, []):
                label = alias_to_display.get(a, a)
                c.node(a, label=label, fillcolor=face[tier_name])

    add_tier("base", "same")
    add_tier("stage", "same")
    add_tier("mart", "same")

    for u, v in edges:
        g.edge(u, v, color="#888888")

    g.attr(label=f"Lineage: {mart_display}", labelloc="t", fontsize="14", fontname="Segoe UI")

    with tempfile.TemporaryDirectory() as td:
        base = Path(td) / "graph"
        outfile = g.render(filename=str(base), cleanup=True)
        # Load to memory and return
        return PILImage.open(outfile)


# ------------------------
# Workbook builder
# ------------------------

def build_workbook(args):
    manifest = load_manifest(Path(args.manifest))
    catalog  = load_catalog(Path(args.catalog)) if args.catalog else {}
    schemas  = load_schemas_yml(args.schemas) if args.schemas else {}

    nodes = manifest.get('nodes', {})
    models: List[Dict[str, Any]] = []
    alias_to_node: Dict[str, Any] = {}

    for unique_id, n in nodes.items():
        if n.get('resource_type') != 'model':
            continue
        name = n.get('name')
        alias = n.get('alias') or name
        alias_lc = (alias or '').lower()
        alias_to_node[alias_lc] = n

        cfg = n.get('config') or {}
        mat = (cfg.get('materialized') or '').lower()
        tags = n.get('tags') or []
        kind = classify_kind(alias or name, tags, mat)

        relation = ".".join(filter(None, [n.get('database'), n.get('schema'), n.get('relation_name') or alias or name]))
        models.append({
            'unique_id': unique_id,
            'name': name,
            'alias': alias,
            'alias_lc': alias_lc,
            'path': n.get('path') or '',
            'fqn': ".".join(n.get('fqn') or []),
            'database': n.get('database') or '',
            'schema': n.get('schema') or '',
            'relation': relation,
            'tags': tags,
            'materialized': mat,
            'kind': kind,
            'description': n.get('description') or ''
        })

    alias_to_display = {m['alias_lc']: m['alias'] for m in models}
    model_kinds      = {m['alias_lc']: m['kind'] for m in models}

    # Merge schemas.yml metadata
    schema_meta: Dict[str, Tuple[Dict[str, Any], List[Dict[str, Any]]]] = {}
    for m in models:
        entry = schemas.get(m['alias_lc']) or schemas.get(m['name'].lower())
        tmeta, cols = read_schema_columns(entry)
        schema_meta[m['alias_lc']] = (tmeta, cols)

    # Build Relationships and Star Map rows
    relationships_rows = []
    star_map_rows = []
    for m in models:
        alias_lc = m['alias_lc']
        disp = alias_to_display.get(alias_lc, m['alias'])
        _tmeta, cols = schema_meta.get(alias_lc, ({}, []))
        for c in cols:
            for rel in (c.get('tests') or []):
                to_model = alias_to_display.get((rel.get('to_model') or '').lower(), rel.get('to_model'))
                to_col   = rel.get('to_column') or ''
                relationships_rows.append({
                    'FromModel': disp,
                    'FromColumn': c['name'],
                    'ToModel': pascalize(to_model or ''),
                    'ToColumn': to_col,
                    'TestName': 'relationships'
                })
                if model_kinds.get(alias_lc) == 'fact':
                    star_map_rows.append({
                        'FactModel': disp,
                        'DimensionModel': pascalize(to_model or ''),
                        'FactFKColumn': c['name'],
                        'DimKeyColumn': to_col
                    })

    df_rels = pd.DataFrame(relationships_rows)
    df_star = pd.DataFrame(star_map_rows).drop_duplicates()

    # Exclusion filters for individual model sheets
    exc_prefixes = [p.strip() for p in (args.exclude_sheet_prefixes or "").split(",") if p.strip()]
    exc_tags     = {t.strip().lower() for t in (args.exclude_sheet_tags or "").split(",") if t.strip()}
    exc_mats     = {m.strip().lower() for m in (args.exclude_sheet_materializations or "").split(",") if m.strip()}
    exc_globs    = [g.strip() for g in (args.exclude_sheet_path_globs or "").split(",") if g.strip()]

    def exclude_sheet(m):
        a = m['alias'].lower()
        if any(a.startswith(p.lower()) for p in exc_prefixes):
            return True
        if exc_tags and (set(t.lower() for t in m['tags']) & exc_tags):
            return True
        if exc_mats and (m['materialized'] in exc_mats):
            return True
        if exc_globs and is_glob_matched(m['path'], exc_globs):
            return True
        return False

    # Create workbook
    wb = Workbook()
    used_names = set()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    # Summary
    ws_sum = wb.create_sheet(safe_sheet_name("Summary", used_names))
    ws_sum.append(["Model", "Kind", "Materialization", "Relation", "Path", "FQN", "Tags", "Description"])
    for m in models:
        ws_sum.append([
            alias_to_display[m['alias_lc']],
            m['kind'], m['materialized'], m['relation'], m['path'], m['fqn'],
            ",".join(m['tags']), m['description']
        ])
    style_table(ws_sum, 1,
                width_map={"A": 36, "B": 14, "C": 18, "D": 44, "E": 32, "F": 48, "G": 24, "H": 64},
                wrap_cols=list("DH"))

    # Relationships
    ws_rel = wb.create_sheet(safe_sheet_name("Relationships", used_names))
    ws_rel.append(["FromModel", "FromColumn", "ToModel", "ToColumn", "TestName"])
    if not df_rels.empty:
        for _, r in df_rels.iterrows():
            ws_rel.append([r['FromModel'], r['FromColumn'], r['ToModel'], r['ToColumn'], r['TestName']])
    style_table(ws_rel, 1,
                width_map={"A": 36, "B": 36, "C": 36, "D": 28, "E": 18},
                wrap_cols=list("ABCD"))

    # Star Map
    ws_star = wb.create_sheet(safe_sheet_name("Star Map", used_names))
    ws_star.append(["FactModel", "DimensionModel", "FactFKColumn", "DimKeyColumn"])
    if not df_star.empty:
        for _, r in df_star.iterrows():
            ws_star.append([r['FactModel'], r['DimensionModel'], r['FactFKColumn'], r['DimKeyColumn']])
    style_table(ws_star, 1,
                width_map={"A": 36, "B": 36, "C": 32, "D": 28},
                wrap_cols=list("ABCD"))

    # Individual model sheets
    for m in models:
        if exclude_sheet(m):
            continue
        alias_lc = m['alias_lc']
        disp = alias_to_display[alias_lc]
        ws = wb.create_sheet(safe_sheet_name(disp, used_names))

        tmeta, cols = schema_meta.get(alias_lc, ({}, []))
        meta_rows = [
            ["Model", disp],
            ["Kind", m['kind'].capitalize()],
            ["Materialization", m['materialized']],
            ["Relation", m['relation']],
            ["Tags", ",".join(m['tags'])],
            ["SurrogateKey", tmeta.get('surrogate_key', '')],
            ["BusinessKey", tmeta.get('business_key', '')],
            ["PrimaryKey", tmeta.get('primary_key', tmeta.get('surrogate_key', ''))],
            ["Description", tmeta.get('description', '')],
        ]
        for row in meta_rows:
            ws.append(row)

        ws.append(["Column", "DataType", "Nullable?", "Description", "Tests", "IsPK?", "IsFK?", "Source"])
        hdr_row = ws.max_row

        for c in cols:
            tests_text = ""
            if c.get('tests'):
                tests_text = "; ".join([f"relationships to={t['to_model']} field={t['to_column']}" for t in c['tests']])
            ws.append([
                c.get('name'), c.get('data_type'), c.get('nullable'),
                c.get('description'), tests_text,
                c.get('is_pk'), c.get('is_fk'), c.get('source')
            ])

        style_table(ws, hdr_row,
                    width_map={"A": 32, "B": 22, "C": 10, "D": 68, "E": 48, "F": 10, "G": 10, "H": 10},
                    wrap_cols=list("DE"))

    # Star diagrams
    if args.star_diagrams:
        facts = sorted(set(df_star['FactModel'])) if not df_star.empty else []
        for fact in facts:
            pick = df_star[df_star['FactModel'] == fact]
            dims = pick['DimensionModel'].tolist()
            fks  = pick['FactFKColumn'].tolist()

            # dedupe role players
            seen = set()
            dim_nodes, fk_labels = [], []
            for d, fk in zip(dims, fks):
                key = d.lower()
                if key in seen:
                    continue
                seen.add(key)
                dim_nodes.append((d, None))
                fk_labels.append(fk)

            pil_img = draw_star_image(
                fact_label=fact,
                dim_nodes=dim_nodes,
                fk_labels=fk_labels,
                two_rings=args.diagram_two_rings,
                font_scale=args.diagram_font_scale,
                color_scheme=args.diagram_color_scheme,
                wrap_labels=not args.diagram_no_wrap_labels,
            )
            tab = safe_sheet_name(f"Star-Fact {fact}", used_names).replace('_', ' ')
            ws_img = wb.create_sheet(tab)
            ws_img.sheet_properties.tabColor = "92D050"
            ws_img.add_image(XLImage(pil_img), "A1")

    # Lineage (text + diagrams)
    if args.lineage:
        parent_map = manifest.get('parent_map') or {}
        id_to_alias = {m['unique_id']: m['alias_lc'] for m in models}

        # Basic tier map
        alias_to_tier = {}
        for m in models:
            if m['kind'] in {'fact', 'dimension'}:
                alias_to_tier[m['alias_lc']] = 'mart'
            elif m['kind'] == 'stage':
                alias_to_tier[m['alias_lc']] = 'stage'
            elif m['kind'] == 'base':
                alias_to_tier[m['alias_lc']] = 'base'
            else:
                alias_to_tier[m['alias_lc']] = 'mart'

        # Text sheet
        ws_lin = wb.create_sheet(safe_sheet_name("Lineage", used_names))
        ws_lin.append(["MartModel", "UpstreamModel", "UpstreamTier"])
        for m in models:
            if alias_to_tier[m['alias_lc']] != 'mart':
                continue
            parents = set()
            for pid in (parent_map.get(m['unique_id']) or []):
                al = id_to_alias.get(pid)
                if al:
                    parents.add(al)
            for p in sorted(parents):
                ws_lin.append([alias_to_display[m['alias_lc']],
                               alias_to_display.get(p, p),
                               alias_to_tier.get(p, '')])
        style_table(ws_lin, 1, width_map={"A": 40, "B": 40, "C": 14}, wrap_cols=list("ABC"))

        # Matplotlib fallback diagrams
        if args.lineage_diagrams and not (args.gv_lineage and graphviz_available()):
            mart_models = sorted({alias_to_display[m['alias_lc']]
                                  for m in models if alias_to_tier[m['alias_lc']] == 'mart'})
            for mart_disp in mart_models:
                mart_alias_lc = mart_disp.lower()
                tiers = {"base": [], "stage": [], "mart": [mart_alias_lc]}
                edges = []

                # From Relationships (dims upstream to fact)
                if 'Relationships' in wb.sheetnames:
                    wr = wb['Relationships']
                    for r in range(2, wr.max_row+1):
                        if (wr.cell(r, 1).value or '').strip() == mart_disp:
                            up = (wr.cell(r, 3).value or '').strip().lower()
                            if up:
                                tiers.setdefault(alias_to_tier.get(up, 'mart'), []).append(up)
                                edges.append((up, mart_alias_lc))

                # Simple block diagram
                fig = plt.figure(figsize=(10, 4), dpi=150)
                ax = plt.gca(); ax.axis('off')
                x = 0.5
                y = {'base': 0.75, 'stage': 0.5, 'mart': 0.25}
                for tier, arr in tiers.items():
                    for i, a in enumerate(sorted(set(arr))):
                        lbl = pascalize(alias_to_display.get(a, a))
                        rect = FancyBboxPatch((x+i*2.2, y[tier]), 1.8, 0.3,
                                              boxstyle="round,pad=0.02,rounding_size=0.08",
                                              linewidth=1.5, edgecolor="#333333",
                                              facecolor=("#F0F0F0" if tier=='base' else "#EAF3FE" if tier=='stage' else "#FFF2CC"))
                        ax.add_patch(rect)
                        ax.text(x+i*2.2+0.9, y[tier]+0.15, wrap_label(lbl, 18), ha='center', va='center', fontsize=9)
                for u, v in edges:
                    yu = y[alias_to_tier.get(u, 'mart')]
                    ax.plot([x+0.9, x+0.9], [yu+0.15, y['mart']+0.15], color="#888888")
                try:
                    fig.tight_layout()
                except Exception:
                    pass
                buf = BytesIO()
                fig.savefig(buf, format="png", dpi=150)
                plt.close(fig)
                buf.seek(0)
                pil_img = PILImage.open(buf)

                tab = safe_sheet_name(f"Lineage-{mart_disp}", used_names)
                ws_img = wb.create_sheet(tab)
                ws_img.add_image(XLImage(pil_img), "A1")

        # Graphviz lineage (robust mart discovery)
        if args.gv_lineage and graphviz_available():
            mart_list = set()
            # 1) Facts from Star Map
            if 'Star Map' in wb.sheetnames:
                wss = wb['Star Map']
                for r in range(2, wss.max_row+1):
                    fac = (wss.cell(r, 1).value or '').strip()
                    if fac:
                        mart_list.add(fac)
            # 2) Tiered 'mart'
            for alias, tier in alias_to_tier.items():
                if tier == 'mart':
                    mart_list.add(alias_to_display.get(alias, alias))
            # 3) Kind facts & dimensions
            for alias, kind in model_kinds.items():
                if kind in {'fact', 'dimension'}:
                    mart_list.add(alias_to_display.get(alias, alias))
            # 4) Tag 'mart'
            for alias, node in ( (m['alias_lc'], next(n for uid, n in manifest['nodes'].items()
                                                  if uid == m['unique_id'])) for m in models ):
                tags = set((node.get('tags') or []))
                if 'mart' in {t.lower() for t in tags}:
                    mart_list.add(alias_to_display.get(alias, alias))

            mart_list = {m for m in (m.strip() for m in mart_list) if m}
            if not mart_list:
                print("NOTE: gv-lineage: no mart models found; nothing to draw.", file=sys.stderr)
            else:
                print(f"gv-lineage: will render {len(mart_list)} mart models", file=sys.stderr)
                for mart_disp in sorted(mart_list):
                    print(f"gv-lineage: rendering {mart_disp}", file=sys.stderr)
                    mart_alias_lc = mart_disp.lower()
                    tiers = {"base": [], "stage": [], "mart": [mart_alias_lc]}
                    edges = []

                    # Pull upstream dims from Relationships
                    if 'Relationships' in wb.sheetnames:
                        wr = wb['Relationships']
                        for r in range(2, wr.max_row+1):
                            if (wr.cell(r, 1).value or '').strip() == mart_disp:
                                up = (wr.cell(r, 3).value or '').strip().lower()
                                if up:
                                    tiers.setdefault(alias_to_tier.get(up, 'mart'), []).append(up)
                                    edges.append((up, mart_alias_lc))

                    pil_img = render_lineage_graphviz_image(mart_disp, tiers, edges, alias_to_display)
                    if pil_img is not None:
                        tab = safe_sheet_name(f"Lineage-GV {mart_disp}", used_names)
                        ws_img = wb.create_sheet(tab)
                        ws_img.add_image(XLImage(pil_img), "A1")

    # Save
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"Wrote: {out_path}")


# ------------------------
# CLI
# ------------------------

def main():
    p = argparse.ArgumentParser(description="Generate a Kimball-style Excel workbook from dbt artifacts.")
    p.add_argument("--manifest", required=True, help="Path to target/manifest.json")
    p.add_argument("--catalog",  required=False, default="", help="Path to target/catalog.json")
    p.add_argument("--schemas",  required=False, default="", help="Comma-separated paths to schemas.yml files")

    # Sheet inclusion/exclusion
    p.add_argument("--include-views", action="store_true", help="Include view materializations in sheets/summary")
    p.add_argument("--materializations", default="", help="Only include models with these materializations (comma-list)")
    p.add_argument("--exclude-sheet-prefixes", default="", help="Comma-list of model name prefixes to exclude")
    p.add_argument("--exclude-sheet-tags", default="", help="Comma-list of tags to exclude")
    p.add_argument("--exclude-sheet-materializations", default="", help="Comma-list of materializations to exclude")
    p.add_argument("--exclude-sheet-path-globs", default="", help="Comma-list of path globs to exclude (e.g. models/base/*)")

    # Star diagrams
    p.add_argument("--star-diagrams", action="store_true", help="Render a star diagram tab per fact")
    p.add_argument("--diagram-shape", choices=["roundrect"], default="roundrect")
    p.add_argument("--diagram-dedupe-roles", action="store_true", help="Deduplicate role-playing dimensions by label")
    p.add_argument("--diagram-two-rings", type=int, default=12, help="Max nodes in outer ring before using inner ring")
    p.add_argument("--diagram-font-scale", type=float, default=0.9)
    p.add_argument("--diagram-dpi", type=int, default=240)
    p.add_argument("--diagram-color-scheme", choices=["soft", "mint"], default="soft")
    p.add_argument("--diagram-legend", action="store_true")
    p.add_argument("--diagram-no-wrap-labels", action="store_true", help="Do not wrap node labels")

    # Lineage
    p.add_argument("--lineage", action="store_true", help="Emit textual lineage table")
    p.add_argument("--lineage-diagrams", action="store_true", help="Emit per-mart lineage diagrams (Matplotlib fallback)")
    p.add_argument("--gv-lineage", action="store_true", help="Use Graphviz for lineage diagrams when available")
    p.add_argument("--gv-format", default="png", help="Graphviz output image format (png recommended)")

    p.add_argument("--out", required=True, help="Output XLSX path")
    args = p.parse_args()

    build_workbook(args)


if __name__ == "__main__":
    main()
