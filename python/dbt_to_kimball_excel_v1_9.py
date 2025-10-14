#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
dbt_to_kimball_excel.py  —  Generate a Kimball-style Excel catalog + star diagrams
=====================================================================

Version: 1.9.0

Overview
--------
This script converts a compiled **dbt Core** project (manifest + catalog) into
a Kimball-style Excel workbook:

1) **Summary** tab
   - One row per model with kind, materialization, relation, tags, etc.

2) **Relationships** tab  (can be omitted with --no-relationships)
   - One row per dbt *relationships* test (fact→dimension) with to/from fields.
   - Useful as a ground-truth, per-test audit.

3) **Star Map** tab
   - Aggregated (FactModel, DimensionModel) pairs with:
     - **FactFKColumn** (FK(s) on the fact side)
     - **DimKeyColumn** (key(s) on the dimension side)
   - By default, excludes relationships whose `to:` dimension model is missing
     from the manifest. With **--include-missing-dims**, those rows are included
     and the DimensionModel is suffixed with **"(missing)"**.

4) **One sheet per model**
   - Table meta at the top (#Model, #Kind, #Materialization, #Relation, #Tags,
     #SurrogateKey, #BusinessKey, #PrimaryKey, #Description)
   - Column catalog below with DataType, Nullable?, Description, Tests, IsPK?, IsFK?, Source
   - Frozen header row, wrapped descriptions, styled header

5) **Optional star-diagram tabs** (one per fact)
   - Matplotlib → PNG (portable)
   - Rounded rectangle for fact; rounded rectangles for dimensions
   - Labels wrap at Camel/Pascal boundaries (no spaces inserted — only line breaks)
   - Spoke labels show the **fact-side FK column names**
   - Colors by dimension family (Date, People, Project, Classification, Other)
   - Sheet name: **"Star-Fact <FactName>"** (fallback to "Star Fact <FactName>")

Data Sources
------------
- **manifest.json** (dbt `target/manifest.json`)
  - Model metadata, tests (including relationships), tags, paths, etc.
- **catalog.json** (dbt `target/catalog.json`)
  - Column types/comments (if available)
- **schemas YAML** (optional; comma-separated globs)
  - e.g., `models/mart/schemas.yml,models/stage/schemas.yml`
  - Overlays per-column data_type/nullable/description and per-model
    `meta.surrogate_key` / `meta.business_key` / `meta.primary_key`.

Installation
------------
Python 3.9+ recommended.

    pip install pandas openpyxl pyyaml matplotlib

(If you don’t need YAML overlays, you can skip `pyyaml`. If `matplotlib` is not
available, diagram tabs are silently skipped.)

Invocation (example)
--------------------
    python scripts/dbt_to_kimball_excel.py ^
      --manifest .\target\manifest.json ^
      --catalog  .\target\catalog.json ^
      --schemas ".\models\mart\schemas.yml,.\models\stage\schemas.yml" ^
      --exclude-sheet-prefixes base_,stage_ ^
      --exclude-sheet-tags stage,logging ^
      --exclude-sheet-materializations view ^
      --exclude-sheet-path-globs "models/base/*,models/stage/*,models/logging/*" ^
      --star-diagrams ^
      --diagram-two-rings 12 ^
      --diagram-font-scale 0.95 ^
      --diagram-dpi 220 ^
      --diagram-color-scheme soft ^
      --out .\documentation\kimball_model_catalog.xlsx

Common Flags
------------
--manifest PATH                      Required. dbt manifest.json
--catalog PATH                       Required. dbt catalog.json
--out PATH                           Required. Output .xlsx
--schemas CSV                        Optional CSV/globs for YAML schemas to overlay
--include-views                      Include view materializations when filtering by --materializations
--materializations CSV               Whitelist: e.g., table,incremental
--exclude-sheet-prefixes CSV         Exclude per-model tabs where alias starts with any prefix
--exclude-sheet-tags CSV             Exclude per-model tabs by tag
--exclude-sheet-materializations CSV Exclude per-model tabs by materialization (e.g., view)
--exclude-sheet-path-globs CSV       Exclude per-model tabs by path glob(s)
--no-relationships                   Omit the Relationships sheet (default: include)
--include-missing-dims               Include rows in Star Map whose target dimension model
                                     is not in the manifest (suffix "(missing)"). Diagrams
                                     still skip missing dimensions. (default: exclude)

Diagram Flags (if --star-diagrams is set)
-----------------------------------------
--star-diagrams                      Enable star diagram tabs (PNG)
--diagram-two-rings INT              Put excess dimensions on outer ring above this count
--diagram-font-scale FLOAT           Overall diagram font scale (default 1.0)
--diagram-dpi INT                    PNG DPI (default 240)
--diagram-color-scheme soft          (currently single palette)
--diagram-legend                     Add a small legend text box to the diagram sheet
--diagram-no-wrap-labels             Disable wrapping (single-line labels)
--diagram-max-label-chars INT        Max chars per line for labels (default 18)

Notes / Limitations
-------------------
- Relationships are inferred from dbt **relationships tests** in manifest.json.
- If a `relationships` test points to a dimension that isn’t in the manifest:
  - Default: excluded from Star Map and diagrams.
  - With --include-missing-dims: included in Star Map as "(missing)";
    diagrams still skip to remain robust.
- Sheet names are kept <= 31 chars and made safe automatically.

---------------------------------------------------------------------
"""

import argparse, json, re, sys, fnmatch, math, tempfile, os
from collections import defaultdict, OrderedDict
from pathlib import Path

import pandas as pd
try:
    import yaml
except Exception:
    yaml = None


# ----------------------------
# Core helpers
# ----------------------------
def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node):
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags):
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', base or 'Sheet')[:31] or 'Sheet'
    candidate = cleaned; i = 1
    while candidate in used:
        suffix = f'_{i}'; candidate = (cleaned[:31-len(suffix)] + suffix); i += 1
    used.add(candidate); return candidate

def _normalize_to_model(raw):
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"""ref\(\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    m = re.match(r"""source\(\s*['"][^'"]+['"]\s*,\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    return s

def extract_tests_for_model(manifest, node_id):
    """
    Return { column_name -> [ {name, details, raw}, ... ] } for tests that depend on node_id.
    Relationship tests include details {'to_model', 'to_field'} when present.
    """
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test':
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends:
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_model = _normalize_to_model(raw_to)
            rel_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    """
    Heuristics for IsPK?/IsFK?/Nullable? columns on per-model sheets.
    """
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')):
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    db, schema, alias = model_relation_identifiers(node)
    key = f'{db}.{schema}.{alias}'.lower()
    for k, v in (catalog.get('nodes') or {}).items():
        if k.lower() == key:
            return v.get('columns') or {}
    return {}

def build_relationship_rows_with_deps(manifest):
    """
    Return relationship edges pulled from test nodes and their 'depends_on' models.
    Each row includes DepModels (models in depends_on), ToModel (target of relationship),
    ToColumn, FromColumn, TestName. We choose a 'FromModel' later.
    """
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test':
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname:
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_model = _normalize_to_model(raw_to)
        to_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
        dep_models = []
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model':
                dep_models.append((n.get('alias') or n.get('name')))
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'DepModels': dep_models,'ToModel': to_model,'ToColumn': to_field,
                     'FromModel': None,'FromColumn': from_column,'TestName': tname})
    return rows

def _as_list(v):
    if v is None: return []
    if isinstance(v, (list, tuple)): return [str(x) for x in v]
    return [str(v)]

def load_yaml_overlays(paths_csv: str):
    """
    Parse one or more schema YAMLs and build overlay:
      overlays[model_name_lower] = {
        'model_meta': {'surrogate_key': [...], 'business_key': [...], 'primary_key': [...] },
        'columns': { col_lower: {'data_type':..., 'nullable':..., 'description':...}, ... }
      }
    """
    overlays = {}
    if not paths_csv or not yaml: return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc: continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname: continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})
                entry['model_meta'] = {
                    'surrogate_key': [s.lower() for s in _as_list(mm.get('surrogate_key'))],
                    'business_key':  [s.lower() for s in _as_list(mm.get('business_key'))],
                    'primary_key':   [s.lower() for s in _as_list(mm.get('primary_key'))],
                }
                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {'data_type': dtype, 'nullable': nullable, 'description': desc}.items() if v is not None}
    return overlays

def parse_csv_set(s: str):
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}

def should_exclude_sheet(alias_lower: str, tags_lower: set, mat_lower: str, path_str: str,
                         prefixes: set, tags_excl: set, mats_excl: set, path_globs: set) -> bool:
    if any(alias_lower.startswith(p) for p in prefixes): return True
    if any(t in tags_lower for t in tags_excl): return True
    if mat_lower in mats_excl: return True
    ps = (path_str or '')
    for g in path_globs:
        if fnmatch.fnmatch(ps.lower(), g): return True
    return False


# ----------------------------
# Diagram utilities — wrap only at camel-case, keep string intact
# ----------------------------
_CAMEL_SPLIT_RE = re.compile(
    r'(?<=[A-Za-z])(?=[A-Z][a-z])|(?<=[a-z])(?=[A-Z])|(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])'
)

def _camel_chunks(s: str):
    """Return list of substrings split at camel/Pascal boundaries. No spaces added."""
    if not s: return []
    idxs = [0]
    for m in _CAMEL_SPLIT_RE.finditer(s):
        idxs.append(m.start())
    idxs.append(len(s))
    return [s[idxs[i]:idxs[i+1]] for i in range(len(idxs)-1)]

def _wrap_preserving_string(s: str, max_chars: int, max_lines: int = 2):
    """
    Wrap camel/PascalCase text by inserting '\n' at boundaries.
    - No spaces are inserted into the string.
    - Produces up to `max_lines` lines.
    - Adds an ellipsis only on the *last* line if chunks remain.
    """
    chunks = _camel_chunks(s)
    if not chunks:
        return ['']

    lines = []
    i = 0  # index of next chunk to consume

    for _line_idx in range(max_lines):
        cur = ''
        # fill this line with as many whole chunks as will fit
        while i < len(chunks):
            cand = cur + chunks[i]
            if len(cand) <= max_chars or cur == '':
                cur = cand
                i += 1
            else:
                break

        if cur:  # we built something for this line
            lines.append(cur)
        else:
            # even a single chunk is longer than max_chars; hard-truncate
            lines.append((chunks[i][:max_chars-1] + '…') if max_chars > 1 else '…')
            i += 1

        # stop early if we consumed everything
        if i >= len(chunks):
            break

    # if leftovers remain, ellipsize the last line to indicate truncation
    if i < len(chunks):
        last = lines[-1]
        if len(last) >= max_chars:
            lines[-1] = last[:max(1, max_chars-1)] + '…'
        else:
            lines[-1] = last + '…'

    return lines

def _clip_to_circle(cx, cy, r, x, y):
    dx, dy = x - cx, y - cy
    d = math.hypot(dx, dy) or 1e-6
    return cx + r * dx / d, cy + r * dy / d

def _clip_to_rect(cx, cy, w, h, x, y):
    dx, dy = x - cx, y - cy
    if dx == 0 and dy == 0:
        return cx, cy
    tx = (w/2) / abs(dx) if dx != 0 else float('inf')
    ty = (h/2) / abs(dy) if dy != 0 else float('inf')
    t = min(tx, ty)
    return cx + dx * t, cy + dy * t

def _family_for_dim(name_lower: str, tags_lower: set):
    if 'date' in name_lower or 'date' in tags_lower: return 'date'
    if 'user' in name_lower or 'person' in name_lower or 'employee' in name_lower: return 'people'
    if 'project' in name_lower: return 'project'
    if any(t in tags_lower for t in ['classification','category','status','type']): return 'classification'
    return 'other'

def _palette_soft():
    return {
        'date': '#D7EBFF',
        'people': '#E9F7D9',
        'project': '#FDEFD0',
        'classification': '#F1E4FF',
        'other': '#D9F2F7',
        'fact': '#FFF2CC',
        'edge': '#8A8A8A',
        'node_edge': '#222222',
        'shadow': (0,0,0,0.08),
        'edge_text': '#333333',
    }

def draw_star_png(
    fact_label, fact_family, spokes, outfile,
    *,
    shape='roundrect',
    two_rings_threshold=None,
    font_scale=1.0,
    dpi=220,
    color_scheme='soft',
    wrap_labels=True,
    max_label_chars=18,
):
    """Draw PNG star. spokes = [{'dim_label','dim_family','edge_labels'}]."""
    try:
        import importlib
        plt = importlib.import_module('matplotlib.pyplot')
        from matplotlib.patches import Circle, FancyBboxPatch
    except Exception:
        return False

    if not spokes:
        return False

    PALETTE = _palette_soft() if color_scheme == 'soft' else _palette_soft()

    f_fact = int(13 * font_scale)
    f_dim  = int(10 * font_scale)
    f_edge = max(8, int(9 * font_scale))

    dims = spokes
    n = len(dims)
    inner_max = None
    if two_rings_threshold and n > two_rings_threshold:
        inner_max = two_rings_threshold
    inner = dims if not inner_max else dims[:inner_max]
    outer = [] if not inner_max else dims[inner_max:]

    inner_radius = 3.4 if len(inner) <= 12 else 3.8
    outer_radius = inner_radius + 1.1 if outer else None

    figsize = (7.5, 7.5)
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111)
    ax.set_aspect('equal'); ax.axis('off')
    cx, cy = 0.0, 0.0

    # ---- Fact node: rounded rect with dynamic width/height; wrap at camel boundaries ----
    fact_lines = (_wrap_preserving_string(fact_label, max_label_chars+4, 2)
                  if wrap_labels else [fact_label])
    longest = max(len(l) for l in fact_lines)
    rect_w = max(2.9, 0.13 * longest + 1.2)
    rect_h = (1.3 if len(fact_lines) == 1 else 1.55 + 0.42*max(0, len(fact_lines)-1))
    shadow_offset = 0.09

    ax.add_patch(FancyBboxPatch(
        (cx - rect_w/2 + shadow_offset, cy - rect_h/2 - shadow_offset), rect_w, rect_h,
        boxstyle="round,pad=0.02,rounding_size=0.25",
        linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
    ax.add_patch(FancyBboxPatch(
        (cx - rect_w/2, cy - rect_h/2), rect_w, rect_h,
        boxstyle="round,pad=0.02,rounding_size=0.25",
        linewidth=1.6, edgecolor=PALETTE['node_edge'], facecolor=PALETTE['fact'], zorder=2))

    if len(fact_lines) == 1:
        ax.text(cx, cy, fact_lines[0], ha='center', va='center', fontsize=f_fact, fontweight='bold', zorder=3)
    else:
        y0 = cy + (0.22 if len(fact_lines) == 2 else 0.33)
        for i, line in enumerate(fact_lines):
            ax.text(cx, y0 - i*0.34, line, ha='center', va='center',
                    fontsize=f_fact, fontweight='bold', zorder=3)

    def ring_positions(count, radius):
        if count <= 0: return []
        angles = [2*math.pi*i/count for i in range(count)]
        return [(cx + radius*math.cos(a), cy + radius*math.sin(a), a) for i, a in enumerate(angles)]

    inner_pos = ring_positions(len(inner), inner_radius)
    outer_pos = ring_positions(len(outer), outer_radius) if outer else []

    def draw_dim_node(x, y, label, family):
        node_edge = PALETTE['node_edge']
        face = PALETTE.get(family, PALETTE['other'])
        lines = (_wrap_preserving_string(label, max_label_chars, 2) if wrap_labels else [label])
        longest = max(len(l) for l in lines)
        w = max(2.1, 0.12 * longest + 0.9)
        h = (1.1 if len(lines) == 1 else 1.35)
        r = 0.35

        ax.add_patch(FancyBboxPatch(
            (x - w/2 + shadow_offset, y - h/2 - shadow_offset), w, h,
            boxstyle=f"round,pad=0.02,rounding_size={r}",
            linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
        ax.add_patch(FancyBboxPatch(
            (x - w/2, y - h/2), w, h,
            boxstyle=f"round,pad=0.02,rounding_size={r}",
            linewidth=1.6, edgecolor=node_edge, facecolor=face, zorder=2))

        if len(lines) == 1:
            ax.text(x, y, lines[0], ha='center', va='center', fontsize=f_dim, zorder=3)
        else:
            y0 = y + 0.16
            for i, line in enumerate(lines):
                ax.text(x, y0 - i*0.28, line, ha='center', va='center', fontsize=f_dim, zorder=3)

        return w, h

    def draw_spokes(positions, items):
        for (x, y, _), spec in zip(positions, items):
            w, h = draw_dim_node(x, y, spec['dim_label'], spec['dim_family'])
            rx, ry = _clip_to_rect(cx, cy, rect_w, rect_h, x, y)
            cx2, cy2 = _clip_to_circle(x, y, min(w, h)/2 - 0.02, cx, cy)
            ax.plot([rx, cx2], [ry, cy2], color=PALETTE['edge'], linewidth=1.3, zorder=1)
            if spec['edge_labels']:
                label = ', '.join([l for l in spec['edge_labels'] if l])
                midx = rx*0.35 + cx2*0.65
                midy = ry*0.35 + cy2*0.65
                dx, dy = cx2 - rx, cy2 - ry
                L = math.hypot(dx, dy) or 1.0
                off = 0.17
                ex = midx - off * dy / L
                ey = midy + off * dx / L
                if len(label) > 32:
                    parts, cur, total = [], [], 0
                    for tok in label.split(', '):
                        if total + len(tok) > 28 and cur:
                            parts.append(', '.join(cur)); cur = [tok]; total = len(tok)
                        else:
                            cur.append(tok); total += (len(tok) + 2)
                    if cur: parts.append(', '.join(cur))
                    label = '\n'.join(parts)
                ax.text(ex, ey, label, ha='center', va='center', fontsize=f_edge, color=PALETTE['edge_text'], zorder=3)

    draw_spokes(inner_pos, inner)
    if outer: draw_spokes(outer_pos, outer)

    fig.tight_layout(pad=0.55)
    fig.savefig(outfile, dpi=dpi)
    plt.close(fig)
    return True


# ----------------------------
# Main
# ----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true')
    parser.add_argument('--materializations', type=str, default='')
    parser.add_argument('--schemas', type=str, default='')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='')
    parser.add_argument('--exclude-sheet-tags', type=str, default='')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='')
    parser.add_argument('--no-relationships', action='store_true',
                        help='Omit the Relationships sheet.')
    parser.add_argument('--include-missing-dims', action='store_true',
                        help='Include rows in Star Map where the target dimension model is '
                             'missing from the manifest, suffixed with "(missing)". Diagrams still skip.')
    parser.add_argument('--star-diagrams', action='store_true')

    # Diagram flags
    parser.add_argument('--diagram-shape', choices=['roundrect','circle'], default='roundrect')
    parser.add_argument('--diagram-dedupe-roles', action='store_true')
    parser.add_argument('--diagram-two-rings', type=int, default=None)
    parser.add_argument('--diagram-font-scale', type=float, default=1.0)
    parser.add_argument('--diagram-dpi', type=int, default=240)
    parser.add_argument('--diagram-color-scheme', choices=['soft'], default='soft')
    parser.add_argument('--diagram-legend', action='store_true')
    parser.add_argument('--diagram-no-wrap-labels', action='store_true')
    parser.add_argument('--diagram-max-label-chars', type=int, default=18)

    args = parser.parse_args()

    manifest = load_json(args.manifest); catalog = load_json(args.catalog)

    # Filter models
    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    # Maps
    alias_to_node, name_to_alias, alias_to_display = {}, {}, {}
    for n in filtered:
        alias_display = (n.get('alias') or n.get('name') or '')
        alias_lower = alias_display.lower()
        name_lower  = (n.get('name') or '').lower()
        alias_to_node[alias_lower] = n
        alias_to_display[alias_lower] = alias_display
        name_to_alias[alias_lower] = alias_lower
        if name_lower:
            name_to_alias[name_lower] = alias_lower

    def canon(token: str):
        if not token: return None
        t = str(token).strip().strip('"').lower()
        return name_to_alias.get(t, t)

    # Relationships from tests
    raw_rel_rows = build_relationship_rows_with_deps(manifest)
    norm_rel_rows = []
    for r in raw_rel_rows:
        dep_aliases = [canon(x) for x in (r.get('DepModels') or [])]
        to_alias    = canon(r.get('ToModel'))
        from_alias = None
        for d in dep_aliases:
            if d and d != to_alias:
                from_alias = d; break
        if not from_alias and dep_aliases:
            from_alias = dep_aliases[0]
        if not to_alias or not from_alias:
            continue
        norm_rel_rows.append({
            'FromModel': from_alias,
            'FromColumn': r.get('FromColumn'),
            'ToModel': to_alias,
            'ToColumn': r.get('ToColumn'),
            'TestName': r.get('TestName'),
        })

    # Spokes per fact (skip missing dims unless include-missing-dims)
    spokes_raw = defaultdict(list)
    for r in norm_rel_rows:
        frm = r['FromModel']; to = r['ToModel']; col = r.get('FromColumn')
        if not frm or not to: 
            continue
        if (to not in alias_to_display) and (not args.include_missing_dims):
            continue  # default behavior: skip missing in spokes
        if to in alias_to_display:
            spokes_raw[frm].append((to, col))
        # if missing dim and include-missing-dims, we still don't add to spokes (diagrams skip)

    # YAML overlays (used for per-model sheets)
    yaml_overlays = load_yaml_overlays(args.schemas)

    # Exclusions for per-model tabs
    ex_prefixes = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags = parse_csv_set(args.exclude_sheet_tags)
    ex_mats = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    # Excel formatting
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XLImage

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    col_widths = {"A": 28, "B": 20, "C": 10, "D": 80, "E": 28, "F": 6, "G": 6, "H": 12}

    # --------------------------
    # Summary / Relationships / Star Map
    # --------------------------
    summary_rows = []
    for alias_lower, n in alias_to_node.items():
        db, schema, alias_disp = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias_lower, n.get('tags', []))
        summary_rows.append({'Model': alias_disp,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias_disp}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    df_summary = pd.DataFrame(summary_rows, columns=['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description'])
    if not df_summary.empty:
        df_summary = df_summary.sort_values(['Kind','Model'])

    if not args.no_relationships:
        df_rel = pd.DataFrame([
            {'FromModel': alias_to_display.get(r['FromModel'], r['FromModel']),
             'FromColumn': r.get('FromColumn'),
             'ToModel': alias_to_display.get(r['ToModel'], r['ToModel']) if r['ToModel'] in alias_to_display else (r['ToModel'] or '(missing)'),
             'ToColumn': r.get('ToColumn'),
             'TestName': r.get('TestName')}
            for r in norm_rel_rows
        ]) if norm_rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])
    else:
        df_rel = None  # omitted

    # --- Star Map with FK/PK columns (aggregated per Fact/Dim pair) ---
    star_map = OrderedDict()
    for r in norm_rel_rows:
        f = alias_to_display.get(r['FromModel'], r['FromModel'])
        if r['ToModel'] in alias_to_display:
            d = alias_to_display.get(r['ToModel'], r['ToModel'])
        else:
            if not args.include_missing_dims:
                continue  # default: drop missing from Star Map
            d_raw = r['ToModel'] or '(missing)'
            d = f"{d_raw}(missing)" if not d_raw.endswith('(missing)') else d_raw

        key = (f, d)
        entry = star_map.setdefault(
            key,
            {'FactModel': f, 'DimensionModel': d, 'FactFKColumn': set(), 'DimKeyColumn': set()}
        )
        if r.get('FromColumn'):
            entry['FactFKColumn'].add(r['FromColumn'])
        if r.get('ToColumn'):
            entry['DimKeyColumn'].add(r['ToColumn'])

    star_rows = []
    for (_f, _d), e in star_map.items():
        star_rows.append({
            'FactModel': e['FactModel'],
            'DimensionModel': e['DimensionModel'],
            'FactFKColumn': ', '.join(sorted(e['FactFKColumn'])) if e['FactFKColumn'] else '',
            'DimKeyColumn': ', '.join(sorted(e['DimKeyColumn'])) if e['DimKeyColumn'] else '',
        })

    df_star = (
        pd.DataFrame(star_rows, columns=['FactModel','DimensionModel','FactFKColumn','DimKeyColumn'])
        if star_rows else
        pd.DataFrame(columns=['FactModel','DimensionModel','FactFKColumn','DimKeyColumn'])
    )

    used_sheet_names = set()

    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        if df_rel is not None and not df_rel.empty:
            df_rel.to_excel(writer, index=False, sheet_name='Relationships')
        if not df_star.empty:
            df_star.to_excel(writer, index=False, sheet_name='Star Map')

        # --------------------------
        # Per-model tabs
        # --------------------------
        for alias_lower, n in alias_to_node.items():
            db, schema, alias_disp = model_relation_identifiers(n)
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''
            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            sheet_name = safe_sheet_name(alias_disp or 'Model', used_sheet_names)

            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            catalog_cols = get_catalog_columns_for_model(catalog, n)
            overlay_entry = (yaml_overlays.get(alias_lower, {}) or {})
            overlay_cols = overlay_entry.get('columns', {}) or {}
            model_meta = overlay_entry.get('model_meta', {}) or {}
            pk_from_model = set(model_meta.get('primary_key', []) or [])
            sk_from_model = set(model_meta.get('surrogate_key', []) or [])
            bk_from_model = set(model_meta.get('business_key', []) or [])
            explicit_pk_names = {*pk_from_model, *sk_from_model}

            rows = []
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or (meta.get('type') or '')
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            info_rows = [
                {'Column': '#Model','DataType': alias_disp,'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Kind','DataType': classify_model_kind(alias_lower, n.get('tags', [])),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Materialization','DataType': (n.get('config') or {}).get('materialized'),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Relation','DataType': f'{db}.{schema}.{alias_disp}','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Tags','DataType': ','.join(n.get('tags') or []),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#SurrogateKey','DataType': ','.join(sorted(sk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#BusinessKey','DataType': ','.join(sorted(bk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#PrimaryKey','DataType': ','.join(sorted(pk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Description','DataType': (n.get('description') or ''),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            df_meta.to_excel(writer, index=False, sheet_name=sheet_name, header=False, startrow=0)
            start_row = len(df_meta.index) + 1
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

            wb = writer.book; ws = wb[sheet_name]
            ws.freeze_panes = f"A{start_row+2}"
            hdr = start_row + 1
            for col_idx in range(1, 8+1):
                c = ws.cell(row=hdr, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = Alignment(vertical="center")
            for r in range(hdr+1, ws.max_row+1):
                ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        # --------------------------
        # Diagram tabs (PNG)
        # --------------------------
        if args.star_diagrams:
            tmpdir = tempfile.mkdtemp(prefix="dbt_star_")
            wb = writer.book
            used = set(wb.sheetnames)

            for center_alias_lower, pairs in sorted(spokes_raw.items()):
                # Dedupe by dimension if requested; aggregate FK labels per dimension
                if args.diagram_dedupe_roles:
                    by_dim = OrderedDict()
                    for dim_alias_lower, fk_col in pairs:
                        by_dim.setdefault(dim_alias_lower, []).append(fk_col or '')
                    spokes = []
                    for dim_alias_lower, fk_cols in by_dim.items():
                        dn = alias_to_node.get(dim_alias_lower)
                        tags_lower = {t.lower() for t in (dn.get('tags') or [])} if dn else set()
                        family = _family_for_dim(dim_alias_lower, tags_lower)
                        spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                       'dim_family': family,
                                       'edge_labels': [c for c in fk_cols if c]})
                else:
                    spokes = []
                    for dim_alias_lower, fk_col in pairs:
                        dn = alias_to_node.get(dim_alias_lower)
                        tags_lower = {t.lower() for t in (dn.get('tags') or [])} if dn else set()
                        family = _family_for_dim(dim_alias_lower, tags_lower)
                        spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                       'dim_family': family,
                                       'edge_labels': [fk_col] if fk_col else []})

                if not spokes:
                    continue

                fact_label = alias_to_display.get(center_alias_lower, center_alias_lower)
                img_path = os.path.join(tmpdir, f"{center_alias_lower}.png")

                drew = draw_star_png(
                    fact_label, 'other', spokes, img_path,
                    shape='roundrect',
                    two_rings_threshold=args.diagram_two_rings,
                    font_scale=args.diagram_font_scale,
                    dpi=args.diagram_dpi,
                    color_scheme=args.diagram_color_scheme,
                    wrap_labels=not args.diagram_no_wrap_labels,
                    max_label_chars=max(10, args.diagram_max_label_chars),
                )
                if not drew:
                    continue

                # Sheet name: prefer "Star-Fact <Fact>", otherwise "Star Fact <Fact>"
                proposed = f"Star-Fact {fact_label}"
                if any(c in proposed for c in r'[]:*?\/'):
                    proposed = f"Star Fact {fact_label}"
                sheet_name = safe_sheet_name(proposed, used)

                ws = wb.create_sheet(title=sheet_name)
                try:
                    img = XLImage(img_path)
                    ws.add_image(img, "A1")
                except Exception:
                    pass

                if args.diagram_legend:
                    ws["J2"] = ("Legend: node colors by family "
                                "(Date, People, Project, Classification, Other).\n"
                                "Spoke text = fact FK column(s).")
                    ws["J2"].alignment = Alignment(wrap_text=True)

    print(f"Wrote: {args.out}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
