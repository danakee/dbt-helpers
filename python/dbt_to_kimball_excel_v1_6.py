
#!/usr/bin/env python3
"""
dbt_to_kimball_excel.py  (v1.6)

What's new in v1.6
------------------
- Per-model sheets now render **table meta** (the `#...` rows) at the **top**, then a blank spacer,
  then the **column header** and column metadata rows.
- Formatting:
  * Header row has a blue background and bold text.
  * Description column is word-wrapped and uses a wider default width.
  * Sensible default column widths for readability.
  * Freeze panes so the **table meta** stays visible while scrolling the column list.
- All v1.5 features remain: YAML overlays, model-level meta (surrogate/business/primary keys),
  sheet-level exclusions, custom materializations, Relationships & Star Map, etc.
"""
import argparse, json, re, sys, fnmatch
from collections import defaultdict
from pathlib import Path

import pandas as pd
try:
    import yaml  # for reading schema overlays
except Exception:
    yaml = None

# ----------------------------
# Helpers (same behavior as v1.5)
# ----------------------------
def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node): 
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags):
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', base or 'Sheet')[:31] or 'Sheet'
    candidate = cleaned; i = 1
    while candidate in used:
        suffix = f'_{i}'; candidate = (cleaned[:31-len(suffix)] + suffix); i += 1
    used.add(candidate); return candidate

def extract_tests_for_model(manifest, node_id):
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test': 
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends: 
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            rel_model = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')): 
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    db, schema, alias = model_relation_identifiers(node)
    key = f'{db}.{schema}.{alias}'.lower()
    for k, v in (catalog.get('nodes') or {}).items():
        if k.lower() == key: 
            return v.get('columns') or {}
    return {}

def build_relationship_rows(manifest):
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test': 
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname: 
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        to_model = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
        from_model_node = None
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model': 
                from_model_node = n; break
        from_model = (from_model_node.get('alias') or from_model_node.get('name')) if from_model_node else None
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'FromModel': from_model,'FromColumn': from_column,'ToModel': to_model,'ToColumn': to_field,'TestName': tname})
    return rows

def guess_fact_groups(rel_rows, model_kinds):
    fact_to_dims = defaultdict(set)
    for r in rel_rows:
        frm = (r.get('FromModel') or '').lower(); to = (r.get('ToModel') or '').lower()
        if not frm or not to: 
            continue
        frm_kind = model_kinds.get(frm, 'Other'); to_kind = model_kinds.get(to, 'Other')
        if frm_kind == 'Fact' and to_kind in {'Dimension','Other'}: 
            fact_to_dims[frm].add(to)
        elif to_kind == 'Fact' and frm_kind in {'Dimension','Other'}: 
            fact_to_dims[to].add(frm)
    return {k: sorted(v) for k, v in fact_to_dims.items()}

def _as_list(v):
    if v is None: 
        return []
    if isinstance(v, (list, tuple)):
        return [str(x) for x in v]
    return [str(v)]

def load_yaml_overlays(paths_csv: str):
    overlays = {}
    if not paths_csv or not yaml:
        return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc:
            continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname:
                    continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})
                entry['model_meta'] = {
                    'surrogate_key': [s.lower() for s in _as_list(mm.get('surrogate_key'))],
                    'business_key':  [s.lower() for s in _as_list(mm.get('business_key'))],
                    'primary_key':   [s.lower() for s in _as_list(mm.get('primary_key'))],
                }
                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: 
                        continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {
                        'data_type': dtype, 'nullable': nullable, 'description': desc
                    }.items() if v is not None}
    return overlays

def parse_csv_set(s: str):
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}

def should_exclude_sheet(alias_lower: str, tags_lower: set, mat_lower: str, path_str: str,
                         prefixes: set, tags_excl: set, mats_excl: set, path_globs: set) -> bool:
    if any(alias_lower.startswith(p) for p in prefixes):
        return True
    if any(t in tags_lower for t in tags_excl):
        return True
    if mat_lower in mats_excl:
        return True
    ps = (path_str or '')
    for g in path_globs:
        if fnmatch.fnmatch(ps.lower(), g):
            return True
    return False

# ----------------------------
# Main
# ----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true', help='Only useful if you also pass --materializations')
    parser.add_argument('--materializations', type=str, default='', help='Comma-separated list; if provided, only include these mats. Default: include all except ephemeral.')
    parser.add_argument('--schemas', type=str, default='', help='Comma-separated YAML paths or globs to overlay column metadata.')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='', help='Comma-separated name prefixes to exclude from per-model sheets.')
    parser.add_argument('--exclude-sheet-tags', type=str, default='', help='Comma-separated tags to exclude from per-model sheets.')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='', help='Comma-separated materializations to exclude from per-model sheets.')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='', help='Comma-separated glob patterns matched against model path.')

    args = parser.parse_args()

    manifest = load_json(args.manifest); catalog = load_json(args.catalog)
    yaml_overlays = load_yaml_overlays(args.schemas)

    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]

    # Filter models
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    # Kinds map
    model_kinds = {}; name_to_node = {}
    for n in filtered:
        alias = (n.get('alias') or n.get('name') or '').lower()
        model_kinds[alias] = classify_model_kind(alias, n.get('tags', []))
        name_to_node[alias] = n

    # Relationships / stars
    rel_rows = build_relationship_rows(manifest); fact_groups = guess_fact_groups(rel_rows, model_kinds)

    # Summary (unchanged)
    summary_rows = []
    for n in filtered:
        db, schema, alias = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias, n.get('tags', []))
        summary_rows.append({'Model': alias,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    summary_cols = ['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description']
    df_summary = pd.DataFrame(summary_rows, columns=summary_cols)
    if not df_summary.empty: 
        df_summary = df_summary.sort_values(['Kind','Model'])

    df_rel = pd.DataFrame(rel_rows) if rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])

    star_rows = []
    for fact, dims in fact_groups.items():
        if not dims: 
            star_rows.append({'FactModel': fact,'DimensionModel': None})
        else:
            for d in dims: 
                star_rows.append({'FactModel': fact,'DimensionModel': d})
    df_star = pd.DataFrame(star_rows) if star_rows else pd.DataFrame(columns=['FactModel','DimensionModel'])

    ex_prefixes = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags = parse_csv_set(args.exclude_sheet_tags)
    ex_mats = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    used_sheet_names = set()
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    col_widths = {"A": 28, "B": 20, "C": 10, "D": 80, "E": 28, "F": 6, "G": 6, "H": 12}

    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        if not df_rel.empty: df_rel.to_excel(writer, index=False, sheet_name='Relationships')
        if not df_star.empty: df_star.to_excel(writer, index=False, sheet_name='Star Map')

        header_row_map = {}

        for n in filtered:
            db, schema, alias = model_relation_identifiers(n)
            alias_lower = (alias or '').lower()
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''

            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            sheet_name = safe_sheet_name(alias or 'Model', used_sheet_names)

            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            catalog_cols = get_catalog_columns_for_model(catalog, n)

            overlay_entry = yaml_overlays.get(alias_lower, {}) or {}
            overlay_cols = overlay_entry.get('columns', {}) or {}
            model_meta = overlay_entry.get('model_meta', {}) or {}
            pk_from_model = set(model_meta.get('primary_key', []) or [])
            sk_from_model = set(model_meta.get('surrogate_key', []) or [])
            bk_from_model = set(model_meta.get('business_key', []) or [])
            explicit_pk_names = {*pk_from_model, *sk_from_model}

            rows = []
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    dtype_catalog = meta.get('type') or ''
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or dtype_catalog
                    nullable = ov.get('nullable')
                    if nullable is None: 
                        nullable = nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable')
                    if nullable is None: 
                        nullable = nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            info_rows = [
                {'Column': '#Model','DataType': alias,'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Kind','DataType': classify_model_kind(alias, n.get('tags', [])),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Materialization','DataType': (n.get('config') or {}).get('materialized'),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Relation','DataType': f'{db}.{schema}.{alias}','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Tags','DataType': ','.join(n.get('tags') or []),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#SurrogateKey','DataType': ','.join(sorted(sk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#BusinessKey','DataType': ','.join(sorted(bk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#PrimaryKey','DataType': ','.join(sorted(pk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Description','DataType': (n.get('description') or ''),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            # write meta (no header)
            df_meta.to_excel(writer, index=False, sheet_name=sheet_name, header=False, startrow=0)
            # leave a blank row, then write df with header
            meta_rows = len(df_meta.index)
            start_row = meta_rows + 1  # zero-based index for pandas
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

            # remember where the header row is for styling (convert to 1-based)
            header_row_map[sheet_name] = start_row + 1

        # Styling
        wb = writer.book
        for sheet_name, header_row in header_row_map.items():
            ws = wb[sheet_name]
            ws.freeze_panes = f"A{header_row}"  # keep table meta visible

            # Header styling
            for col_idx in range(1, 9):  # A..H
                cell = ws.cell(row=header_row, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")

            # Wrap description cells
            last_row = ws.max_row
            for r in range(header_row+1, last_row+1):
                c = ws.cell(row=r, column=4)  # column D
                c.alignment = Alignment(wrap_text=True, vertical="top")

            # Column widths
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

    print(f'Wrote: {args.out}')

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
