#!/usr/bin/env python3
r"""
dbt_to_kimball_excel.py  (v1.7.3)

Fixes vs 1.7.2
--------------
• Accurately determines the FROM side of dbt `relationships` tests by comparing the test's
  `to:` (normalized ref()/source()) against the test's dependent model nodes and selecting
  the "other" model as the child/from model.
• Canonicalizes names/aliases consistently before grouping.
• Star tabs are created for any model with outgoing relationships (not only those tagged Fact).
• Skips missing/WIP dimensions and skips creating an empty star (no more single-box stars).
• Still supports YAML using either `tests:` or `data_tests:` (we read compiled manifest tests).
"""

import argparse, json, re, sys, fnmatch, math, tempfile, os
from collections import defaultdict
from pathlib import Path

import pandas as pd
try:
    import yaml  # overlay support
except Exception:
    yaml = None


# ----------------------------
# Helpers
# ----------------------------
def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node):
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags):
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', base or 'Sheet')[:31] or 'Sheet'
    candidate = cleaned; i = 1
    while candidate in used:
        suffix = f'_{i}'; candidate = (cleaned[:31-len(suffix)] + suffix); i += 1
    used.add(candidate); return candidate

def _normalize_to_model(raw):
    """Normalize ref()/source() strings to plain names."""
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"""ref\(\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    m = re.match(r"""source\(\s*['"][^'"]+['"]\s*,\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    return s

def extract_tests_for_model(manifest, node_id):
    """column -> [tests] for tests depending on `node_id` (compiled manifest)."""
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test':
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends:
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_model = _normalize_to_model(raw_to)
            rel_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')):
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    db, schema, alias = model_relation_identifiers(node)
    key = f'{db}.{schema}.{alias}'.lower()
    for k, v in (catalog.get('nodes') or {}).items():
        if k.lower() == key:
            return v.get('columns') or {}
    return {}

# Return raw rel rows + the set of dependent model node ids for each test
def build_relationship_rows_with_deps(manifest):
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test':
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname:
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_model = _normalize_to_model(raw_to)
        to_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
        # Collect all dependent model aliases/names (raw for now)
        dep_models = []
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model':
                dep_models.append((n.get('alias') or n.get('name')))
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'DepModels': dep_models,'ToModel': to_model,'ToColumn': to_field,
                     'FromModel': None,'FromColumn': from_column,'TestName': tname})
    return rows

def _as_list(v):
    if v is None: return []
    if isinstance(v, (list, tuple)): return [str(x) for x in v]
    return [str(v)]

def load_yaml_overlays(paths_csv: str):
    overlays = {}
    if not paths_csv or not yaml: return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc: continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname: continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})
                entry['model_meta'] = {
                    'surrogate_key': [s.lower() for s in _as_list(mm.get('surrogate_key'))],
                    'business_key':  [s.lower() for s in _as_list(mm.get('business_key'))],
                    'primary_key':   [s.lower() for s in _as_list(mm.get('primary_key'))],
                }
                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {
                        'data_type': dtype, 'nullable': nullable, 'description': desc
                    }.items() if v is not None}
    return overlays

def parse_csv_set(s: str):
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}

def should_exclude_sheet(alias_lower: str, tags_lower: set, mat_lower: str, path_str: str,
                         prefixes: set, tags_excl: set, mats_excl: set, path_globs: set) -> bool:
    if any(alias_lower.startswith(p) for p in prefixes): return True
    if any(t in tags_lower for t in tags_excl): return True
    if mat_lower in mats_excl: return True
    ps = (path_str or '')
    for g in path_globs:
        if fnmatch.fnmatch(ps.lower(), g): return True
    return False

# ---------- Diagram drawing ----------
def draw_star_png(fact_name, dim_names, outfile, figsize=(6.5,6.5)):
    try:
        import importlib
        plt = importlib.import_module('matplotlib.pyplot')
        from matplotlib.patches import Circle, Rectangle
    except Exception:
        return False

    if not dim_names:  # nothing to draw
        return False

    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111)
    ax.set_aspect('equal'); ax.axis('off')

    center = (0,0); radius = 3.0
    n = len(dim_names)
    angles = [2*math.pi*i/n for i in range(n)]
    positions = [(center[0]+radius*math.cos(a), center[1]+radius*math.sin(a)) for a in angles]

    for (x,y) in positions:
        ax.plot([center[0], x], [center[1], y], color="#777777", linewidth=1)

    rect_w, rect_h = 2.4, 1.2
    fact_rect = Rectangle((center[0]-rect_w/2, center[1]-rect_h/2), rect_w, rect_h,
                          linewidth=1.5, edgecolor='black', facecolor='#FFF2CC')
    ax.add_patch(fact_rect)
    ax.text(center[0], center[1], fact_name, ha='center', va='center', fontsize=12, fontweight='bold')

    for (x,y), label in zip(positions, dim_names):
        circ = Circle((x,y), 1.0, linewidth=1.5, edgecolor='black', facecolor='#CCF2F9')
        ax.add_patch(circ)
        ax.text(x, y, label, ha='center', va='center', fontsize=10)

    fig.tight_layout(pad=0.5)
    fig.savefig(outfile, dpi=160)
    plt.close(fig)
    return True


# ----------------------------
# Main
# ----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true')
    parser.add_argument('--materializations', type=str, default='')
    parser.add_argument('--schemas', type=str, default='')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='')
    parser.add_argument('--exclude-sheet-tags', type=str, default='')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='')
    parser.add_argument('--star-diagrams', action='store_true', help='Add a tab per model with outgoing relationships')

    args = parser.parse_args()

    manifest = load_json(args.manifest); catalog = load_json(args.catalog)

    # Filter models
    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    # Kinds + lookups (name→alias mapping)
    model_kinds = {}
    name_to_node = {}
    name_to_alias = {}

    for n in filtered:
        alias = (n.get('alias') or n.get('name') or '').lower()
        name  = (n.get('name') or '').lower()
        model_kinds[alias] = classify_model_kind(alias, n.get('tags', []))
        name_to_node[alias] = n
        name_to_alias[alias] = alias
        if name:
            name_to_alias[name] = alias

    def canonicalize_model_token(token: str):
        if not token:
            return None
        t = str(token).strip().strip('"').lower()
        return name_to_alias.get(t, t)

    # Relationships (use dependents to select true FROM model)
    raw_rel_rows = build_relationship_rows_with_deps(manifest)
    norm_rel_rows = []
    for r in raw_rel_rows:
        dep_aliases = [canonicalize_model_token(x) for x in (r.get('DepModels') or [])]
        to_alias    = canonicalize_model_token(r.get('ToModel'))
        # pick FROM as the first dep that is not the TO
        from_alias = None
        for d in dep_aliases:
            if d and d != to_alias:
                from_alias = d
                break
        # fallback: if only one dep or couldn't find non-TO, use whatever we have
        if not from_alias and dep_aliases:
            from_alias = dep_aliases[0]

        if not to_alias or not from_alias:
            continue  # nothing usable

        norm_rel_rows.append({
            'FromModel': from_alias,
            'FromColumn': r.get('FromColumn'),
            'ToModel': to_alias,
            'ToColumn': r.get('ToColumn'),
            'TestName': r.get('TestName'),
        })

    # Build groups: any model with outgoing relationships becomes a center
    fact_to_dims = defaultdict(set)
    for r in norm_rel_rows:
        frm = r['FromModel']; to = r['ToModel']
        if not frm or not to:
            continue
        # Only include dimensions we actually know about (skip WIP/missing quietly)
        if to not in name_to_alias.values():
            continue
        fact_to_dims[frm].add(to)

    # YAML overlays
    yaml_overlays = load_yaml_overlays(args.schemas)

    # Exclusions
    ex_prefixes = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags = parse_csv_set(args.exclude_sheet_tags)
    ex_mats = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    # Formatting
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XLImage

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    col_widths = {"A": 28, "B": 20, "C": 10, "D": 80, "E": 28, "F": 6, "G": 6, "H": 12}

    # Summary / Relationships / Star Map data
    summary_rows = []
    for n in filtered:
        db, schema, alias = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias, n.get('tags', []))
        summary_rows.append({'Model': alias,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    df_summary = pd.DataFrame(summary_rows, columns=['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description'])
    if not df_summary.empty:
        df_summary = df_summary.sort_values(['Kind','Model'])

    df_rel = pd.DataFrame(norm_rel_rows) if norm_rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])

    star_rows = []
    for center, dims in fact_to_dims.items():
        for d in sorted(dims):
            star_rows.append({'FactModel': center, 'DimensionModel': d})
    df_star = pd.DataFrame(star_rows) if star_rows else pd.DataFrame(columns=['FactModel','DimensionModel'])

    used_sheet_names = set()

    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        # summary tabs
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        if not df_rel.empty: df_rel.to_excel(writer, index=False, sheet_name='Relationships')
        if not df_star.empty: df_star.to_excel(writer, index=False, sheet_name='Star Map')

        # per-model tabs (same layout/format as before)
        for n in filtered:
            db, schema, alias = model_relation_identifiers(n)
            alias_lower = (alias or '').lower()
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''

            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            sheet_name = safe_sheet_name(alias or 'Model', used_sheet_names)

            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            catalog_cols = get_catalog_columns_for_model(catalog, n)
            overlay_entry = yaml_overlays.get(alias_lower, {}) or {}
            overlay_cols = overlay_entry.get('columns', {}) or {}
            model_meta = overlay_entry.get('model_meta', {}) or {}
            pk_from_model = set(model_meta.get('primary_key', []) or [])
            sk_from_model = set(model_meta.get('surrogate_key', []) or [])
            bk_from_model = set(model_meta.get('business_key', []) or [])
            explicit_pk_names = {*pk_from_model, *sk_from_model}

            rows = []
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or (meta.get('type') or '')
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            info_rows = [
                {'Column': '#Model','DataType': alias,'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Kind','DataType': classify_model_kind(alias, n.get('tags', [])),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Materialization','DataType': (n.get('config') or {}).get('materialized'),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Relation','DataType': f'{db}.{schema}.{alias}','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Tags','DataType': ','.join(n.get('tags') or []),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#SurrogateKey','DataType': ','.join(sorted(sk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#BusinessKey','DataType': ','.join(sorted(bk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#PrimaryKey','DataType': ','.join(sorted(pk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Description','DataType': (n.get('description') or ''),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            # write meta then data
            df_meta.to_excel(writer, index=False, sheet_name=sheet_name, header=False, startrow=0)
            start_row = len(df_meta.index) + 1
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

            # style
            wb = writer.book; ws = wb[sheet_name]
            ws.freeze_panes = f"A{start_row+2}"  # freeze meta + header
            hdr = start_row + 1
            for col_idx in range(1, 8+1):
                c = ws.cell(row=hdr, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = Alignment(vertical="center")
            for r in range(hdr+1, ws.max_row+1):
                ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        # Star diagram tabs (skip if libs missing; skip empty stars)
        if args.star_diagrams:
            tmpdir = tempfile.mkdtemp(prefix="dbt_star_")
            wb = writer.book
            for center, dims in sorted(fact_to_dims.items()):
                dims_known = [d for d in sorted(dims) if d in name_to_alias.values()]
                if not dims_known:
                    continue
                img_path = os.path.join(tmpdir, f"{center}.png")
                disp_dims = [d.replace('_',' ').title() for d in dims_known]
                drew = draw_star_png(center.replace('_',' ').title(), disp_dims, img_path, figsize=(6.5,6.5))
                if not drew:
                    continue
                sheet_name = safe_sheet_name(f"Star: {center}", set(wb.sheetnames))
                ws = wb.create_sheet(title=sheet_name)
                try:
                    from openpyxl.drawing.image import Image as XLImage
                    img = XLImage(img_path)
                    ws.add_image(img, "A1")
                except Exception:
                    pass

    print(f"Wrote: {args.out}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
