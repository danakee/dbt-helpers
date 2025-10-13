#!/usr/bin/env python3
r"""
dbt_to_kimball_excel.py  (v1.7.6)

High-quality Kimball-style workbook from a dbt project:
- Per-model tabs with table meta block + column metadata (frozen header, wrapped descriptions)
- Summary / Relationships / Star Map tabs
- Star diagrams (optionally): rounded-rect or circle nodes, clipped lines, FK labels on spokes
- Optional dedupe of repeated dimension roles (e.g., many DimDate roles → one bubble)
- Optional two-ring layout for busy stars, font scaling, DPI control, soft color palette
- Uses exact alias casing; gracefully skips missing/WIP dims in relationships

Inputs: dbt `manifest.json`, `catalog.json`, and optional YAML overlays with column meta
"""

import argparse, json, re, sys, fnmatch, math, tempfile, os
from collections import defaultdict, OrderedDict
from pathlib import Path

import pandas as pd
try:
    import yaml
except Exception:
    yaml = None


# ----------------------------
# Core helpers
# ----------------------------
def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node):
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags):
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', base or 'Sheet')[:31] or 'Sheet'
    candidate = cleaned; i = 1
    while candidate in used:
        suffix = f'_{i}'; candidate = (cleaned[:31-len(suffix)] + suffix); i += 1
    used.add(candidate); return candidate

def _normalize_to_model(raw):
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"""ref\(\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    m = re.match(r"""source\(\s*['"][^'"]+['"]\s*,\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    return s

def extract_tests_for_model(manifest, node_id):
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test':
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends:
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_model = _normalize_to_model(raw_to)
            rel_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')):
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    db, schema, alias = model_relation_identifiers(node)
    key = f'{db}.{schema}.{alias}'.lower()
    for k, v in (catalog.get('nodes') or {}).items():
        if k.lower() == key:
            return v.get('columns') or {}
    return {}

def build_relationship_rows_with_deps(manifest):
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test':
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname:
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_model = _normalize_to_model(raw_to)
        to_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
        dep_models = []
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model':
                dep_models.append((n.get('alias') or n.get('name')))
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'DepModels': dep_models,'ToModel': to_model,'ToColumn': to_field,
                     'FromModel': None,'FromColumn': from_column,'TestName': tname})
    return rows

def _as_list(v):
    if v is None: return []
    if isinstance(v, (list, tuple)): return [str(x) for x in v]
    return [str(v)]

def load_yaml_overlays(paths_csv: str):
    overlays = {}
    if not paths_csv or not yaml: return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc: continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname: continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})
                entry['model_meta'] = {
                    'surrogate_key': [s.lower() for s in _as_list(mm.get('surrogate_key'))],
                    'business_key':  [s.lower() for s in _as_list(mm.get('business_key'))],
                    'primary_key':   [s.lower() for s in _as_list(mm.get('primary_key'))],
                }
                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {'data_type': dtype, 'nullable': nullable, 'description': desc}.items() if v is not None}
    return overlays

def parse_csv_set(s: str):
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}

def should_exclude_sheet(alias_lower: str, tags_lower: set, mat_lower: str, path_str: str,
                         prefixes: set, tags_excl: set, mats_excl: set, path_globs: set) -> bool:
    if any(alias_lower.startswith(p) for p in prefixes): return True
    if any(t in tags_lower for t in tags_excl): return True
    if mat_lower in mats_excl: return True
    ps = (path_str or '')
    for g in path_globs:
        if fnmatch.fnmatch(ps.lower(), g): return True
    return False


# ----------------------------
# Diagram utilities
# ----------------------------
def _clip_to_circle(cx, cy, r, x, y):
    dx, dy = x - cx, y - cy
    d = math.hypot(dx, dy) or 1e-6
    return cx + r * dx / d, cy + r * dy / d

def _clip_to_rect(cx, cy, w, h, x, y):
    dx, dy = x - cx, y - cy
    if dx == 0 and dy == 0:
        return cx, cy
    tx = (w/2) / abs(dx) if dx != 0 else float('inf')
    ty = (h/2) / abs(dy) if dy != 0 else float('inf')
    t = min(tx, ty)
    return cx + dx * t, cy + dy * t

def _family_for_dim(name_lower: str, tags_lower: set):
    if 'date' in name_lower or 'date' in tags_lower: return 'date'
    if 'user' in name_lower or 'person' in name_lower or 'employee' in name_lower: return 'people'
    if 'project' in name_lower: return 'project'
    if any(t in tags_lower for t in ['classification','category','status','type']): return 'classification'
    return 'other'

def _palette_soft():
    return {
        'date': '#D7EBFF',
        'people': '#E9F7D9',
        'project': '#FDEFD0',
        'classification': '#F1E4FF',
        'other': '#D9F2F7',
        'fact': '#FFF2CC',
        'edge': '#8A8A8A',
        'node_edge': '#222222',
        'shadow': (0,0,0,0.08),
        'edge_text': '#333333',
    }

def draw_star_png(
    fact_label, fact_family, spokes, outfile,
    *,
    shape='roundrect',
    two_rings_threshold=None,
    font_scale=1.0,
    dpi=220,
    color_scheme='soft',
):
    """Draws PNG star. Spokes = list of {'dim_label','dim_family','edge_labels'}."""
    try:
        import importlib
        plt = importlib.import_module('matplotlib.pyplot')
        from matplotlib.patches import Circle, Rectangle, FancyBboxPatch
    except Exception:
        return False

    if not spokes:
        return False

    PALETTE = _palette_soft() if color_scheme == 'soft' else _palette_soft()

    f_fact = int(13 * font_scale)
    f_dim  = int(10 * font_scale)
    f_edge = max(8, int(9 * font_scale))

    dims = spokes
    n = len(dims)
    inner_max = None
    if two_rings_threshold and n > two_rings_threshold:
        inner_max = two_rings_threshold
    inner = dims if not inner_max else dims[:inner_max]
    outer = [] if not inner_max else dims[inner_max:]

    inner_radius = 3.4 if len(inner) <= 12 else 3.8
    outer_radius = inner_radius + 1.1 if outer else None

    figsize = (7.5, 7.5)
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111)
    ax.set_aspect('equal'); ax.axis('off')
    cx, cy = 0.0, 0.0

    rect_w, rect_h = 2.7, 1.3
    shadow_offset = 0.09
    shadow = Rectangle((cx-rect_w/2 + shadow_offset, cy-rect_h/2 - shadow_offset),
                       rect_w, rect_h, linewidth=0, facecolor=PALETTE['shadow'], zorder=1)
    ax.add_patch(shadow)
    fact_rect = Rectangle((cx-rect_w/2, cy-rect_h/2), rect_w, rect_h,
                          linewidth=1.6, edgecolor=PALETTE['node_edge'], facecolor=PALETTE['fact'], zorder=2)
    ax.add_patch(fact_rect)
    ax.text(cx, cy, fact_label, ha='center', va='center', fontsize=f_fact, fontweight='bold', zorder=3)

    def ring_positions(count, radius):
        if count <= 0: return []
        angles = [2*math.pi*i/count for i in range(count)]
        return [(cx + radius*math.cos(a), cy + radius*math.sin(a), a) for a in angles]

    inner_pos = ring_positions(len(inner), inner_radius)
    outer_pos = ring_positions(len(outer), outer_radius) if outer else []

    def draw_dim_node(x, y, label, family):
        node_edge = PALETTE['node_edge']
        face = PALETTE.get(family, PALETTE['other'])
        circ_r = 1.06
        if shape == 'circle':
            ax.add_patch(Circle((x+shadow_offset, y-shadow_offset), circ_r, linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
            ax.add_patch(Circle((x, y), circ_r, linewidth=1.6, edgecolor=node_edge, facecolor=face, zorder=2))
        else:
            w, h, r = 2.1, 1.1, 0.35
            ax.add_patch(FancyBboxPatch(
                (x - w/2 + shadow_offset, y - h/2 - shadow_offset), w, h,
                boxstyle=f"round,pad=0.02,rounding_size={r}",
                linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
            ax.add_patch(FancyBboxPatch(
                (x - w/2, y - h/2), w, h,
                boxstyle=f"round,pad=0.02,rounding_size={r}",
                linewidth=1.6, edgecolor=node_edge, facecolor=face, zorder=2))
        ax.text(x, y, label, ha='center', va='center', fontsize=f_dim, zorder=3)

    def draw_spokes(positions, items):
        for (x, y, ang), spec in zip(positions, items):
            draw_dim_node(x, y, spec['dim_label'], spec['dim_family'])
            if shape == 'circle':
                rx, ry = _clip_to_rect(cx, cy, rect_w, rect_h, x, y)
                cx2, cy2 = _clip_to_circle(x, y, 1.06, cx, cy)
            else:
                rx, ry = _clip_to_rect(cx, cy, rect_w, rect_h, x, y)
                cx2, cy2 = _clip_to_circle(x, y, 0.75, cx, cy)
            ax.plot([rx, cx2], [ry, cy2], color=PALETTE['edge'], linewidth=1.3, zorder=1)
            if spec['edge_labels']:
                label = ', '.join([l for l in spec['edge_labels'] if l])
                midx = rx*0.35 + cx2*0.65
                midy = ry*0.35 + cy2*0.65
                dx, dy = cx2 - rx, cy2 - ry
                L = math.hypot(dx, dy) or 1.0
                off = 0.17
                ex = midx - off * dy / L
                ey = midy + off * dx / L
                if len(label) > 32:
                    parts, cur, total = [], [], 0
                    for tok in label.split(', '):
                        if total + len(tok) > 28 and cur:
                            parts.append(', '.join(cur)); cur = [tok]; total = len(tok)
                        else:
                            cur.append(tok); total += (len(tok) + 2)
                    if cur: parts.append(', '.join(cur))
                    label = '\n'.join(parts)
                ax.text(ex, ey, label, ha='center', va='center', fontsize=f_edge, color=PALETTE['edge_text'], zorder=3)

    draw_spokes(inner_pos, inner)
    if outer: draw_spokes(outer_pos, outer)

    fig.tight_layout(pad=0.55)
    fig.savefig(outfile, dpi=dpi)
    plt.close(fig)
    return True


# ----------------------------
# Main
# ----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true')
    parser.add_argument('--materializations', type=str, default='')
    parser.add_argument('--schemas', type=str, default='')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='')
    parser.add_argument('--exclude-sheet-tags', type=str, default='')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='')
    parser.add_argument('--star-diagrams', action='store_true')

    # Diagram flags
    parser.add_argument('--diagram-shape', choices=['roundrect','circle'], default='roundrect')
    parser.add_argument('--diagram-dedupe-roles', action='store_true')
    parser.add_argument('--diagram-two-rings', type=int, default=None)
    parser.add_argument('--diagram-font-scale', type=float, default=1.0)
    parser.add_argument('--diagram-dpi', type=int, default=240)
    parser.add_argument('--diagram-color-scheme', choices=['soft'], default='soft')
    parser.add_argument('--diagram-legend', action='store_true')

    args = parser.parse_args()

    manifest = load_json(args.manifest); catalog = load_json(args.catalog)

    # Filter models
    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    # Maps
    model_kinds = {}
    alias_to_node = {}
    name_to_alias = {}
    alias_to_display = {}

    for n in filtered:
        alias_display = (n.get('alias') or n.get('name') or '')
        alias_lower = alias_display.lower()
        name_lower  = (n.get('name') or '').lower()
        model_kinds[alias_lower] = classify_model_kind(alias_lower, n.get('tags', []))
        alias_to_node[alias_lower] = n
        alias_to_display[alias_lower] = alias_display
        name_to_alias[alias_lower] = alias_lower
        if name_lower:
            name_to_alias[name_lower] = alias_lower

    def canon(token: str):
        if not token: return None
        t = str(token).strip().strip('"').lower()
        return name_to_alias.get(t, t)

    # Relationships
    raw_rel_rows = build_relationship_rows_with_deps(manifest)
    norm_rel_rows = []
    for r in raw_rel_rows:
        dep_aliases = [canon(x) for x in (r.get('DepModels') or [])]
        to_alias    = canon(r.get('ToModel'))
        from_alias = None
        for d in dep_aliases:
            if d and d != to_alias:
                from_alias = d; break
        if not from_alias and dep_aliases:
            from_alias = dep_aliases[0]
        if not to_alias or not from_alias:
            continue
        norm_rel_rows.append({
            'FromModel': from_alias,
            'FromColumn': r.get('FromColumn'),
            'ToModel': to_alias,
            'ToColumn': r.get('ToColumn'),
            'TestName': r.get('TestName'),
        })

    # Spokes per fact
    spokes_raw = defaultdict(list)  # center -> list[(dim, fk)]
    for r in norm_rel_rows:
        frm = r['FromModel']; to = r['ToModel']; col = r.get('FromColumn')
        if not frm or not to: continue
        if to not in alias_to_display:  # skip WIP dims
            continue
        spokes_raw[frm].append((to, col))

    # YAML overlays
    yaml_overlays = load_yaml_overlays(args.schemas)

    # Exclusions
    ex_prefixes = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags = parse_csv_set(args.exclude_sheet_tags)
    ex_mats = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    # Excel formatting
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XLImage

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    col_widths = {"A": 28, "B": 20, "C": 10, "D": 80, "E": 28, "F": 6, "G": 6, "H": 12}

    # Summary / Relationships / Star Map
    summary_rows = []
    for alias_lower, n in alias_to_node.items():
        db, schema, alias_disp = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias_lower, n.get('tags', []))
        summary_rows.append({'Model': alias_disp,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias_disp}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    df_summary = pd.DataFrame(summary_rows, columns=['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description'])
    if not df_summary.empty:
        df_summary = df_summary.sort_values(['Kind','Model'])

    df_rel = pd.DataFrame([
        {'FromModel': alias_to_display.get(r['FromModel'], r['FromModel']),
         'FromColumn': r.get('FromColumn'),
         'ToModel': alias_to_display.get(r['ToModel'], r['ToModel']),
         'ToColumn': r.get('ToColumn'),
         'TestName': r.get('TestName')}
        for r in norm_rel_rows
    ]) if norm_rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])

    star_rows = []
    for center, pairs in spokes_raw.items():
        for (d, _) in pairs:
            star_rows.append({'FactModel': alias_to_display.get(center, center),
                              'DimensionModel': alias_to_display.get(d, d)})
    df_star = pd.DataFrame(star_rows) if star_rows else pd.DataFrame(columns=['FactModel','DimensionModel'])

    used_sheet_names = set()

    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        if not df_rel.empty: df_rel.to_excel(writer, index=False, sheet_name='Relationships')
        if not df_star.empty: df_star.to_excel(writer, index=False, sheet_name='Star Map')

        # Per-model tabs
        for alias_lower, n in alias_to_node.items():
            db, schema, alias_disp = model_relation_identifiers(n)
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''
            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            sheet_name = safe_sheet_name(alias_disp or 'Model', used_sheet_names)

            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            catalog_cols = get_catalog_columns_for_model(catalog, n)
            overlay_entry = (yaml_overlays.get(alias_lower, {}) or {})
            overlay_cols = overlay_entry.get('columns', {}) or {}
            model_meta = overlay_entry.get('model_meta', {}) or {}
            pk_from_model = set(model_meta.get('primary_key', []) or [])
            sk_from_model = set(model_meta.get('surrogate_key', []) or [])
            bk_from_model = set(model_meta.get('business_key', []) or [])
            explicit_pk_names = {*pk_from_model, *sk_from_model}

            rows = []
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or (meta.get('type') or '')
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            info_rows = [
                {'Column': '#Model','DataType': alias_disp,'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Kind','DataType': classify_model_kind(alias_lower, n.get('tags', [])),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Materialization','DataType': (n.get('config') or {}).get('materialized'),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Relation','DataType': f'{db}.{schema}.{alias_disp}','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Tags','DataType': ','.join(n.get('tags') or []),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#SurrogateKey','DataType': ','.join(sorted(sk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#BusinessKey','DataType': ','.join(sorted(bk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#PrimaryKey','DataType': ','.join(sorted(pk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Description','DataType': (n.get('description') or ''),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            df_meta.to_excel(writer, index=False, sheet_name=sheet_name, header=False, startrow=0)
            start_row = len(df_meta.index) + 1
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

            wb = writer.book; ws = wb[sheet_name]
            ws.freeze_panes = f"A{start_row+2}"
            hdr = start_row + 1
            for col_idx in range(1, 8+1):
                c = ws.cell(row=hdr, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = Alignment(vertical="center")
            for r in range(hdr+1, ws.max_row+1):
                ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        # Diagram tabs
        if args.star_diagrams:
            tmpdir = tempfile.mkdtemp(prefix="dbt_star_")
            wb = writer.book
            used = set(wb.sheetnames)

            for center_alias_lower, pairs in sorted(spokes_raw.items()):
                # Build spokes (dedupe optional)
                if args.diagram_dedupe_roles:
                    by_dim = OrderedDict()
                    for dim_alias_lower, fk_col in pairs:
                        by_dim.setdefault(dim_alias_lower, []).append(fk_col or '')
                    spokes = []
                    for dim_alias_lower, fk_cols in by_dim.items():
                        dn = alias_to_node.get(dim_alias_lower)
                        tags_lower = {t.lower() for t in (dn.get('tags') or [])} if dn else set()
                        family = _family_for_dim(dim_alias_lower, tags_lower)
                        spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                       'dim_family': family,
                                       'edge_labels': [c for c in fk_cols if c]})
                else:
                    spokes = []
                    for dim_alias_lower, fk_col in pairs:
                        dn = alias_to_node.get(dim_alias_lower)
                        tags_lower = {t.lower() for t in (dn.get('tags') or [])} if dn else set()
                        family = _family_for_dim(dim_alias_lower, tags_lower)
                        spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                       'dim_family': family,
                                       'edge_labels': [fk_col] if fk_col else []})

                if not spokes:
                    continue

                fact_label = alias_to_display.get(center_alias_lower, center_alias_lower)
                img_path = os.path.join(tmpdir, f"{center_alias_lower}.png")

                drew = draw_star_png(
                    fact_label, 'other', spokes, img_path,
                    shape=args.diagram_shape,
                    two_rings_threshold=args.diagram_two_rings,
                    font_scale=args.diagram_font_scale,
                    dpi=args.diagram_dpi,
                    color_scheme=args.diagram_color_scheme,
                )
                if not drew:
                    continue

                sheet_name = safe_sheet_name(f"Star: {fact_label}", used)
                ws = wb.create_sheet(title=sheet_name)
                try:
                    img = XLImage(img_path)
                    ws.add_image(img, "A1")
                except Exception:
                    pass

                if args.diagram_legend:
                    ws["J2"] = ("Legend: node colors by family "
                                "(Date, People, Project, Classification, Other).\n"
                                "Spoke text = fact FK column(s).")
                    ws["J2"].alignment = Alignment(wrap_text=True)

    print(f"Wrote: {args.out}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
