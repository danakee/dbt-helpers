#!/usr/bin/env python3
r"""
dbt_to_kimball_excel.py  (v1.7.4)

Adds better star diagrams:
- Lines clip to the circle/rectangle edges (not to centers).
- Uses original alias casing from manifest (no forced Title Case).
- Labels each spoke with the FACT-side FK column name.
- Visual styling tweaks (shadows, thicker strokes, spacing).
Keeps v1.7.3 behavior for relationships detection, aliases, and graceful skips.
"""

import argparse, json, re, sys, fnmatch, math, tempfile, os
from collections import defaultdict
from pathlib import Path

import pandas as pd
try:
    import yaml
except Exception:
    yaml = None


# ----------------------------
# Helpers (unchanged basics)
# ----------------------------
def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node):
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags):
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', base or 'Sheet')[:31] or 'Sheet'
    candidate = cleaned; i = 1
    while candidate in used:
        suffix = f'_{i}'; candidate = (cleaned[:31-len(suffix)] + suffix); i += 1
    used.add(candidate); return candidate

def _normalize_to_model(raw):
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"""ref\(\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    m = re.match(r"""source\(\s*['"][^'"]+['"]\s*,\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    return s

def extract_tests_for_model(manifest, node_id):
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test':
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends:
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_model = _normalize_to_model(raw_to)
            rel_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')):
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    db, schema, alias = model_relation_identifiers(node)
    key = f'{db}.{schema}.{alias}'.lower()
    for k, v in (catalog.get('nodes') or {}).items():
        if k.lower() == key:
            return v.get('columns') or {}
    return {}

def build_relationship_rows_with_deps(manifest):
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test':
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname:
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_model = _normalize_to_model(raw_to)
        to_field = kwargs.get('field') or kwargs.get('to_field') or kwargs.get('column')
        dep_models = []
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model':
                dep_models.append((n.get('alias') or n.get('name')))
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'DepModels': dep_models,'ToModel': to_model,'ToColumn': to_field,
                     'FromModel': None,'FromColumn': from_column,'TestName': tname})
    return rows

def _as_list(v):
    if v is None: return []
    if isinstance(v, (list, tuple)): return [str(x) for x in v]
    return [str(v)]

def load_yaml_overlays(paths_csv: str):
    overlays = {}
    if not paths_csv or not yaml: return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc: continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname: continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})
                entry['model_meta'] = {
                    'surrogate_key': [s.lower() for s in _as_list(mm.get('surrogate_key'))],
                    'business_key':  [s.lower() for s in _as_list(mm.get('business_key'))],
                    'primary_key':   [s.lower() for s in _as_list(mm.get('primary_key'))],
                }
                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {'data_type': dtype, 'nullable': nullable, 'description': desc}.items() if v is not None}
    return overlays

def parse_csv_set(s: str):
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}

def should_exclude_sheet(alias_lower: str, tags_lower: set, mat_lower: str, path_str: str,
                         prefixes: set, tags_excl: set, mats_excl: set, path_globs: set) -> bool:
    if any(alias_lower.startswith(p) for p in prefixes): return True
    if any(t in tags_lower for t in tags_excl): return True
    if mat_lower in mats_excl: return True
    ps = (path_str or '')
    for g in path_globs:
        if fnmatch.fnmatch(ps.lower(), g): return True
    return False


# ---------- Geometry helpers for diagram ----------
def _clip_to_circle(cx, cy, r, x, y):
    """Return point on circle perimeter from (cx,cy) toward (x,y)."""
    dx, dy = x - cx, y - cy
    d = math.hypot(dx, dy) or 1e-6
    return cx + r * dx / d, cy + r * dy / d

def _clip_to_rect(cx, cy, w, h, x, y):
    """
    Clip a line from rect center (cx,cy) to point (x,y) so it exits at rect boundary.
    Rect is axis-aligned with width w and height h.
    """
    dx, dy = x - cx, y - cy
    if dx == 0 and dy == 0:
        return cx, cy
    # scale to the first rectangle side hit
    tx = (w/2) / abs(dx) if dx != 0 else float('inf')
    ty = (h/2) / abs(dy) if dy != 0 else float('inf')
    t = min(tx, ty)
    return cx + dx * t, cy + dy * t


# ---------- Diagram drawing with labels & styling ----------
def draw_star_png(fact_label, spokes, outfile, figsize=(7.2,7.2)):
    """
    fact_label: display text for the fact
    spokes: list of (dim_label, edge_label)  # edge_label = fact FK column
    """
    try:
        import importlib
        plt = importlib.import_module('matplotlib.pyplot')
        from matplotlib.patches import Circle, Rectangle
    except Exception:
        return False

    if not spokes:
        return False

    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111)
    ax.set_aspect('equal'); ax.axis('off')

    # layout
    center = (0.0, 0.0)
    n = len(spokes)
    radius = 3.4 if n <= 6 else 3.8 if n <= 10 else 4.2
    angles = [2*math.pi*i/n for i in range(n)]
    positions = [(center[0]+radius*math.cos(a), center[1]+radius*math.sin(a)) for a in angles]

    # style
    line_color = "#8a8a8a"
    node_edge = "#1f1f1f"
    dim_fill = "#D9F2F7"
    fact_fill = "#FFF2CC"

    # fact rectangle (with subtle shadow)
    rect_w, rect_h = 2.6, 1.25
    fx, fy = center
    shadow_offset = 0.08
    shadow = Rectangle((fx-rect_w/2 + shadow_offset, fy-rect_h/2 - shadow_offset),
                       rect_w, rect_h, linewidth=0, facecolor="#000000", alpha=0.08, zorder=1)
    ax.add_patch(shadow)
    fact_rect = Rectangle((fx-rect_w/2, fy-rect_h/2), rect_w, rect_h,
                          linewidth=1.6, edgecolor=node_edge, facecolor=fact_fill, zorder=2)
    ax.add_patch(fact_rect)
    ax.text(fx, fy, fact_label, ha='center', va='center', fontsize=13, fontweight='bold', zorder=3)

    # draw spokes
    circle_r = 1.05
    for (x, y), (dim_label, edge_label) in zip(positions, spokes):
        # circle with shadow
        cshadow = Circle((x+shadow_offset, y-shadow_offset), circle_r, linewidth=0, facecolor="#000000", alpha=0.08, zorder=1)
        ax.add_patch(cshadow)
        circ = Circle((x, y), circle_r, linewidth=1.6, edgecolor=node_edge, facecolor=dim_fill, zorder=2)
        ax.add_patch(circ)

        # compute clipped line endpoints
        # from fact rect edge to circle edge
        rx, ry = _clip_to_rect(fx, fy, rect_w, rect_h, x, y)
        cx, cy = _clip_to_circle(x, y, circle_r, fx, fy)
        ax.plot([rx, cx], [ry, cy], color=line_color, linewidth=1.3, zorder=1)

        # labels
        ax.text(x, y, dim_label, ha='center', va='center', fontsize=11, zorder=3)

        # edge label at 60% along the segment, offset a bit
        midx = rx * 0.4 + cx * 0.6
        midy = ry * 0.4 + cy * 0.6
        # perpendicular offset for readability
        dx, dy = cx - rx, cy - ry
        L = math.hypot(dx, dy) or 1.0
        off = 0.15
        ex = midx - off * dy / L
        ey = midy + off * dx / L
        if edge_label:
            ax.text(ex, ey, edge_label, ha='center', va='center', fontsize=9, color="#333333", zorder=3)

    fig.tight_layout(pad=0.5)
    fig.savefig(outfile, dpi=180)
    plt.close(fig)
    return True


# ----------------------------
# Main (relationship logic & Excel generation)
# ----------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true')
    parser.add_argument('--materializations', type=str, default='')
    parser.add_argument('--schemas', type=str, default='')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='')
    parser.add_argument('--exclude-sheet-tags', type=str, default='')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='')
    parser.add_argument('--star-diagrams', action='store_true')

    args = parser.parse_args()

    manifest = load_json(args.manifest); catalog = load_json(args.catalog)

    # Filter models
    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    # Maps: alias_lower -> node ; name/alias -> alias_lower ; alias_lower -> display (original case)
    model_kinds = {}
    alias_to_node = {}
    name_to_alias = {}
    alias_to_display = {}

    for n in filtered:
        alias_display = (n.get('alias') or n.get('name') or '')
        alias_lower = alias_display.lower()
        name_lower  = (n.get('name') or '').lower()

        model_kinds[alias_lower] = classify_model_kind(alias_lower, n.get('tags', []))
        alias_to_node[alias_lower] = n
        alias_to_display[alias_lower] = alias_display  # preserve case
        name_to_alias[alias_lower] = alias_lower
        if name_lower:
            name_to_alias[name_lower] = alias_lower

    def canon(token: str):
        if not token: return None
        t = str(token).strip().strip('"').lower()
        return name_to_alias.get(t, t)

    # Build relationship rows with correct FROM selection
    raw_rel_rows = build_relationship_rows_with_deps(manifest)
    norm_rel_rows = []
    for r in raw_rel_rows:
        dep_aliases = [canon(x) for x in (r.get('DepModels') or [])]
        to_alias    = canon(r.get('ToModel'))
        from_alias = None
        for d in dep_aliases:
            if d and d != to_alias:
                from_alias = d; break
        if not from_alias and dep_aliases:
            from_alias = dep_aliases[0]
        if not to_alias or not from_alias:
            continue
        norm_rel_rows.append({
            'FromModel': from_alias,
            'FromColumn': r.get('FromColumn'),
            'ToModel': to_alias,
            'ToColumn': r.get('ToColumn'),
            'TestName': r.get('TestName'),
        })

    # Build star spokes: center -> list[(dim, from_column)]
    star_spokes = defaultdict(list)
    for r in norm_rel_rows:
        frm = r['FromModel']; to = r['ToModel']; col = r.get('FromColumn')
        if not frm or not to:
            continue
        # Only keep dimensions we know about (skip WIP/missing quietly)
        if to not in alias_to_display:
            continue
        star_spokes[frm].append((to, col))

    # YAML overlays (used elsewhere for sheets)
    yaml_overlays = load_yaml_overlays(args.schemas)

    # Exclusions
    ex_prefixes = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags = parse_csv_set(args.exclude_sheet_tags)
    ex_mats = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    # Formatting setup for Excel
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XLImage

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    col_widths = {"A": 28, "B": 20, "C": 10, "D": 80, "E": 28, "F": 6, "G": 6, "H": 12}

    # Summary / Relationships / Star Map sheets
    summary_rows = []
    for alias_lower, n in alias_to_node.items():
        db, schema, alias_disp = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias_lower, n.get('tags', []))
        summary_rows.append({'Model': alias_disp,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias_disp}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    df_summary = pd.DataFrame(summary_rows, columns=['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description'])
    if not df_summary.empty:
        df_summary = df_summary.sort_values(['Kind','Model'])

    df_rel = pd.DataFrame([
        {'FromModel': alias_to_display.get(r['FromModel'], r['FromModel']),
         'FromColumn': r.get('FromColumn'),
         'ToModel': alias_to_display.get(r['ToModel'], r['ToModel']),
         'ToColumn': r.get('ToColumn'),
         'TestName': r.get('TestName')}
        for r in norm_rel_rows
    ]) if norm_rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])

    star_rows = []
    for center, dims in star_spokes.items():
        for (d, _) in dims:
            star_rows.append({'FactModel': alias_to_display.get(center, center),
                              'DimensionModel': alias_to_display.get(d, d)})
    df_star = pd.DataFrame(star_rows) if star_rows else pd.DataFrame(columns=['FactModel','DimensionModel'])

    used_sheet_names = set()

    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        if not df_rel.empty: df_rel.to_excel(writer, index=False, sheet_name='Relationships')
        if not df_star.empty: df_star.to_excel(writer, index=False, sheet_name='Star Map')

        # per-model tabs (same as earlier versions)
        for alias_lower, n in alias_to_node.items():
            db, schema, alias_disp = model_relation_identifiers(n)
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''

            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            sheet_name = safe_sheet_name(alias_disp or 'Model', used_sheet_names)

            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            catalog_cols = get_catalog_columns_for_model(catalog, n)
            overlay_entry = (yaml_overlays.get(alias_lower, {}) or {})
            overlay_cols = overlay_entry.get('columns', {}) or {}
            model_meta = overlay_entry.get('model_meta', {}) or {}
            pk_from_model = set(model_meta.get('primary_key', []) or [])
            sk_from_model = set(model_meta.get('surrogate_key', []) or [])
            bk_from_model = set(model_meta.get('business_key', []) or [])
            explicit_pk_names = {*pk_from_model, *sk_from_model}

            rows = []
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or (meta.get('type') or '')
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            info_rows = [
                {'Column': '#Model','DataType': alias_disp,'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Kind','DataType': classify_model_kind(alias_lower, n.get('tags', [])),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Materialization','DataType': (n.get('config') or {}).get('materialized'),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Relation','DataType': f'{db}.{schema}.{alias_disp}','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Tags','DataType': ','.join(n.get('tags') or []),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#SurrogateKey','DataType': ','.join(sorted(sk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#BusinessKey','DataType': ','.join(sorted(bk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#PrimaryKey','DataType': ','.join(sorted(pk_from_model)) or '','Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
                {'Column': '#Description','DataType': (n.get('description') or ''),'Nullable?':'','Description': '','Tests': '','IsPK?': '','IsFK?': '','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            df_meta.to_excel(writer, index=False, sheet_name=sheet_name, header=False, startrow=0)
            start_row = len(df_meta.index) + 1
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

            wb = writer.book; ws = wb[sheet_name]
            ws.freeze_panes = f"A{start_row+2}"
            hdr = start_row + 1
            for col_idx in range(1, 8+1):
                c = ws.cell(row=hdr, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = Alignment(vertical="center")
            for r in range(hdr+1, ws.max_row+1):
                ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        # Star diagram tabs
        if args.star_diagrams:
            tmpdir = tempfile.mkdtemp(prefix="dbt_star_")
            wb = writer.book
            used = set(wb.sheetnames)
            for center_alias_lower, pairs in sorted(star_spokes.items()):
                # Convert to display strings & group
                spokes = []
                for dim_alias_lower, fk_col in pairs:
                    dim_label = alias_to_display.get(dim_alias_lower, dim_alias_lower)
                    edge_label = fk_col or ''
                    spokes.append((dim_label, edge_label))
                if not spokes:
                    continue
                fact_label = alias_to_display.get(center_alias_lower, center_alias_lower)
                img_path = os.path.join(tmpdir, f"{center_alias_lower}.png")
                if not draw_star_png(fact_label, spokes, img_path, figsize=(7.2,7.2)):
                    continue
                sheet_name = safe_sheet_name(f"Star: {fact_label}", used)
                ws = wb.create_sheet(title=sheet_name)
                try:
                    img = XLImage(img_path)
                    ws.add_image(img, "A1")
                except Exception:
                    pass

    print(f"Wrote: {args.out}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
