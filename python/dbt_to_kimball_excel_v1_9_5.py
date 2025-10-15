#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
dbt_to_kimball_excel.py  —  Generate a Kimball-style Excel catalog + star diagrams
==================================================================================

Version: 1.9.5  # CHANGE: bumped version to reflect edits

What this does
--------------
Turns a compiled **dbt Core** project into a Kimball-style Excel workbook with:

1) Summary tab
   • One row per model: Kind, Materialization, Relation, FQN, Tags, Description.
   • Frozen header, blue header bg, bold header, auto-filter, wrapped long text.

2) (Optional) Relationships tab  [default: included — disable with --no-relationships]
   • One row per dbt `relationships` test (FromModel/FromColumn → ToModel/ToColumn).
   • Mirrors exactly what dbt tests assert.

3) Star Map tab
   • Aggregated rows of (FactModel, DimensionModel, FactFKColumn(s), DimKeyColumn(s)).
   • This is the canonical mapping that the diagrams use (after options).

4) One tab per model (Dimension/Fact/Other)
   • **Table meta block** at the top (labels **without '#'** and **bold**):
       Model, Kind, Materialization, Relation, Tags, SurrogateKey, BusinessKey, PrimaryKey, Description
     - Surrogate/Business/Primary use **original casing** from your YAML.
     - If PrimaryKey is missing in YAML, we **fall back to SurrogateKey** for display.
   • **Column grid**: Column, DataType, Nullable?, Description, Tests, IsPK?, IsFK?, Source
     - Heuristic PK/FK from tests; Nullable? from not_null; YAML overlay for data_type/nullable/description.
   • Frozen column header, blue header + bold, auto-filter, wrapped Description/Tests, “green-bar” banding.

5) (Optional) Star-diagram tabs (PNG via Matplotlib, embedded in new sheets)
   • Per fact model: fact = rounded yellow rectangle, dims = rounded **light green** rectangles.
   • **Bold** labels for both fact and dimensions.
   • FK column names printed along spokes (wrapped if long).
   • Two-ring layout if many dimensions (configurable).
   • Sheet titles like **“Star-Fact FactOrderLine”** (safe & ≤31 chars).

Sheet tab style
---------------
We can’t change the rounded “pill” shape from Excel UI, but we do color tabs:
  - Summary/Relationships/Star Map: blue
  - Dimension model tabs: purple
  - Fact model tabs: peach
  - Diagram tabs: green

Inputs
------
• --manifest  target/manifest.json   (required)
• --catalog   target/catalog.json    (required)
• --schemas   CSV/globs of schema YAMLs (optional) to overlay:
    - Reads model-level `meta.surrogate_key`, `meta.business_key`, `meta.primary_key`
      (keeps original case for display; also stores lowercase for matching).
    - Reads column-level `meta.data_type`, `meta.nullable`, and `description`.

Install
-------
    pip install pandas openpyxl pyyaml matplotlib

Typical use
-----------
    python scripts/dbt_to_kimball_excel.py ^
      --manifest .\target\manifest.json ^
      --catalog  .\target\catalog.json ^
      --schemas ".\models\mart\schemas.yml,.\models\stage\schemas.yml" ^
      --exclude-sheet-prefixes base_,stage_ ^
      --exclude-sheet-tags stage,logging ^
      --exclude-sheet-materializations view ^
      --exclude-sheet-path-globs "models/base/*,models/stage/*,models/logging/*" ^
      --star-diagrams --diagram-two-rings 12 --diagram-font-scale 0.95 --diagram-dpi 220 ^
      --out .\documentation\kimball_model_catalog.xlsx

Important flags & behavior
--------------------------
• --include-views + --materializations TABLE,INCREMENTAL: combined filter logic for per-model tabs.
• Exclude per-model tabs via:
  --exclude-sheet-prefixes / --exclude-sheet-tags / --exclude-sheet-materializations / --exclude-sheet-path-globs
• --no-relationships            : omit the Relationships sheet. (Star Map & diagrams still generated)
• --include-missing-dims        : keep Star Map rows whose target dim isn’t in manifest (suffix “(missing)”).
                                  Diagrams still skip missing dims (robust).
• --star-diagrams               : enable diagram tabs.
• --diagram-two-rings N         : inner ring up to N dims, rest outer ring.
• --diagram-no-wrap-labels      : disable label wrapping in diagrams.
• --diagram-max-label-chars INT : wrapping width per line (default 18).

Limitations / assumptions
-------------------------
• Relationship edges are derived from dbt `relationships` tests in `manifest.json`.
• If a test references a dimension not present in the manifest:
  - Default: skip in Star Map & diagrams.
  - With --include-missing-dims: include in Star Map (suffix “(missing)”); diagrams still skip.

Changelog highlights
--------------------
1.9.5
  - Fixed Excel column letter access via `openpyxl.utils` (version-robust).
  - Seeded sheet-name de-dup set with existing sheetnames to avoid collisions.
  - Fallback for catalog key matching without database part.
  - Consistent relationships kwargs parsing (`to_field` preference).
  - Optional fact-sheet note listing missing dimensions when `--include-missing-dims`.
  - Temporary diagram directory now auto-cleans with `TemporaryDirectory`.
  - Warn if `--schemas` provided but PyYAML is not installed.

1.9.4
  - Colored sheet tabs by type (Summary/Relationships/Star Map = blue; Dimensions = purple; Facts = peach; Diagrams = green).
  - Fixed “_1” suffix on sheet names (no double-safe-naming; case-insensitive uniqueness).
  - Kept all 1.9.3 formatting and diagram improvements (bold labels, single dim color, etc.).

1.9.3
  - Thorough comments; polished overview tab formatting; diagram label wrapping at camel boundaries.

"""

import argparse, json, re, sys, fnmatch, math, tempfile, os
from collections import defaultdict, OrderedDict
from pathlib import Path

import pandas as pd
try:
    import yaml
except Exception:
    yaml = None


# =============================================================================
# Basic helpers
# =============================================================================

def load_json(path: Path):
    """Open and parse a JSON file with UTF-8 encoding."""
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)

def node_is_model(node: dict) -> bool:
    """True if a manifest node is a dbt model."""
    return node.get('resource_type') == 'model'

def classify_model_kind(name: str, tags) -> str:
    """Heuristic to tag a model as Fact / Dimension / Other when tags/names help."""
    name_lower = (name or '').lower()
    tags_lower = {t.lower() for t in (tags or [])}
    if any(t in tags_lower for t in ['fact','facts']): return 'Fact'
    if any(t in tags_lower for t in ['dim','dimension','dimensions']): return 'Dimension'
    if name_lower.startswith('fact_') or name_lower.endswith('_fact'): return 'Fact'
    if name_lower.startswith('dim_') or name_lower.endswith('_dim'): return 'Dimension'
    return 'Other'

def safe_sheet_name(base: str, used: set):
    """
    Return a safe, <=31-char sheet name unique in workbook.
    Uniqueness is case-insensitive (Excel treats names case-insensitively).
    """
    raw = (base or 'Sheet').strip()
    cleaned = re.sub(r'[:\\\/\?\*\[\]]', '_', raw)[:31] or 'Sheet'
    used_lower = {u.lower() for u in used}
    candidate = cleaned; i = 1
    while candidate.lower() in used_lower:
        suffix = f'_{i}'
        candidate = (cleaned[:31 - len(suffix)] + suffix)
        i += 1
    used.add(candidate)  # keep original casing in set
    return candidate

def _normalize_to_model(raw):
    """Normalize strings like ref('DimUser') to 'DimUser'; otherwise return as-is."""
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"""ref\(\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    m = re.match(r"""source\(\s*['"][^'"]+['"]\s*,\s*['"]([^'"]+)['"]\s*\)""", s, re.IGNORECASE)
    if m: return m.group(1)
    return s

def extract_tests_for_model(manifest, node_id):
    """
    Build: { column_name -> [test, ...] } for tests that depend_on this model.
    Stores relationship details when present.
    """
    results = defaultdict(list)
    for _, test in manifest.get('nodes', {}).items():
        if test.get('resource_type') != 'test':
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        if node_id not in depends:
            continue
        meta = test.get('test_metadata') or {}; kwargs = meta.get('kwargs') or {}
        col = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column') or None
        test_name = (meta.get('name') or test.get('name') or 'test').lower()
        rel = None
        if 'relationship' in test_name or 'relationships' in test_name:
            raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
            rel_model = _normalize_to_model(raw_to)
            # CHANGE: prefer to_field, then field, then column for the TO side
            rel_field = kwargs.get('to_field') or kwargs.get('field') or kwargs.get('column')
            rel = {'to_model': rel_model, 'to_field': rel_field}
        results[col].append({'name': test_name, 'details': rel, 'raw': test})
    return results

def infer_pk_fk_and_nullable(columns_tests):
    """
    Heuristics to mark columns as PK/FK/Nullable based on presence of tests.
    - PK: has unique + not_null (or unique and name looks like id/key).
    - FK: has relationships test.
    - Nullable? is '' (blank) if not_null present, else 'Y' (conservative).
    """
    inferred = {}
    for col, tests in columns_tests.items():
        names = [t['name'] for t in tests]
        is_unique = any('unique' in n for n in names)
        is_not_null = any(n in ('not_null','not-null','not null') or 'not_null' in n for n in names)
        has_rel = any('relationship' in n or 'relationships' in n for n in names)
        is_pk = bool(is_unique and is_not_null)
        if not is_pk and col:
            cl = col.lower()
            if is_unique and (cl.endswith('key') or cl.endswith('id')):
                is_pk = True
        inferred[col] = {'is_pk': is_pk, 'is_fk': has_rel, 'nullable_from_tests': ('' if is_not_null else 'Y')}
    return inferred

def model_relation_identifiers(node):
    """Return (db, schema, alias) for a model; alias reflects the final object name."""
    db = (node.get('database') or node.get('schema') or '').strip('"')
    schema = (node.get('schema') or '').strip('"')
    alias = (node.get('alias') or node.get('name') or '').strip('"')
    return db, schema, alias

def get_catalog_columns_for_model(catalog, node):
    """Find the matching catalog entry for a node and return its columns dict."""
    db, schema, alias = model_relation_identifiers(node)
    key_db = f'{db}.{schema}.{alias}'.lower()
    key_no_db = f'{schema}.{alias}'.lower()  # CHANGE: fallback key without database
    for k, v in (catalog.get('nodes') or {}).items():
        kl = k.lower()
        if kl == key_db or kl == key_no_db:  # CHANGE
            return v.get('columns') or {}
    return {}

def build_relationship_rows_with_deps(manifest):
    """
    Pull relationship tests and capture both:
      - DepModels: models in test.depends_on.nodes (used to choose FromModel)
      - ToModel / ToColumn: relationship target
      - FromColumn        : fk column on the fact
    """
    rows = []; nodes = manifest.get('nodes', {})
    for _, test in nodes.items():
        if test.get('resource_type') != 'test':
            continue
        meta = test.get('test_metadata') or {}
        tname = (meta.get('name') or test.get('name') or '').lower()
        if 'relationship' not in tname and 'relationships' not in tname:
            continue
        depends = (test.get('depends_on') or {}).get('nodes') or []
        kwargs = meta.get('kwargs') or {}
        raw_to = kwargs.get('to') or kwargs.get('model') or kwargs.get('to_model')
        to_model = _normalize_to_model(raw_to)
        # CHANGE: prefer to_field first for the TO side
        to_field = kwargs.get('to_field') or kwargs.get('field') or kwargs.get('column')
        dep_models = []
        for nid in depends:
            n = nodes.get(nid) or {}
            if n.get('resource_type') == 'model':
                dep_models.append((n.get('alias') or n.get('name')))
        # CHANGE: prefer column_name for the FROM side
        from_column = kwargs.get('column_name') or kwargs.get('field') or kwargs.get('column')
        rows.append({'DepModels': dep_models,'ToModel': to_model,'ToColumn': to_field,
                     'FromModel': None,'FromColumn': from_column,'TestName': tname})
    return rows

def parse_csv_set(s: str):
    """CSV string -> set of lowercased tokens; empty on falsy input."""
    return {t.strip().lower() for t in (s.split(',') if s else []) if t.strip()}


# =============================================================================
# YAML overlays (preserve display casing + lowercase for matching)
# =============================================================================

def load_yaml_overlays(paths_csv: str):
    """
    Build:
      overlays[model_name_lower] = {
        'model_meta_display': {'surrogate_key': [OrigCase...], 'business_key': [...], 'primary_key': [...] },
        'model_meta_match':   {'surrogate_key': [lower...],      'business_key': [...], 'primary_key': [...] },
        'columns': { col_lower: {'data_type':..., 'nullable':..., 'description':...}, ... }
      }
    """
    overlays = {}
    if not paths_csv or not yaml: return overlays
    import glob
    paths = []
    for token in paths_csv.split(','):
        token = token.strip()
        if token:
            paths.extend(glob.glob(token))
    for p in paths:
        try:
            with open(p, 'r', encoding='utf-8') as f:
                doc = yaml.safe_load(f)
        except Exception:
            continue
        if not doc: continue
        for section in ('models','sources','snapshots','semantic_models'):
            for item in (doc.get(section) or []):
                mname = (item.get('name') or '').lower()
                if not mname: continue
                entry = overlays.setdefault(mname, {})
                mm = (item.get('meta') or {})

                def _norm_list(v):
                    if v is None: return []
                    if isinstance(v, (list, tuple)): return [str(x) for x in v]
                    return [str(v)]

                sk_disp = _norm_list(mm.get('surrogate_key'))
                bk_disp = _norm_list(mm.get('business_key'))
                pk_disp = _norm_list(mm.get('primary_key'))

                entry['model_meta_display'] = {
                    'surrogate_key': sk_disp,
                    'business_key':  bk_disp,
                    'primary_key':   pk_disp,
                }
                entry['model_meta_match'] = {
                    'surrogate_key': [s.lower() for s in sk_disp],
                    'business_key':  [s.lower() for s in bk_disp],
                    'primary_key':   [s.lower() for s in pk_disp],
                }

                colmap = entry.setdefault('columns', {})
                for col in (item.get('columns') or []):
                    cname = (col.get('name') or '').lower()
                    if not cname: continue
                    meta = col.get('meta') or {}
                    dtype = meta.get('data_type') or col.get('data_type') or None
                    nullable = meta.get('nullable') if 'nullable' in meta else col.get('nullable')
                    desc = col.get('description') or None
                    colmap[cname] = {k:v for k,v in {'data_type': dtype, 'nullable': nullable, 'description': desc}.items() if v is not None}
    return overlays


# =============================================================================
# Diagram utilities — wrap only at camel-case boundaries (no spaces inserted)
# =============================================================================

_CAMEL_SPLIT_RE = re.compile(
    r'(?<=[A-Za-z])(?=[A-Z][a-z])|(?<=[a-z])(?=[A-Z])|(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])'
)

def _camel_chunks(s: str):
    if not s: return []
    idxs = [0]
    for m in _CAMEL_SPLIT_RE.finditer(s):
        idxs.append(m.start())
    idxs.append(len(s))
    return [s[idxs[i]:idxs[i+1]] for i in range(len(idxs)-1)]

def _wrap_preserving_string(s: str, max_chars: int, max_lines: int = 2):
    """Return up to max_lines with soft breaks at camel boundaries; adds '…' if truncated."""
    chunks = _camel_chunks(s)
    if not chunks:
        return ['']
    lines, i = [], 0
    for _ in range(max_lines):
        cur = ''
        while i < len(chunks):
            cand = cur + chunks[i]
            if len(cand) <= max_chars or cur == '':
                cur = cand; i += 1
            else:
                break
        if cur:
            lines.append(cur)
        else:
            lines.append((chunks[i][:max_chars-1] + '…') if max_chars > 1 else '…'); i += 1
        if i >= len(chunks): break
    if i < len(chunks):  # overflow — elide
        last = lines[-1]
        lines[-1] = (last[:max(1, max_chars-1)] + '…') if len(last) >= max_chars else (last + '…')
    return lines

def _clip_to_circle(cx, cy, r, x, y):
    """Point on circle (x,y) projected to circle boundary center (cx,cy), radius r."""
    dx, dy = x - cx, y - cy
    d = math.hypot(dx, dy) or 1e-6
    return cx + r * dx / d, cy + r * dy / d

def _clip_to_rect(cx, cy, w, h, x, y):
    """Point where a line from center to (x,y) hits rectangle boundary (width w, height h)."""
    dx, dy = x - cx, y - cy
    if dx == 0 and dy == 0:
        return cx, cy
    tx = (w/2) / abs(dx) if dx != 0 else float('inf')
    ty = (h/2) / abs(dy) if dy != 0 else float('inf')
    t = min(tx, ty)
    return cx + dx * t, cy + dy * t

def _palette_soft():
    """Centralized diagram palette for easy retheme."""
    return {
        'fact': '#FFF2CC',         # light yellow
        'dim_uniform': '#DFF5E1',  # single light-green color for all dimension nodes
        'edge': '#8A8A8A',
        'node_edge': '#222222',
        'shadow': (0,0,0,0.08),
        'edge_text': '#333333',
    }

def draw_star_png(
    fact_label, fact_family, spokes, outfile,
    *,
    shape='roundrect',
    two_rings_threshold=None,
    font_scale=1.0,
    dpi=220,
    color_scheme='soft',
    wrap_labels=True,
    max_label_chars=18,
):
    """
    Render a star as a PNG.
    spokes: list of dicts: { 'dim_label': str, 'dim_family': str, 'edge_labels': [fk, ...] }
    Returns True if an image was written; False if Matplotlib missing or no spokes.
    """
    try:
        import importlib
        plt = importlib.import_module('matplotlib.pyplot')
        from matplotlib.patches import FancyBboxPatch
    except Exception:
        return False

    if not spokes:
        return False

    PALETTE = _palette_soft()

    # Font sizes scale coherently with font_scale
    f_fact = int(13 * font_scale)
    f_dim  = int(10 * font_scale)
    f_edge = max(8, int(9 * font_scale))

    # Two-ring logic (inner ring feels balanced up to ~12 dimensions)
    dims = spokes; n = len(dims)
    inner_max = two_rings_threshold if (two_rings_threshold and n > two_rings_threshold) else None
    inner = dims if not inner_max else dims[:inner_max]
    outer = [] if not inner_max else dims[inner_max:]

    inner_radius = 3.4 if len(inner) <= 12 else 3.8
    outer_radius = inner_radius + 1.1 if outer else None

    # Canvas
    figsize = (7.5, 7.5)
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111)
    ax.set_aspect('equal'); ax.axis('off')
    cx, cy = 0.0, 0.0
    shadow_offset = 0.09

    # ---- Fact node: rounded rect, yellow, bold label (wrapped) ----
    fact_lines = (_wrap_preserving_string(fact_label, max_label_chars+4, 2)
                  if wrap_labels else [fact_label])
    longest = max(len(l) for l in fact_lines)
    rect_w = max(2.9, 0.13 * longest + 1.2)
    rect_h = (1.3 if len(fact_lines) == 1 else 1.55 + 0.42*max(0, len(fact_lines)-1))

    ax.add_patch(FancyBboxPatch(
        (cx - rect_w/2 + shadow_offset, cy - rect_h/2 - shadow_offset), rect_w, rect_h,
        boxstyle="round,pad=0.02,rounding_size=0.25",
        linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
    ax.add_patch(FancyBboxPatch(
        (cx - rect_w/2, cy - rect_h/2), rect_w, rect_h,
        boxstyle="round,pad=0.02,rounding_size=0.25",
        linewidth=1.6, edgecolor=PALETTE['node_edge'], facecolor=PALETTE['fact'], zorder=2))

    if len(fact_lines) == 1:
        ax.text(cx, cy, fact_lines[0], ha='center', va='center', fontsize=f_fact, fontweight='bold', zorder=3)
    else:
        y0 = cy + (0.22 if len(fact_lines) == 2 else 0.33)
        for i, line in enumerate(fact_lines):
            ax.text(cx, y0 - i*0.34, line, ha='center', va='center',
                    fontsize=f_fact, fontweight='bold', zorder=3)

    def ring_positions(count, radius):
        """Evenly spaced positions on a circle (center cx,cy; given radius)."""
        if count <= 0: return []
        angles = [2*math.pi*i/count for i in range(count)]
        return [(cx + radius*math.cos(a), cy + radius*math.sin(a), a) for i, a in enumerate(angles)]

    inner_pos = ring_positions(len(inner), inner_radius)
    outer_pos = ring_positions(len(outer), outer_radius) if outer else []

    def draw_dim_node(x, y, label):
        """Rounded rectangle node for a dimension; bold wrapped label."""
        from matplotlib.patches import FancyBboxPatch  # local alias for clarity
        node_edge = PALETTE['node_edge']
        face = PALETTE['dim_uniform']  # single color for all dims
        lines = (_wrap_preserving_string(label, max_label_chars, 2) if wrap_labels else [label])
        longest = max(len(l) for l in lines)
        w = max(2.1, 0.12 * longest + 0.9)
        h = (1.1 if len(lines) == 1 else 1.35)
        r = 0.35

        ax.add_patch(FancyBboxPatch(
            (x - w/2 + shadow_offset, y - h/2 - shadow_offset), w, h,
            boxstyle=f"round,pad=0.02,rounding_size={r}",
            linewidth=0, facecolor=PALETTE['shadow'], zorder=1))
        ax.add_patch(FancyBboxPatch(
            (x - w/2, y - h/2), w, h,
            boxstyle=f"round,pad=0.02,rounding_size={r}",
            linewidth=1.6, edgecolor=node_edge, facecolor=face, zorder=2))

        if len(lines) == 1:
            ax.text(x, y, lines[0], ha='center', va='center', fontsize=f_dim, fontweight='bold', zorder=3)
        else:
            y0 = y + 0.16
            for i, line in enumerate(lines):
                ax.text(x, y0 - i*0.28, line, ha='center', va='center', fontsize=f_dim, fontweight='bold', zorder=3)

        return w, h

    def draw_spokes(positions, items):
        """Draw edges from fact to each dimension and lay down edge labels."""
        for (x, y, _), spec in zip(positions, items):
            w, h = draw_dim_node(x, y, spec['dim_label'])
            rx, ry = _clip_to_rect(cx, cy, rect_w, rect_h, x, y)
            cx2, cy2 = _clip_to_circle(x, y, min(w, h)/2 - 0.02, cx, cy)
            ax.plot([rx, cx2], [ry, cy2], color=PALETTE['edge'], linewidth=1.3, zorder=1)
            if spec['edge_labels']:
                label = ', '.join([l for l in spec['edge_labels'] if l])
                midx = rx*0.35 + cx2*0.65
                midy = ry*0.35 + cy2*0.65
                dx, dy = cx2 - rx, cy2 - ry
                L = math.hypot(dx, dy) or 1.0
                off = 0.17
                ex = midx - off * dy / L
                ey = midy + off * dx / L
                if len(label) > 32:
                    parts, cur, total = [], [], 0
                    for tok in label.split(', '):
                        if total + len(tok) > 28 and cur:
                            parts.append(', '.join(cur)); cur = [tok]; total = len(tok)
                        else:
                            cur.append(tok); total += (len(tok) + 2)
                    if cur: parts.append(', '.join(cur))
                    label = '\n'.join(parts)
                ax.text(ex, ey, label, ha='center', va='center', fontsize=f_edge, color=PALETTE['edge_text'], zorder=3)

    draw_spokes(inner_pos, inner)
    if outer: draw_spokes(outer_pos, outer)

    fig.tight_layout(pad=0.55)
    fig.savefig(outfile, dpi=dpi)
    plt.close(fig)
    return True


# =============================================================================
# Main
# =============================================================================

def main():
    # ------------------------
    # CLI definition
    # ------------------------
    parser = argparse.ArgumentParser()
    parser.add_argument('--manifest', required=True, type=Path)
    parser.add_argument('--catalog', required=True, type=Path)
    parser.add_argument('--out', required=True, type=Path)
    parser.add_argument('--include-views', action='store_true')
    parser.add_argument('--materializations', type=str, default='')
    parser.add_argument('--schemas', type=str, default='')
    parser.add_argument('--exclude-sheet-prefixes', type=str, default='')
    parser.add_argument('--exclude-sheet-tags', type=str, default='')
    parser.add_argument('--exclude-sheet-materializations', type=str, default='')
    parser.add_argument('--exclude-sheet-path-globs', type=str, default='')
    parser.add_argument('--no-relationships', action='store_true',
                        help='Omit the Relationships sheet.')
    parser.add_argument('--include-missing-dims', action='store_true',
                        help='Include rows in Star Map where the target dimension model is '
                             'missing from the manifest, suffixed with "(missing)". Diagrams still skip.')
    parser.add_argument('--star-diagrams', action='store_true')

    # Diagram flags
    parser.add_argument('--diagram-shape', choices=['roundrect','circle'], default='roundrect')
    parser.add_argument('--diagram-dedupe-roles', action='store_true')
    parser.add_argument('--diagram-two-rings', type=int, default=None)
    parser.add_argument('--diagram-font-scale', type=float, default=1.0)
    parser.add_argument('--diagram-dpi', type=int, default=240)
    parser.add_argument('--diagram-color-scheme', choices=['soft'], default='soft')
    parser.add_argument('--diagram-legend', action='store_true')
    parser.add_argument('--diagram-no-wrap-labels', action='store_true')
    parser.add_argument('--diagram-max-label-chars', type=int, default=18)

    args = parser.parse_args()

    # CHANGE: warn if --schemas was supplied but PyYAML isn't available
    if args.schemas and yaml is None:
        print("WARNING: --schemas was supplied but PyYAML is not installed; YAML overlays will be skipped.",
              file=sys.stderr)

    # ------------------------
    # Load sources
    # ------------------------
    manifest = load_json(args.manifest); catalog = load_json(args.catalog)

    # Model filter (materializations + include/exclude views)
    nodes = manifest.get('nodes', {}); models = [n for n in nodes.values() if node_is_model(n)]
    mats_arg = [m.strip().lower() for m in args.materializations.split(',') if m.strip()]
    filtered = []
    for n in models:
        mat = ((n.get('config') or {}).get('materialized') or '').lower()
        if mats_arg:
            if mat in mats_arg or (args.include_views and mat == 'view'):
                filtered.append(n)
        else:
            if mat != 'ephemeral':
                filtered.append(n)

    # Canonical name maps
    alias_to_node, name_to_alias, alias_to_display = {}, {}, {}
    for n in filtered:
        alias_display = (n.get('alias') or n.get('name') or '')
        alias_lower = alias_display.lower()
        name_lower  = (n.get('name') or '').lower()
        alias_to_node[alias_lower] = n
        alias_to_display[alias_lower] = alias_display
        name_to_alias[alias_lower] = alias_lower
        if name_lower:
            name_to_alias[name_lower] = alias_lower

    def canon(token: str):
        """Normalize tokens to alias_lower when possible."""
        if not token: return None
        t = str(token).strip().strip('"').lower()
        return name_to_alias.get(t, t)

    # ------------------------
    # Relationships from tests
    # ------------------------
    raw_rel_rows = build_relationship_rows_with_deps(manifest)

    # Choose FromModel based on depends_on list; normalize both ends
    norm_rel_rows = []
    for r in raw_rel_rows:
        dep_aliases = [canon(x) for x in (r.get('DepModels') or [])]
        to_alias    = canon(r.get('ToModel'))
        from_alias = None
        for d in dep_aliases:
            if d and d != to_alias:
                from_alias = d; break
        if not from_alias and dep_aliases:
            from_alias = dep_aliases[0]
        if not to_alias or not from_alias:
            continue
        norm_rel_rows.append({
            'FromModel': from_alias,
            'FromColumn': r.get('FromColumn'),
            'ToModel': to_alias,
            'ToColumn': r.get('ToColumn'),
            'TestName': r.get('TestName'),
        })

    # CHANGE: Build a map of missing dimensions per fact (for notes on fact sheets)
    missing_dims_by_fact = defaultdict(set)
    if args.include_missing_dims:
        for r in norm_rel_rows:
            frm = r['FromModel']
            to = r['ToModel']
            if frm and to and (to not in alias_to_display):
                missing_dims_by_fact[frm].add(to)

    # Spokes per fact for diagrams (skip missing dims unless include-missing-dims requested)
    spokes_raw = defaultdict(list)
    for r in norm_rel_rows:
        frm = r['FromModel']; to = r['ToModel']; col = r.get('FromColumn')
        if not frm or not to:
            continue
        if (to not in alias_to_display) and (not args.include_missing_dims):
            continue
        if to in alias_to_display:
            spokes_raw[frm].append((to, col))

    # YAML overlays for per-model sheets
    yaml_overlays = load_yaml_overlays(args.schemas)

    # Per-model sheet exclusions
    ex_prefixes = parse_csv_set(args.exclude_sheet_prefixes)
    ex_tags = parse_csv_set(args.exclude_sheet_tags)
    ex_mats = parse_csv_set(args.exclude_sheet_materializations)
    ex_path_globs = parse_csv_set(args.exclude_sheet_path_globs)

    # ------------------------
    # Excel styling helpers
    # ------------------------
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XLImage
    # CHANGE: robust column helpers (fix column_letter/index issues)
    from openpyxl.utils import get_column_letter, column_index_from_string

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # blue header bg
    header_font = Font(bold=True)
    alt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")    # green-bar

    # ---- Sheet tab colors (hex RGB without '#') ----
    TAB_BLUE   = "5B9BD5"  # Summary / Relationships / Star Map / Others
    TAB_PURPLE = "C9C2F3"  # Dimension model tabs
    TAB_PEACH  = "F8CBAD"  # Fact model tabs
    TAB_GREEN  = "A9D18E"  # Star diagram tabs

    def style_table(ws, top_header_row: int, width_map: dict, wrap_cols: list, freeze=True):
        """
        Apply common styling to a sheet that already has data:
          - optional freeze panes under header
          - blue header bg + bold
          - auto-filter over full used range
          - word-wrap for requested columns
          - set column widths
        wrap_cols: list of column letters by header (e.g., ['A','F','G'])
        """
        if freeze:
            ws.freeze_panes = f"A{top_header_row+1}"

        max_col = ws.max_column

        # header style
        for col_idx in range(1, max_col+1):
            c = ws.cell(row=top_header_row, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="center")  # CHANGE: wrap long headers

        # auto-filter
        last_col_letter = get_column_letter(max_col)  # CHANGE
        ws.auto_filter.ref = f"A{top_header_row}:{last_col_letter}{ws.max_row}"

        # word-wrap selected columns by letter
        for r in range(top_header_row+1, ws.max_row+1):
            for letter in wrap_cols:
                col_index = column_index_from_string(letter)  # CHANGE
                ws.cell(row=r, column=col_index).alignment = Alignment(wrap_text=True, vertical="top")

        # set widths
        for letter, width in width_map.items():
            ws.column_dimensions[letter].width = width

    # Column width presets for overview tabs
    widths_summary = {"A": 26, "B": 12, "C": 16, "D": 16, "E": 16, "F": 36, "G": 28, "H": 36, "I": 24, "J": 50}
    widths_rel     = {"A": 28, "B": 30, "C": 28, "D": 30, "E": 26}
    widths_star    = {"A": 28, "B": 28, "C": 36, "D": 30}

    # =============================================================================
    # Build DataFrames for the overview tabs
    # =============================================================================
    summary_rows = []
    for alias_lower, n in alias_to_node.items():
        db, schema, alias_disp = model_relation_identifiers(n)
        mat = (n.get('config') or {}).get('materialized')
        fqn = '.'.join(n.get('fqn') or [])
        tags = ','.join(n.get('tags') or [])
        desc = n.get('description') or ''
        kind = classify_model_kind(alias_lower, n.get('tags', []))
        summary_rows.append({'Model': alias_disp,'Kind': kind,'Materialization': mat,'Database': db,'Schema': schema,
                             'Relation': f'{db}.{schema}.{alias_disp}','Path': n.get('path'),'FQN': fqn,'Tags': tags,'Description': desc})
    df_summary = pd.DataFrame(summary_rows, columns=['Model','Kind','Materialization','Database','Schema','Relation','Path','FQN','Tags','Description'])
    if not df_summary.empty:
        df_summary = df_summary.sort_values(['Kind','Model'])

    # Relationships tab (may be omitted later via --no-relationships)
    df_rel = pd.DataFrame([
        {'FromModel': alias_to_display.get(r['FromModel'], r['FromModel']),
         'FromColumn': r.get('FromColumn'),
         'ToModel': alias_to_display.get(r['ToModel'], r['ToModel']) if r['ToModel'] in alias_to_display else (r['ToModel'] or '(missing)'),
         'ToColumn': r.get('ToColumn'),
         'TestName': r.get('TestName')}
        for r in norm_rel_rows
    ]) if norm_rel_rows else pd.DataFrame(columns=['FromModel','FromColumn','ToModel','ToColumn','TestName'])

    # Star Map (aggregated per Fact/Dim with FK/PK columns composed)
    star_map = OrderedDict()
    for r in norm_rel_rows:
        f = alias_to_display.get(r['FromModel'], r['FromModel'])
        if r['ToModel'] in alias_to_display:
            d = alias_to_display.get(r['ToModel'], r['ToModel'])
        else:
            if not args.include_missing_dims:
                continue
            d_raw = r['ToModel'] or '(missing)'
            d = f"{d_raw}(missing)" if not d_raw.endswith('(missing)') else d_raw

        e = star_map.setdefault((f,d), {'FactModel': f, 'DimensionModel': d, 'FactFKColumn': set(), 'DimKeyColumn': set()})
        if r.get('FromColumn'):
            e['FactFKColumn'].add(r['FromColumn'])
        if r.get('ToColumn'):
            e['DimKeyColumn'].add(r['ToColumn'])

    df_star = (
        pd.DataFrame([
            {'FactModel': e['FactModel'],
             'DimensionModel': e['DimensionModel'],
             'FactFKColumn': ', '.join(sorted(e['FactFKColumn'])) if e['FactFKColumn'] else '',
             'DimKeyColumn': ', '.join(sorted(e['DimKeyColumn'])) if e['DimKeyColumn'] else ''}
            for e in star_map.values()
        ], columns=['FactModel','DimensionModel','FactFKColumn','DimKeyColumn'])
        if star_map else
        pd.DataFrame(columns=['FactModel','DimensionModel','FactFKColumn','DimKeyColumn'])
    )

    used_sheet_names = set()  # will be re-seeded after overview tabs (see below)

    # =============================================================================
    # Write Excel workbook
    # =============================================================================
    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        # ----- Summary -----
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        wb = writer.book; ws = wb['Summary']
        ws.sheet_properties.tabColor = TAB_BLUE
        style_table(ws, top_header_row=1, width_map=widths_summary,
                    wrap_cols=['A','F','G','H','I','J'], freeze=True)

        # ----- Relationships -----
        if not args.no_relationships and (df_rel is not None) and (not df_rel.empty):
            df_rel.to_excel(writer, index=False, sheet_name='Relationships')
            ws = wb['Relationships']
            ws.sheet_properties.tabColor = TAB_BLUE
            style_table(ws, top_header_row=1, width_map=widths_rel,
                        wrap_cols=['A','B','C','D','E'], freeze=True)

        # ----- Star Map -----
        if not df_star.empty:
            df_star.to_excel(writer, index=False, sheet_name='Star Map')
            ws = wb['Star Map']
            ws.sheet_properties.tabColor = TAB_BLUE
            style_table(ws, top_header_row=1, width_map=widths_star,
                        wrap_cols=['A','B','C','D'], freeze=True)

        # CHANGE: seed the used sheet-name set with the workbook's current sheetnames
        used_sheet_names = set(writer.book.sheetnames)

        # ----- Per-model tabs -----
        col_widths = {"A": 28, "B": 20, "C": 10, "D": 80, "E": 28, "F": 6, "G": 6, "H": 12}

        for alias_lower, n in alias_to_node.items():
            db, schema, alias_disp = model_relation_identifiers(n)
            tags_lower = {t.lower() for t in (n.get('tags') or [])}
            mat_lower = ((n.get('config') or {}).get('materialized') or '').lower()
            path_str = n.get('path') or ''

            # small predicate to apply all four exclusion mechanisms
            def should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                     prefixes, tags_excl, mats_excl, path_globs) -> bool:
                if any(alias_lower.startswith(p) for p in prefixes): return True
                if any(t in tags_lower for t in tags_excl): return True
                if mat_lower in mats_excl: return True
                ps = (path_str or '')
                for g in path_globs:
                    if fnmatch.fnmatch(ps.lower(), g): return True
                return False

            if should_exclude_sheet(alias_lower, tags_lower, mat_lower, path_str,
                                    ex_prefixes, ex_tags, ex_mats, ex_path_globs):
                continue

            # Reserve a unique sheet name ONCE (fixes unwanted “_1” suffix)
            model_sheet_name = safe_sheet_name(alias_disp or 'Model', used_sheet_names)

            # Tests + heuristic flags
            col_tests = extract_tests_for_model(manifest, n.get('unique_id'))
            flags = infer_pk_fk_and_nullable(col_tests)

            # Overlays
            overlay_entry = (yaml_overlays.get(alias_lower, {}) or {})
            overlay_cols = overlay_entry.get('columns', {}) or {}
            meta_disp = overlay_entry.get('model_meta_display', {}) or {}
            meta_match = overlay_entry.get('model_meta_match', {}) or {}

            # Display lists preserve original case; match sets are lowercase
            sk_display = meta_disp.get('surrogate_key', []) or []
            bk_display = meta_disp.get('business_key', []) or []
            pk_display = meta_disp.get('primary_key', []) or []
            pk_display_final = pk_display if pk_display else sk_display  # display fallback

            pk_match = set((meta_match.get('primary_key', []) or []))
            sk_match = set((meta_match.get('surrogate_key', []) or []))
            explicit_pk_names = {*pk_match, *sk_match}  # assists column PK inference

            # Build columns grid (prefer catalog metadata; fallback to manifest metadata)
            rows = []
            catalog_cols = get_catalog_columns_for_model(catalog, n)
            if catalog_cols:
                ordered = sorted(catalog_cols.items(), key=lambda kv: (kv[1].get('index') or 0))
                for col, meta in ordered:
                    key = (col or '').lower()
                    desc_catalog = meta.get('comment') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or (meta.get('type') or '')
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_catalog
                    source = 'Merged' if ov else 'Catalog'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})
            else:
                manifest_cols = (n.get('columns') or {})
                for col, meta in manifest_cols.items():
                    key = (col or '').lower()
                    desc_manifest = meta.get('description') or ''
                    tests = col_tests.get(col) or []
                    test_list = ', '.join(sorted({t['name'] for t in tests})) if tests else ''
                    is_pk = flags.get(col, {}).get('is_pk', False) or (key in explicit_pk_names)
                    is_fk = flags.get(col, {}).get('is_fk', False)
                    nullable_by_test = flags.get(col, {}).get('nullable_from_tests', '')
                    ov = overlay_cols.get(key, {})
                    dtype = ov.get('data_type') or ''
                    nullable = ov.get('nullable');  nullable = nullable if nullable is not None else nullable_by_test
                    desc = ov.get('description') or desc_manifest
                    source = 'YAML' if ov else 'Manifest'
                    rows.append({'Column': col,'DataType': dtype,'Nullable?': ('Y' if nullable in (True,'Y','y','yes','true',1) else ('' if nullable=='' else 'N')),
                                 'Description': desc,'Tests': test_list,'IsPK?': 'Y' if is_pk else '','IsFK?': 'Y' if is_fk else '','Source': source})

            df = pd.DataFrame(rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            # Table meta block (labels without '#'; original-case lists; PK fallback)
            db, schema, alias_disp = model_relation_identifiers(n)
            info_rows = [
                {'Column': 'Model',         'DataType': alias_disp,                                'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Kind',          'DataType': classify_model_kind(alias_lower, n.get('tags', [])), 'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Materialization','DataType': (n.get('config') or {}).get('materialized'),        'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Relation',      'DataType': f'{db}.{schema}.{alias_disp}',             'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Tags',          'DataType': ','.join(n.get('tags') or []),             'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'SurrogateKey',  'DataType': ', '.join(sk_display),                     'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'BusinessKey',   'DataType': ', '.join(bk_display),                     'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'PrimaryKey',    'DataType': ', '.join(pk_display_final),               'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
                {'Column': 'Description',   'DataType': (n.get('description') or ''),              'Nullable?':'','Description':'','Tests':'','IsPK?':'','IsFK?':'','Source':''},
            ]
            df_meta = pd.DataFrame(info_rows, columns=['Column','DataType','Nullable?','Description','Tests','IsPK?','IsFK?','Source'])

            # Write per-model tab
            df_meta.to_excel(writer, index=False, sheet_name=model_sheet_name, header=False, startrow=0)
            start_row = len(df_meta.index) + 1
            df.to_excel(writer, index=False, sheet_name=model_sheet_name, startrow=start_row)

            # Per-model formatting (freeze, header style, wrap, banding, widths)
            ws = wb[model_sheet_name]
            # Color tab by kind
            kind = classify_model_kind(alias_lower, n.get('tags', []))
            ws.sheet_properties.tabColor = TAB_PEACH if kind == 'Fact' else (TAB_PURPLE if kind == 'Dimension' else TAB_BLUE)

            ws.freeze_panes = f"A{start_row+2}"                         # keep the columns header visible
            hdr = start_row + 1                                         # first row of the grid

            # style header row (blue background + bold)
            for col_idx in range(1, 8+1):
                c = ws.cell(row=hdr, column=col_idx)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(vertical="center")

            # auto-filter over the entire data range
            from openpyxl.utils import get_column_letter as _gcl  # local alias okay
            last_col_letter = _gcl(8)  # CHANGE: robust column letter
            ws.auto_filter.ref = f"A{hdr}:{last_col_letter}{ws.max_row}"

            # wrap Description (D) and Tests (E)
            for r in range(hdr+1, ws.max_row+1):
                ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")

            # bold meta labels in the top block (Column column only)
            from openpyxl.styles import Font as _Font
            for r in range(1, len(df_meta.index)+1):
                ws.cell(row=r, column=1).font = _Font(bold=True)

            # green-bar banding on the grid rows
            for r in range(hdr+1, ws.max_row+1):
                if (r - (hdr+1)) % 2 == 1:
                    for c in range(1, 8+1):
                        ws.cell(row=r, column=c).fill = alt_fill

            # column widths
            for col_letter, width in {"A":28,"B":20,"C":10,"D":80,"E":28,"F":6,"G":6,"H":12}.items():
                ws.column_dimensions[col_letter].width = width

            # CHANGE: If requested, append a note listing missing dimensions for FACT sheets
            if args.include_missing_dims and (kind == 'Fact'):
                missing = sorted(missing_dims_by_fact.get(alias_lower, []))
                if missing:
                    note_row = ws.max_row + 2
                    ws.cell(row=note_row, column=1, value="Notes")
                    ws.cell(row=note_row, column=1).font = header_font  # bold label
                    ws.cell(row=note_row + 1, column=1, value="Some dimensions referenced by dbt relationships "
                                                              "weren't found in the manifest and were omitted from diagrams.")
                    ws.cell(row=note_row + 2, column=1, value="Missing dimensions:")
                    ws.cell(row=note_row + 3, column=1, value=", ".join(missing))
                    from openpyxl.styles import Alignment as _Alignment
                    for r_ in range(note_row + 1, note_row + 4):
                        ws.cell(row=r_, column=1).alignment = _Alignment(wrap_text=True, vertical="top")
                    if ws.column_dimensions["A"].width < 40:
                        ws.column_dimensions["A"].width = 40

        # ----- Diagram tabs -----
        if args.star_diagrams:
            from tempfile import TemporaryDirectory  # CHANGE: auto-clean temp dir
            wb = writer.book
            used = set(wb.sheetnames)

            # CHANGE: use TemporaryDirectory so it cleans up even on exceptions
            with TemporaryDirectory(prefix="dbt_star_") as tmpdir:
                for center_alias_lower, pairs in sorted(spokes_raw.items()):
                    # Aggregate FK labels per dimension (dedupe roles if requested)
                    if args.diagram_dedupe_roles:
                        by_dim = OrderedDict()
                        for dim_alias_lower, fk_col in pairs:
                            by_dim.setdefault(dim_alias_lower, []).append(fk_col or '')
                        spokes = []
                        for dim_alias_lower, fk_cols in by_dim.items():
                            spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                           'dim_family': 'uniform',
                                           'edge_labels': [c for c in fk_cols if c]})
                    else:
                        spokes = []
                        for dim_alias_lower, fk_col in pairs:
                            spokes.append({'dim_label': alias_to_display.get(dim_alias_lower, dim_alias_lower),
                                           'dim_family': 'uniform',
                                           'edge_labels': [fk_col] if fk_col else []})

                    if not spokes:
                        continue  # nothing to draw

                    fact_label = alias_to_display.get(center_alias_lower, center_alias_lower)
                    img_path = os.path.join(tmpdir, f"{center_alias_lower}.png")

                    drew = draw_star_png(
                        fact_label, 'other', spokes, img_path,
                        shape='roundrect',
                        two_rings_threshold=args.diagram_two_rings,
                        font_scale=args.diagram_font_scale,
                        dpi=args.diagram_dpi,
                        color_scheme=args.diagram_color_scheme,
                        wrap_labels=not args.diagram_no_wrap_labels,
                        max_label_chars=max(10, args.diagram_max_label_chars),
                    )
                    if not drew:
                        continue

                    # Prefer "Star-Fact <Fact>"; if unsafe characters, fall back to "Star Fact <Fact>"
                    proposed = f"Star-Fact {fact_label}"
                    if any(c in proposed for c in r'[]:*?\/'):
                        proposed = f"Star Fact {fact_label}"
                    sheet_name = safe_sheet_name(proposed, used)

                    # Create diagram sheet and drop the PNG at A1
                    ws = wb.create_sheet(title=sheet_name)
                    ws.sheet_properties.tabColor = TAB_GREEN
                    try:
                        img = XLImage(img_path)
                        ws.add_image(img, "A1")
                    except Exception:
                        pass

                    if args.diagram_legend:
                        ws["J2"] = ("Legend: dimensions shown in light green; fact in light yellow.\n"
                                    "Spoke text = fact FK column(s).")
                        from openpyxl.styles import Alignment as _Alignment
                        ws["J2"].alignment = _Alignment(wrap_text=True)

    print(f"Wrote: {args.out}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr); sys.exit(1)
